import argparse
import json
import unittest
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

matplotlib.use("Agg")
import matplotlib.pyplot as plt


PALETTE = {
    "blue": "#5271AE",
    "light_blue": "#70ACDE",
    "yellow": "#F5CC7D",
    "orange": "#FFA660",
    "red": "#D85B59",
}
PRODUCTION_LEVELS = [72.0, 63.0, 54.0, 45.0, 36.0]
REPRESENTATIVE_PRODUCTION = 36.0
PARK_COUNTS = [1, 5, 10, 20]


def calculate_exchange_metrics(exchange_power):
    values = np.asarray(exchange_power, dtype=float).reshape(-1)
    if values.size == 0 or not np.all(np.isfinite(values)):
        raise ValueError("Exchange profile must contain finite values")
    ramps = np.abs(np.diff(values))
    return {
        "购电量(MWh)": float(np.clip(values, 0.0, None).sum()),
        "售电量(MWh)": float(np.clip(-values, 0.0, None).sum()),
        "最大购电功率(MW)": float(max(values.max(), 0.0)),
        "最大反送功率(MW)": float(max(-values.min(), 0.0)),
        "交换功率峰谷差(MW)": float(values.max() - values.min()),
        "最大爬坡功率(MW/h)": float(ramps.max()) if ramps.size else 0.0,
        "交换功率标准差(MW)": float(values.std(ddof=0)),
    }


def calculate_aggregate_scale_metrics(profiles, park_counts):
    values = np.asarray(profiles, dtype=float)
    if values.ndim != 2 or values.shape[0] == 0 or values.shape[1] == 0:
        raise ValueError("Profiles must be a non-empty scenario-by-hour matrix")
    scenario_metrics = pd.DataFrame(
        [calculate_exchange_metrics(profile) for profile in values]
    )
    diversified_base = calculate_exchange_metrics(values.mean(axis=0))
    synchronized_base = scenario_metrics.max(axis=0).to_dict()
    rows = []
    for count in park_counts:
        if count <= 0:
            raise ValueError("Equivalent park count must be positive")
        for mode, base in (
            ("场景多样化期望", diversified_base),
            ("同步运行上界", synchronized_base),
        ):
            row = {"等效园区数量": int(count), "聚合方式": mode}
            row.update({key: float(value) * count for key, value in base.items()})
            rows.append(row)
    return pd.DataFrame(rows)


def configure_plotting():
    plt.rcParams.update(
        {
            "font.sans-serif": ["Microsoft YaHei", "SimHei", "DejaVu Sans"],
            "axes.unicode_minus": False,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "figure.dpi": 140,
            "savefig.dpi": 300,
        }
    )


def require_columns(frame, columns, source_name):
    missing = [column for column in columns if column not in frame.columns]
    if missing:
        raise ValueError(f"{source_name} missing columns: {missing}")


def load_hourly_results(path, model_name):
    frame = pd.read_csv(path)
    required = [
        "场景",
        "风电场景",
        "光伏场景",
        "日氨产量(t/day)",
        "时段序号(h)",
        "总负荷功率(MW)",
        "风电功率(MW)",
        "光伏功率(MW)",
        "购电功率(MW)",
        "售电功率(MW)",
    ]
    require_columns(frame, required, model_name)
    numeric_columns = [column for column in required if column not in {"场景"}]
    for column in numeric_columns:
        frame[column] = pd.to_numeric(frame[column], errors="raise")
    frame["模型"] = model_name
    frame["净交换功率(MW)"] = frame["购电功率(MW)"] - frame["售电功率(MW)"]
    frame["功率平衡残差(MW)"] = (
        frame["总负荷功率(MW)"]
        + frame["售电功率(MW)"]
        - frame["风电功率(MW)"]
        - frame["光伏功率(MW)"]
        - frame["购电功率(MW)"]
    )
    return frame


def build_scenario_metrics(hourly):
    group_columns = ["模型", "场景", "风电场景", "光伏场景", "日氨产量(t/day)"]
    rows = []
    for keys, group in hourly.groupby(group_columns, sort=True):
        group = group.sort_values("时段序号(h)")
        if len(group) != 24 or group["时段序号(h)"].nunique() != 24:
            raise ValueError(f"Hourly group must contain 24 unique periods: {keys}")
        row = dict(zip(group_columns, keys))
        row.update(calculate_exchange_metrics(group["净交换功率(MW)"].to_numpy()))
        rows.append(row)
    return pd.DataFrame(rows)


def compare_models(problem_two_metrics, problem_three_metrics):
    keys = ["场景", "风电场景", "光伏场景", "日氨产量(t/day)"]
    metric_columns = [
        "购电量(MWh)",
        "售电量(MWh)",
        "最大购电功率(MW)",
        "最大反送功率(MW)",
        "交换功率峰谷差(MW)",
        "最大爬坡功率(MW/h)",
        "交换功率标准差(MW)",
    ]
    left = problem_two_metrics[keys + metric_columns].copy()
    right = problem_three_metrics[keys + metric_columns].copy()
    merged = left.merge(right, on=keys, suffixes=("_问题二", "_问题三"), validate="one_to_one")
    for column in metric_columns:
        merged[f"{column}_变化"] = merged[f"{column}_问题三"] - merged[f"{column}_问题二"]
    return merged


def build_production_robustness(comparison):
    change_columns = {
        "购电量(MWh)_变化": ("购电平均变化(MWh)", "购电减少场景数"),
        "售电量(MWh)_变化": ("售电平均变化(MWh)", "售电减少场景数"),
        "交换功率峰谷差(MW)_变化": ("峰谷差平均变化(MW)", "峰谷差减少场景数"),
        "最大爬坡功率(MW/h)_变化": ("最大爬坡平均变化(MW/h)", "爬坡减少场景数"),
    }
    require_columns(comparison, ["日氨产量(t/day)", *change_columns], "Model comparison")
    rows = []
    for production, group in comparison.groupby("日氨产量(t/day)", sort=False):
        row = {"日氨产量(t/day)": float(production), "场景数": int(len(group))}
        for source, (mean_name, count_name) in change_columns.items():
            values = pd.to_numeric(group[source], errors="raise")
            row[mean_name] = float(values.mean())
            row[count_name] = int((values < -1e-9).sum())
        rows.append(row)
    return pd.DataFrame(rows).sort_values("日氨产量(t/day)", ascending=False).reset_index(drop=True)


def build_extreme_scenarios(problem_three_metrics):
    selected = problem_three_metrics[
        np.isclose(problem_three_metrics["日氨产量(t/day)"], REPRESENTATIVE_PRODUCTION)
    ]
    if selected.empty:
        raise ValueError("Representative production has no scenario metrics")
    definitions = [
        ("最大购电功率", "最大购电功率(MW)", "MW"),
        ("最大反送功率", "最大反送功率(MW)", "MW"),
        ("最大交换峰谷差", "交换功率峰谷差(MW)", "MW"),
        ("最大爬坡功率", "最大爬坡功率(MW/h)", "MW/h"),
    ]
    rows = []
    for label, column, unit in definitions:
        index = selected[column].idxmax()
        source = selected.loc[index]
        rows.append(
            {
                "代表日氨产量(t/day)": REPRESENTATIVE_PRODUCTION,
                "极值指标": label,
                "极值": float(source[column]),
                "单位": unit,
                "对应场景": source["场景"],
                "风电场景": int(source["风电场景"]),
                "光伏场景": int(source["光伏场景"]),
                "口径说明": "各指标分别取极值，极值不要求来自同一场景",
            }
        )
    return pd.DataFrame(rows)


def build_literature_sources():
    return pd.DataFrame(
        [
            [
                "L1",
                "政策文件",
                "国家发展改革委、国家能源局",
                "关于有序推动绿电直连发展有关事项的通知（发改能源〔2025〕650号）",
                "要求源荷匹配、交换功率不超过申报容量，鼓励储能和负荷柔性，并落实可观可测可调可控、电能质量和费用责任。",
                "https://www.nea.gov.cn/20250530/2d67a6e49c044f2eacabe1fddf48d20f/c.html",
            ],
            [
                "L2",
                "国内研究",
                "孔令国等",
                "基于功率偏差分摊的源氢氨多元耦合系统参与电网调峰优化调度",
                "源氢氨系统参与调峰后可降低深度调峰量与净负荷方差，支持按实际调峰绩效给予辅助服务补偿。",
                "https://dgjsxb.ces-transaction.com/CN/10.19595/j.cnki.1000-6753.tces.241938",
            ],
            [
                "L3",
                "国内研究",
                "潘禹等",
                "含混合电解槽的绿电-绿氢系统能量管理策略研究",
                "混合电解槽及其能量管理可利用不同设备效率与动态特性提升可再生能源适应能力。",
                "https://www.tynxb.org.cn/CN/10.19912/j.0254-0096.tynxb.2023-1278",
            ],
            [
                "L4",
                "国外研究",
                "Armijo and Philibert",
                "Flexible production of green hydrogen and ammonia from variable solar and wind energy",
                "风光互补和生产柔性有助于降低可再生能源时序波动及其成本。",
                "https://www.sciencedirect.com/science/article/pii/S0360319919342089",
            ],
            [
                "L5",
                "国外研究",
                "Salmon and Bañares-Alcántara",
                "Impact of process flexibility and imperfect forecasting on the operation and design of Haber-Bosch green ammonia",
                "基于LP与MPC说明合成氨过程柔性和预测信息共同影响离网绿氨系统的稳健运行与设计。",
                "https://pubs.rsc.org/doi/d3su00067b",
            ],
            [
                "L6",
                "国外研究",
                "Göke et al.",
                "How flexible electrification can integrate fluctuating renewables",
                "灵活电气化尤其是柔性电解槽可降低剩余峰值负荷、剩余需求和过剩发电。",
                "https://www.sciencedirect.com/science/article/pii/S0360544223012264",
            ],
        ],
        columns=["文献编号", "类型", "作者或机构", "题名", "与本题相关结论", "链接"],
    )


def build_system_impact_summary():
    return pd.DataFrame(
        [
            ["正面影响", "B1", "促进新能源就地消纳", "园区内部消纳风光电量，降低部分余电外送需求", "E1、E2、E4", "L1、L4"],
            ["正面影响", "B2", "提供柔性负荷与调峰能力", "连续调节将制氢制氨负荷迁移至风光富余或低价时段", "E3、E4", "L2、L3、L5、L6"],
            ["正面影响", "B3", "缓解部分远距离输电需求", "绿电在园区就近转化为氢氨产品，可减少部分新能源跨区输送", "E1、E2", "L1、L4"],
            ["负面影响", "D1", "同步购电峰和爬坡压力", "多园区运行相关性高时，购电峰与爬坡功率按接入规模放大", "E3、E9", "L1、L2、L6"],
            ["负面影响", "D2", "集中反送与反向潮流压力", "富余风光集中上网可能形成较大的反送功率边界", "E1、E9", "L1"],
            ["负面影响", "D3", "备用与调节需求增加", "风光预测误差和负荷调节误差需要公共电网提供备用支撑", "E3、E8、E9", "L1、L2、L5、L6"],
            ["负面影响", "D4", "电能质量与保护协调风险", "大规模电力电子设备和双向潮流提高无功、电能质量及保护配合要求", "E9（仅边界功率）", "L1"],
            ["负面影响", "D5", "系统成本与责任分担问题", "园区依赖电网备用却可能未完全承担容量和系统运行成本", "E8、E9", "L1"],
        ],
        columns=["影响类型", "影响编号", "影响", "作用机理", "本题量化依据", "文献政策依据"],
    )


def build_aggregate_results(problem_three_hourly):
    selected = problem_three_hourly[
        np.isclose(problem_three_hourly["日氨产量(t/day)"], REPRESENTATIVE_PRODUCTION)
    ].copy()
    pivot = selected.pivot(index="场景", columns="时段序号(h)", values="净交换功率(MW)")
    pivot = pivot.sort_index(axis=0).sort_index(axis=1)
    if pivot.shape != (24, 24) or pivot.isna().any().any():
        raise ValueError(f"Expected a complete 24x24 scenario profile matrix, got {pivot.shape}")
    result = calculate_aggregate_scale_metrics(pivot.to_numpy(), PARK_COUNTS)
    result.insert(0, "代表日氨产量(t/day)", REPRESENTATIVE_PRODUCTION)
    metric_pairs = ["最大购电功率(MW)", "最大反送功率(MW)"]
    for metric in metric_pairs:
        synchronized = result[result["聚合方式"] == "同步运行上界"].set_index("等效园区数量")[metric]
        diversified_mask = result["聚合方式"] == "场景多样化期望"
        result.loc[diversified_mask, f"{metric}较同步削减率(%)"] = result.loc[
            diversified_mask, "等效园区数量"
        ].map(lambda count: 100.0 * (1.0 - result.loc[
            diversified_mask & (result["等效园区数量"] == count), metric
        ].iloc[0] / synchronized.loc[count]))
        result.loc[~diversified_mask, f"{metric}较同步削减率(%)"] = 0.0
    return result, pivot


def build_mean_exchange_profiles(problem_two_hourly, problem_three_hourly):
    rows = []
    for model_name, frame in (("问题二离散调节", problem_two_hourly), ("问题三连续调节", problem_three_hourly)):
        selected = frame[np.isclose(frame["日氨产量(t/day)"], REPRESENTATIVE_PRODUCTION)]
        profile = selected.groupby("时段序号(h)", as_index=False)["净交换功率(MW)"].mean()
        profile.insert(0, "模型", model_name)
        rows.append(profile)
    return pd.concat(rows, ignore_index=True)


def read_indicator_map(path, value_column="数值"):
    frame = pd.read_csv(path)
    require_columns(frame, ["指标", value_column], path.name)
    return frame.set_index("指标")[value_column].astype(float).to_dict()


def build_evidence_tables(project_root, aggregate_results, extreme_scenarios):
    q1_dir = project_root / "outputs" / "问题一计算结果"
    q3_dir = project_root / "outputs" / "问题三计算结果"
    q4_dir = project_root / "outputs" / "问题四计算结果"
    q1 = read_indicator_map(q1_dir / "问题一汇总结果.csv")
    green = read_indicator_map(q1_dir / "问题一绿电指标.csv", "计算值")
    flexibility = pd.read_csv(q3_dir / "问题三柔性价值.csv")
    qualification = pd.read_csv(q3_dir / "问题二三全年指标变化.csv")
    q4_annual = pd.read_csv(q4_dir / "问题四年度汇总.csv")
    storage_compare = pd.read_csv(q4_dir / "问题四储能前后比较.csv")
    autonomy = pd.read_csv(q4_dir / "问题四能源自治容量边界.csv")

    flex_36 = flexibility[np.isclose(flexibility["日氨产量(t/day)"], 36.0)].iloc[0]
    q4_before = q4_annual[q4_annual["运行模式"] == "离网无储能"].iloc[0]
    q4_after = q4_annual[q4_annual["运行模式"] == "离网配置储能"].iloc[0]
    q4_grid = q4_annual[q4_annual["运行模式"] == "并网同产量"].iloc[0]
    storage_capacity = float(storage_compare["储能容量(MWh)_储能后"].max())
    support_value = (
        q4_after["年综合净成本(¥/year)"] - q4_grid["年综合净成本(¥/year)"]
    ) / q4_after["年氨产量(t/year)"]
    unqualified_two = int(qualification["问题二全部不达标天数(day)"].sum())
    unqualified_three = int(qualification["问题三全部不达标天数(day)"].sum())
    n20_sync = aggregate_results[
        (aggregate_results["等效园区数量"] == 20) & (aggregate_results["聚合方式"] == "同步运行上界")
    ].iloc[0]
    n20_div = aggregate_results[
        (aggregate_results["等效园区数量"] == 20) & (aggregate_results["聚合方式"] == "场景多样化期望")
    ].iloc[0]
    extreme_index = extreme_scenarios.set_index("极值指标")
    import_scene = extreme_index.loc["最大购电功率", "对应场景"]
    export_scene = extreme_index.loc["最大反送功率", "对应场景"]

    evidence_rows = [
        {
            "证据编号": "E1",
            "量化证据": f"典型日R1={100 * green['新能源自发自用比例']:.2f}%，R3={100 * green['上网电量比例']:.2f}%，购电14 h、上网10 h",
            "系统含义": "总量充裕仍可能同时形成购电峰和反送峰",
            "来源": "问题一汇总结果、问题一绿电指标",
        },
        {
            "证据编号": "E2",
            "量化证据": f"新能源总量充裕系数={q1['新能源总量充裕系数']:.4f}，购电量={q1['购电量']:.3f} MWh，上网量={q1['上网电量']:.3f} MWh",
            "系统含义": "园区核心矛盾是时序错配而非单纯电量不足",
            "来源": "问题一汇总结果",
        },
        {
            "证据编号": "E3",
            "量化证据": f"36 t/day时连续调节使日均购电变化{flex_36['平均购电量变化(MWh)']:.3f} MWh、售电变化{flex_36['平均售电量变化(MWh)']:.3f} MWh，柔性价值{flex_36['单位吨氨柔性价值(¥/t)']:.2f} ¥/t",
            "系统含义": "柔性生产负荷能够平滑园区与电网交换并产生经济价值",
            "来源": "问题三柔性价值",
        },
        {
            "证据编号": "E4",
            "量化证据": f"五档产量合计全部不达标天数由{unqualified_two} day降至{unqualified_three} day",
            "系统含义": "连续调节显著改善绿电指标合规性",
            "来源": "问题二三全年指标变化",
        },
        {
            "证据编号": "E5",
            "量化证据": f"既有装机离网无储能年产氨{q4_before['年氨产量(t/year)']:.1f} t，仅为满产25920 t的{100 * q4_before['年氨产量(t/year)'] / 25920.0:.2f}%",
            "系统含义": "既有装机难以依靠离网方式维持高产能利用率",
            "来源": "问题四年度汇总",
        },
        {
            "证据编号": "E6",
            "量化证据": f"全场景严格72 t/day自治的最小组合为风电{autonomy.iloc[0]['风电装机容量(MW)']:.3f} MW、光伏{autonomy.iloc[0]['光伏装机容量(MW)']:.3f} MW",
            "系统含义": "强制绝对自治会导致显著过度扩容",
            "来源": "问题四能源自治容量边界",
        },
        {
            "证据编号": "E7",
            "量化证据": f"配置{storage_capacity:.3f} MWh储能后，年产氨增加{q4_after['年氨产量(t/year)'] - q4_before['年氨产量(t/year)']:.3f} t、弃电减少{q4_before['全年弃电量(MWh/year)'] - q4_after['全年弃电量(MWh/year)']:.3f} MWh、吨氨成本下降{q4_before['综合净吨氨成本(¥/t)'] - q4_after['综合净吨氨成本(¥/t)']:.3f} ¥/t",
            "系统含义": "储能具有价值，但在现有价格与成本条件下边际收益有限",
            "来源": "问题四年度汇总、问题四储能前后比较",
        },
        {
            "证据编号": "E8",
            "量化证据": f"并网同产量相对离网储能方案节约{q4_after['年综合净成本(¥/year)'] - q4_grid['年综合净成本(¥/year)']:.2f} ¥/year，电网支撑价值{support_value:.2f} ¥/t",
            "系统含义": "电网提供的调峰和备用支撑具有较大隐性价值",
            "来源": "问题四年度汇总",
        },
        {
            "证据编号": "E9",
            "量化证据": f"20个等效园区同步上界最大购电/反送为{n20_sync['最大购电功率(MW)']:.2f}/{n20_sync['最大反送功率(MW)']:.2f} MW，分别来自{import_scene}/{export_scene}；多样化期望为{n20_div['最大购电功率(MW)']:.2f}/{n20_div['最大反送功率(MW)']:.2f} MW",
            "系统含义": "园区集中接入规模与场景相关性共同决定电网边界压力",
            "来源": "问题五多园区聚合计算",
        },
    ]
    evidence = pd.DataFrame(evidence_rows)
    policy = pd.DataFrame(
        [
            ["P1", "完善绿电直连计量、认证与考核机制，统一自发自用和上网比例口径", "E1、E2", "L1", "避免指标强相关造成重复考核，并真实反映时序交换"],
            ["P2", "将最大购电、最大反送、峰谷差和爬坡率纳入接入审查并设置交换功率边界", "E1、E2、E9", "L1", "把接入评价从年电量扩展至功率与时序风险"],
            ["P3", "建立需求响应和辅助服务补偿，按实际削峰、填谷及爬坡改善量结算", "E3、E4", "L2、L6", "使柔性制氢制氨负荷的正外部性获得市场回报"],
            ["P4", "储能支持由固定容量补贴转向弃电削减、失供避免和调峰绩效补贴", "E7", "L1、L4、L5", "避免无效堆砌容量，提高财政资金效率"],
            ["P5", "建立公平的电网容量和系统运行成本分担机制，并允许辅助服务收益抵扣", "E8、E9", "L1、L2", "回收备用支撑成本，同时避免重复收费"],
            ["P6", "实行并网型、弱并网型和自治型分类管理，不把完全离网作为普遍要求", "E5、E6", "L1、L4、L5", "兼顾新能源消纳、园区经济性和供电可靠性"],
            ["P7", "建立功率预测报送、可观可测可调可控和保护协调技术标准", "E1、E2、E9", "L1、L5", "提升大规模接入后的运行可控性"],
        ],
        columns=["政策编号", "政策建议", "本题量化依据", "文献政策依据", "实施逻辑"],
    )
    key_values = {
        "flexibility_value": float(flex_36["单位吨氨柔性价值(¥/t)"]),
        "unqualified_two": unqualified_two,
        "unqualified_three": unqualified_three,
        "storage_capacity": storage_capacity,
        "support_value": float(support_value),
        "n20_sync_import": float(n20_sync["最大购电功率(MW)"]),
        "n20_div_import": float(n20_div["最大购电功率(MW)"]),
        "n20_sync_export": float(n20_sync["最大反送功率(MW)"]),
        "n20_div_export": float(n20_div["最大反送功率(MW)"]),
    }
    return evidence, policy, key_values


def build_validation(
    problem_two_hourly,
    problem_three_hourly,
    metrics_two,
    metrics_three,
    aggregate,
    evidence,
    robustness,
    extreme_scenarios,
    literature,
    impacts,
    policy,
):
    rows = []

    def add(name, value, target, passed):
        rows.append({"检验项目": name, "计算值": value, "目标或容许上限": target, "结论": "通过" if passed else "未通过"})

    add("问题二逐时结果行数", len(problem_two_hourly), 2880, len(problem_two_hourly) == 2880)
    add("问题三逐时结果行数", len(problem_three_hourly), 2880, len(problem_three_hourly) == 2880)
    add("问题二场景产量组合数", len(metrics_two), 120, len(metrics_two) == 120)
    add("问题三场景产量组合数", len(metrics_three), 120, len(metrics_three) == 120)
    for name, frame in (("问题二", problem_two_hourly), ("问题三", problem_three_hourly)):
        balance = float(frame["功率平衡残差(MW)"].abs().max())
        simultaneous = float((frame["购电功率(MW)"] * frame["售电功率(MW)"]).abs().max())
        add(f"{name}最大功率平衡残差(MW)", balance, 1e-7, balance <= 1e-7)
        add(f"{name}购售电互斥残差(MW^2)", simultaneous, 1e-7, simultaneous <= 1e-7)
    selected = problem_three_hourly[np.isclose(problem_three_hourly["日氨产量(t/day)"], 36.0)]
    add("36 t/day场景数", selected["场景"].nunique(), 24, selected["场景"].nunique() == 24)
    scale_error = 0.0
    metric_columns = ["最大购电功率(MW)", "最大反送功率(MW)", "交换功率峰谷差(MW)"]
    for mode, group in aggregate.groupby("聚合方式"):
        base = group[group["等效园区数量"] == 1].iloc[0]
        for _, row in group.iterrows():
            count = row["等效园区数量"]
            scale_error = max(scale_error, *(abs(row[column] - count * base[column]) for column in metric_columns))
    add("多园区线性尺度最大残差(MW)", scale_error, 1e-9, scale_error <= 1e-9)
    add("量化证据条数", len(evidence), 9, len(evidence) == 9)
    add("五档产量稳健性行数", len(robustness), 5, len(robustness) == 5)
    add("极值来源记录数", len(extreme_scenarios), 4, len(extreme_scenarios) == 4)
    add("国内外文献及政策来源数", len(literature), 6, len(literature) >= 6)
    add("正负影响类型数", impacts["影响类型"].nunique(), 2, impacts["影响类型"].nunique() == 2)
    dual_evidence_complete = bool(
        policy["本题量化依据"].astype(str).str.len().gt(0).all()
        and policy["文献政策依据"].astype(str).str.len().gt(0).all()
    )
    add("政策双重依据完整性", int(dual_evidence_complete), 1, dual_evidence_complete)
    penetration_columns = sum("渗透率" in str(column) for column in aggregate.columns)
    add("渗透率数值指标列数", penetration_columns, 0, penetration_columns == 0)
    return pd.DataFrame(rows)


def save_figure(fig, directory, stem):
    fig.savefig(directory / f"{stem}.png", bbox_inches="tight")
    fig.savefig(directory / f"{stem}.pdf", bbox_inches="tight")
    plt.close(fig)


def plot_mean_exchange(profile, figure_dir):
    fig, ax = plt.subplots(figsize=(9.2, 4.8))
    styles = [("问题二离散调节", PALETTE["blue"], "--"), ("问题三连续调节", PALETTE["orange"], "-")]
    for model, color, linestyle in styles:
        selected = profile[profile["模型"] == model]
        ax.plot(
            selected["时段序号(h)"],
            selected["净交换功率(MW)"],
            color=color,
            linewidth=2.5,
            linestyle=linestyle,
            marker="o",
            markersize=3.5,
            label=model,
        )
    ax.axhline(0.0, color="#666666", linewidth=1.0)
    ax.fill_between([1, 24], 0, ax.get_ylim()[1], color=PALETTE["yellow"], alpha=0.08)
    ax.fill_between([1, 24], ax.get_ylim()[0], 0, color=PALETTE["light_blue"], alpha=0.08)
    ax.set_xlabel("时段 / h")
    ax.set_ylabel("24场景平均净交换功率 / MW")
    ax.set_xticks(range(1, 25, 2))
    ax.grid(axis="y", color="#D9DEE7", linewidth=0.8)
    ax.legend(frameon=False, ncol=2, loc="best")
    fig.tight_layout()
    save_figure(fig, figure_dir, "问题五连续调节净交换功率")


def plot_aggregate_impacts(aggregate, figure_dir):
    fig, axes = plt.subplots(1, 2, figsize=(10.4, 4.2), sharex=True)
    panels = [("最大购电功率(MW)", "最大购电功率 / MW"), ("最大反送功率(MW)", "最大反送功率 / MW")]
    styles = [("同步运行上界", PALETTE["red"], "o"), ("场景多样化期望", PALETTE["blue"], "s")]
    for ax, (column, ylabel) in zip(axes, panels):
        for mode, color, marker in styles:
            selected = aggregate[aggregate["聚合方式"] == mode]
            ax.plot(
                selected["等效园区数量"],
                selected[column],
                color=color,
                linewidth=2.3,
                marker=marker,
                markersize=5,
                label=mode,
            )
        ax.set_xlabel("等效园区数量 / 个")
        ax.set_ylabel(ylabel)
        ax.set_xticks(PARK_COUNTS)
        ax.grid(axis="y", color="#D9DEE7", linewidth=0.8)
    axes[0].legend(frameon=False, loc="upper left")
    fig.tight_layout()
    save_figure(fig, figure_dir, "问题五多园区聚合影响")


def write_excel(path, sheets):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="5271AE")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    workbook.save(path)


def write_report(
    path,
    comparison,
    aggregate,
    robustness,
    extreme_scenarios,
    impacts,
    evidence,
    literature,
    policy,
    validation,
    key_values,
):
    comparison_36 = comparison[np.isclose(comparison["日氨产量(t/day)"], 36.0)]
    average_changes = comparison_36[
        [
            "购电量(MWh)_变化",
            "售电量(MWh)_变化",
            "交换功率峰谷差(MW)_变化",
            "最大爬坡功率(MW/h)_变化",
        ]
    ].mean()
    aggregate_display = aggregate[
        [
            "等效园区数量",
            "聚合方式",
            "最大购电功率(MW)",
            "最大反送功率(MW)",
            "交换功率峰谷差(MW)",
            "最大爬坡功率(MW/h)",
        ]
    ]
    robustness_display = robustness[
        [
            "日氨产量(t/day)",
            "购电平均变化(MWh)",
            "售电平均变化(MWh)",
            "峰谷差平均变化(MW)",
            "最大爬坡平均变化(MW/h)",
            "购电减少场景数",
            "售电减少场景数",
            "峰谷差减少场景数",
            "爬坡减少场景数",
        ]
    ]
    robustness_36 = robustness[np.isclose(robustness["日氨产量(t/day)"], 36.0)].iloc[0]
    literature_display = literature.copy()
    literature_display["题名及链接"] = literature_display.apply(
        lambda row: f"[{row['题名']}]({row['链接']})", axis=1
    )
    literature_display = literature_display[
        ["文献编号", "类型", "作者或机构", "题名及链接", "与本题相关结论"]
    ]
    text = fr"""# 问题五计算结果

## 1. 计算定位

问题五属于机理分析与政策评价。代码不重新求解前四问，而是读取既有逐时调度和年度结果，构造园区对电网的净交换功率

\[
P_t^{{ex}}=P_t^{{buy}}-P_t^{{sell}},
\]

其中正值表示购电，负值表示向电网反送。随后计算最大购电、最大反送、交换功率峰谷差、最大爬坡和标准差，并用等效园区数量分析规模效应。

### 交换功率渗透率的概念口径

若已知区域电网峰值负荷，可将最大购电或最大反送功率与区域峰值负荷之比称为交换功率渗透率，用于描述园区交换功率相对区域电网规模的大小。本题未提供区域峰值负荷、变电站容量或区域年用电量，因此**交换功率渗透率仅作为概念表述，不作为计算指标，不给出任何百分比结果**。本文实际计算的仅为等效园区数量及其聚合交换功率，等效园区数量不等同于区域渗透率。

## 2. 连续调节对电网交换的影响

以36 t/day为代表口径，对24个风光场景取均值。问题三相对问题二的场景平均变化为：

- 购电量：{average_changes['购电量(MWh)_变化']:.3f} MWh/day；
- 售电量：{average_changes['售电量(MWh)_变化']:.3f} MWh/day；
- 交换功率峰谷差：{average_changes['交换功率峰谷差(MW)_变化']:.3f} MW；
- 最大爬坡：{average_changes['最大爬坡功率(MW/h)_变化']:.3f} MW/h；
- 单位吨氨柔性价值：{key_values['flexibility_value']:.2f} ¥/t。

负号表示连续调节在场景平均意义下降低相应指标。36 t/day下，购电、售电、峰谷差和爬坡分别在{int(robustness_36['购电减少场景数'])}/24、{int(robustness_36['售电减少场景数'])}/24、{int(robustness_36['峰谷差减少场景数'])}/24和{int(robustness_36['爬坡减少场景数'])}/24个场景中减少，因此不能表述为“所有场景均改善”。全部五档产量合计的“全部不达标”天数由{key_values['unqualified_two']} day降至{key_values['unqualified_three']} day。

五档产量的场景均值和改善场景数如下：

{robustness_display.to_markdown(index=False, floatfmt='.3f')}

72 t/day下两种调节方式没有实质差异；其余四档产量的平均购电量均下降，但售电、峰谷差和爬坡的改善具有场景差异。

## 3. 多园区聚合影响

“同步运行上界”表示各指标分别取24个单园区场景中的最大值再按数量线性扩展；“场景多样化期望”表示24个场景等权平均后再按等效园区数量扩展。这里的数量是规模参数，不等同于区域新能源渗透率。

{aggregate_display.to_markdown(index=False, floatfmt='.3f')}

当等效园区数量为20时，同步上界最大购电和最大反送分别为{key_values['n20_sync_import']:.2f} MW和{key_values['n20_sync_export']:.2f} MW；场景多样化期望分别为{key_values['n20_div_import']:.2f} MW和{key_values['n20_div_export']:.2f} MW。因此，大规模接入影响不仅取决于园区数量，也取决于风光场景及运行行为的相关性。该结果是园区边界功率的线性规模分析，不是含网络约束的区域潮流或稳定性仿真。

同步上界的各项极值来源如下。最大购电和最大反送允许来自不同场景，不能解释为同一时刻同时出现：

{extreme_scenarios.to_markdown(index=False, floatfmt='.3f')}

## 4. 对电力系统运行的正负影响

{impacts.to_markdown(index=False)}

电压、频率、电能质量和保护协调缺少区域电网拓扑、线路参数、短路容量及惯量数据，只作有文献政策依据的定性分析，不生成虚假数值。

## 5. 前四问量化证据

{evidence.to_markdown(index=False)}

## 6. 国内外文献与政策依据

{literature_display.to_markdown(index=False)}

上述资料分别支撑源荷匹配、交换功率边界、灵活调节、预测控制、电能质量、系统费用和辅助服务等判断；本题数据用于验证这些机理在给定园区中的具体表现，不将文献算例数值直接套用于本园区。

## 7. 政策建议映射

{policy.to_markdown(index=False)}

## 8. 一致性校验

{validation.to_markdown(index=False)}

所有“通过”仅表示计算文件、功率平衡、购售电互斥与聚合尺度关系满足当前模型口径。由于题目未提供区域电网拓扑、线路参数、短路容量和惯量数据，本问不虚构电压、频率、潮流或稳定性数值。

## 9. 输出文件与图表

- `outputs/问题五计算结果/问题五五档产量稳健性.csv`：五档产量的平均变化及改善场景数；
- `outputs/问题五计算结果/问题五极值对应场景.csv`：同步上界各极值的独立来源；
- `outputs/问题五计算结果/问题五电力系统影响汇总.csv`：正面与负面影响及双重依据；
- `outputs/问题五计算结果/问题五文献与政策依据.csv`：国内外文献和政策文件；

- `figures/问题五计算结果/问题五连续调节净交换功率.pdf`：36 t/day下问题二、三的24场景平均净交换功率；
- `figures/问题五计算结果/问题五多园区聚合影响.pdf`：不同等效园区数量下的最大购电与最大反送功率。

本次不增加新的图形，避免问题五的定性论述被装饰性可视化稀释。

## 10. 可复现运行方式

```powershell
python code/问题五.py
python code/问题五.py --self-test
```
"""
    path.write_text(text, encoding="utf-8")


def run(project_root):
    output_dir = project_root / "outputs" / "问题五计算结果"
    figure_dir = project_root / "figures" / "问题五计算结果"
    report_path = project_root / "reports" / "问题五计算结果.md"
    output_dir.mkdir(parents=True, exist_ok=True)
    figure_dir.mkdir(parents=True, exist_ok=True)

    problem_two_hourly = load_hourly_results(
        project_root / "outputs" / "问题二计算结果" / "问题二全部场景逐时排班.csv",
        "问题二离散调节",
    )
    problem_three_hourly = load_hourly_results(
        project_root / "outputs" / "问题三计算结果" / "问题三全部场景逐时调度.csv",
        "问题三连续调节",
    )
    metrics_two = build_scenario_metrics(problem_two_hourly)
    metrics_three = build_scenario_metrics(problem_three_hourly)
    all_metrics = pd.concat([metrics_two, metrics_three], ignore_index=True)
    comparison = compare_models(metrics_two, metrics_three)
    robustness = build_production_robustness(comparison)
    aggregate, scenario_profiles = build_aggregate_results(problem_three_hourly)
    extreme_scenarios = build_extreme_scenarios(metrics_three)
    mean_profiles = build_mean_exchange_profiles(problem_two_hourly, problem_three_hourly)
    literature = build_literature_sources()
    impacts = build_system_impact_summary()
    evidence, policy, key_values = build_evidence_tables(project_root, aggregate, extreme_scenarios)
    validation = build_validation(
        problem_two_hourly,
        problem_three_hourly,
        metrics_two,
        metrics_three,
        aggregate,
        evidence,
        robustness,
        extreme_scenarios,
        literature,
        impacts,
        policy,
    )
    if not validation["结论"].eq("通过").all():
        failed = validation[validation["结论"] != "通过"]
        raise RuntimeError(f"Validation failed:\n{failed.to_string(index=False)}")

    frames = {
        "场景电网影响指标": all_metrics,
        "连续调节影响比较": comparison,
        "五档产量稳健性": robustness,
        "多园区聚合影响": aggregate,
        "极值对应场景": extreme_scenarios,
        "平均净交换曲线": mean_profiles,
        "电力系统影响汇总": impacts,
        "量化证据": evidence,
        "文献与政策依据": literature,
        "政策建议映射": policy,
        "一致性校验": validation,
    }
    filenames = {
        "场景电网影响指标": "问题五电网影响指标.csv",
        "连续调节影响比较": "问题五连续调节影响比较.csv",
        "五档产量稳健性": "问题五五档产量稳健性.csv",
        "多园区聚合影响": "问题五多园区聚合影响.csv",
        "极值对应场景": "问题五极值对应场景.csv",
        "平均净交换曲线": "问题五平均净交换曲线.csv",
        "电力系统影响汇总": "问题五电力系统影响汇总.csv",
        "量化证据": "问题五量化证据.csv",
        "文献与政策依据": "问题五文献与政策依据.csv",
        "政策建议映射": "问题五政策建议映射.csv",
        "一致性校验": "问题五约束校验.csv",
    }
    for name, frame in frames.items():
        frame.to_csv(output_dir / filenames[name], index=False, encoding="utf-8-sig")
    write_excel(output_dir / "问题五计算结果.xlsx", frames)
    payload = {
        "口径": {
            "净交换功率": "购电功率-售电功率，正值购电，负值反送",
            "代表日氨产量(t/day)": REPRESENTATIVE_PRODUCTION,
            "等效园区数量": PARK_COUNTS,
            "场景权重": "24个风光组合等权",
            "交换功率渗透率": "仅作为概念表述；因缺少区域电网基准，不作为计算指标",
        },
        "关键结果": key_values,
        "五档产量稳健性": json.loads(robustness.to_json(orient="records", force_ascii=False)),
        "多园区聚合影响": json.loads(aggregate.to_json(orient="records", force_ascii=False)),
        "极值对应场景": json.loads(extreme_scenarios.to_json(orient="records", force_ascii=False)),
        "电力系统影响汇总": json.loads(impacts.to_json(orient="records", force_ascii=False)),
        "文献与政策依据": json.loads(literature.to_json(orient="records", force_ascii=False)),
        "一致性校验": json.loads(validation.to_json(orient="records", force_ascii=False)),
    }
    (output_dir / "问题五完整计算结果.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    configure_plotting()
    plot_mean_exchange(mean_profiles, figure_dir)
    plot_aggregate_impacts(aggregate, figure_dir)
    write_report(
        report_path,
        comparison,
        aggregate,
        robustness,
        extreme_scenarios,
        impacts,
        evidence,
        literature,
        policy,
        validation,
        key_values,
    )

    return {
        "outputs": output_dir,
        "figures": figure_dir,
        "report": report_path,
        "validation": validation,
        "scenario_profiles": scenario_profiles,
        "key_values": key_values,
    }


class ProblemFiveCoreTests(unittest.TestCase):
    def test_exchange_metrics_follow_sign_convention(self):
        metrics = calculate_exchange_metrics(np.array([2.0, -3.0, 1.0]))
        self.assertAlmostEqual(metrics["购电量(MWh)"], 3.0)
        self.assertAlmostEqual(metrics["售电量(MWh)"], 3.0)
        self.assertAlmostEqual(metrics["最大购电功率(MW)"], 2.0)
        self.assertAlmostEqual(metrics["最大反送功率(MW)"], 3.0)
        self.assertAlmostEqual(metrics["交换功率峰谷差(MW)"], 5.0)
        self.assertAlmostEqual(metrics["最大爬坡功率(MW/h)"], 5.0)

    def test_aggregate_metrics_scale_linearly(self):
        profiles = np.array([[1.0, -1.0], [3.0, -3.0]])
        result = calculate_aggregate_scale_metrics(profiles, [1, 5])
        diversified = result[result["聚合方式"] == "场景多样化期望"].reset_index(drop=True)
        synchronized = result[result["聚合方式"] == "同步运行上界"].reset_index(drop=True)
        self.assertAlmostEqual(diversified.loc[0, "最大购电功率(MW)"], 2.0)
        self.assertAlmostEqual(diversified.loc[1, "最大购电功率(MW)"], 10.0)
        self.assertAlmostEqual(synchronized.loc[0, "最大购电功率(MW)"], 3.0)
        self.assertAlmostEqual(synchronized.loc[1, "最大反送功率(MW)"], 15.0)

    def test_production_robustness_reports_average_and_improvement_counts(self):
        comparison = pd.DataFrame(
            {
                "日氨产量(t/day)": [36.0, 36.0],
                "购电量(MWh)_变化": [-1.0, 0.0],
                "售电量(MWh)_变化": [-2.0, 1.0],
                "交换功率峰谷差(MW)_变化": [-3.0, -1.0],
                "最大爬坡功率(MW/h)_变化": [1.0, -1.0],
            }
        )
        result = build_production_robustness(comparison).iloc[0]
        self.assertAlmostEqual(result["峰谷差平均变化(MW)"], -2.0)
        self.assertEqual(result["购电减少场景数"], 1)
        self.assertEqual(result["售电减少场景数"], 1)
        self.assertEqual(result["峰谷差减少场景数"], 2)
        self.assertEqual(result["爬坡减少场景数"], 1)

    def test_extreme_scenarios_preserve_separate_sources(self):
        metrics = pd.DataFrame(
            {
                "场景": ["风电1-光伏1", "风电2-光伏2"],
                "风电场景": [1, 2],
                "光伏场景": [1, 2],
                "日氨产量(t/day)": [36.0, 36.0],
                "最大购电功率(MW)": [10.0, 5.0],
                "最大反送功率(MW)": [2.0, 20.0],
                "交换功率峰谷差(MW)": [12.0, 25.0],
                "最大爬坡功率(MW/h)": [3.0, 8.0],
            }
        )
        result = build_extreme_scenarios(metrics).set_index("极值指标")
        self.assertEqual(result.loc["最大购电功率", "对应场景"], "风电1-光伏1")
        self.assertEqual(result.loc["最大反送功率", "对应场景"], "风电2-光伏2")

    def test_literature_table_contains_policy_and_domestic_foreign_research(self):
        result = build_literature_sources()
        self.assertGreaterEqual(len(result), 6)
        self.assertTrue({"政策文件", "国内研究", "国外研究"}.issubset(set(result["类型"])))
        self.assertTrue(result["链接"].str.startswith("http").all())

    def test_impact_summary_has_positive_negative_and_dual_evidence(self):
        result = build_system_impact_summary()
        self.assertTrue({"正面影响", "负面影响"}.issubset(set(result["影响类型"])))
        self.assertTrue(result["本题量化依据"].str.len().gt(0).all())
        self.assertTrue(result["文献政策依据"].str.len().gt(0).all())


def run_self_tests():
    suite = unittest.defaultTestLoader.loadTestsFromTestCase(ProblemFiveCoreTests)
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    if not result.wasSuccessful():
        raise SystemExit(1)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--project-root", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()
    if args.self_test:
        run_self_tests()
        return
    result = run(args.project_root.resolve())
    print(f"Outputs: {result['outputs']}")
    print(f"Figures: {result['figures']}")
    print(f"Report: {result['report']}")
    print(result["validation"].to_string(index=False))


if __name__ == "__main__":
    main()
