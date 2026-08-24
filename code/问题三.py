import argparse
import json
import math
import re
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import coo_matrix, csr_matrix, vstack

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap


HOURS = 24
TIME_STEP_H = 1.0
CONVENTIONAL_LOAD_BASE_MW = 6.0
WIND_CAPACITY_MW = 40.0
PV_CAPACITY_MW = 64.0
ALK_MAX_MW = 20.0
PEM_MAX_MW = 20.0
AMMONIA_MAX_MW = 1.5
ALK_MIN_MW = 2.0
PEM_MIN_MW = 2.0
AMMONIA_MIN_MW = 0.15
HYDROGEN_RATE_KG_PER_H = 600.0
DAILY_OUTPUTS_T = (72, 63, 54, 45, 36)
SCENARIO_DAYS = 15
STRICT_TOLERANCE = 1e-9
PALETTE = {
    "blue": "#5271AE",
    "light_blue": "#70ACDE",
    "yellow": "#F5CC7D",
    "orange": "#FFA660",
    "red": "#D85B59",
}


def find_attachment(project_root, prefix):
    matches = [
        path
        for path in Path(project_root).joinpath("题目").rglob("*.xlsx")
        if not path.name.startswith("~$") and path.name.startswith(prefix)
    ]
    if len(matches) != 1:
        raise FileNotFoundError(f"Expected one {prefix} workbook, found {len(matches)}")
    return matches[0]


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell.value for cell in row) for row in workbook.active.iter_rows()]
    finally:
        workbook.close()


def extract_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        raise ValueError(f"No numeric value found in {value!r}")
    return float(match.group())


def hourly_buy_prices(rows):
    peak_price = extract_number(rows[1][1])
    flat_price = extract_number(rows[2][1])
    valley_price = extract_number(rows[3][1])
    prices = []
    for hour in range(HOURS):
        if 10 <= hour < 15 or 18 <= hour < 21:
            prices.append(peak_price)
        elif 7 <= hour < 10 or 15 <= hour < 18 or 21 <= hour < 23:
            prices.append(flat_price)
        else:
            prices.append(valley_price)
    return np.asarray(prices, dtype=float)


def load_inputs(project_root):
    load_rows = read_rows(find_attachment(project_root, "附件1"))
    typical_rows = read_rows(find_attachment(project_root, "附件2"))
    wind_rows = read_rows(find_attachment(project_root, "附件3"))
    pv_rows = read_rows(find_attachment(project_root, "附件4"))
    equipment_rows = read_rows(find_attachment(project_root, "附件5"))
    ammonia_rows = read_rows(find_attachment(project_root, "附件6"))
    buy_rows = read_rows(find_attachment(project_root, "附件7"))
    sell_rows = read_rows(find_attachment(project_root, "附件8"))

    time_labels = [str(row[0]) for row in load_rows[1:]]
    load_pu = np.asarray([row[1] for row in load_rows[1:]], dtype=float)
    typical_times = [str(row[0]) for row in typical_rows[1:]]
    wind_times = [str(row[0]) for row in wind_rows[1:]]
    pv_times = [str(row[0]) for row in pv_rows[1:]]
    if not (
        len(time_labels) == HOURS
        and time_labels == typical_times
        and time_labels == wind_times
        and time_labels == pv_times
    ):
        raise ValueError("Hourly input tables must contain the same 24 periods")

    wind_scenarios_pu = np.asarray([row[1:7] for row in wind_rows[1:]], dtype=float).T
    pv_scenarios_pu = np.asarray([row[1:5] for row in pv_rows[1:]], dtype=float).T
    if wind_scenarios_pu.shape != (6, HOURS) or pv_scenarios_pu.shape != (4, HOURS):
        raise ValueError("Scenario tables must have dimensions 6x24 and 4x24")

    sell_wind_price = extract_number(sell_rows[1][1])
    sell_pv_price = extract_number(sell_rows[2][1])
    if not np.isclose(sell_wind_price, sell_pv_price):
        raise ValueError("Aggregate surplus requires a common feed-in price")

    return {
        "time_labels": time_labels,
        "conventional_load": CONVENTIONAL_LOAD_BASE_MW * load_pu,
        "wind_scenarios": WIND_CAPACITY_MW * wind_scenarios_pu,
        "pv_scenarios": PV_CAPACITY_MW * pv_scenarios_pu,
        "buy_prices": hourly_buy_prices(buy_rows),
        "sell_price": sell_wind_price,
        "wind_cost": extract_number(equipment_rows[1][1]),
        "pv_cost": extract_number(equipment_rows[1][2]),
        "alk_om": extract_number(equipment_rows[2][3]),
        "pem_om": extract_number(equipment_rows[2][4]),
        "ammonia_investment": extract_number(ammonia_rows[1][2]),
        "ammonia_om": extract_number(ammonia_rows[2][2]),
        "ammonia_lifetime": extract_number(ammonia_rows[3][2]),
    }


def calculate_green_metrics(load_energy, generation_energy, buy_energy, sell_energy):
    if load_energy <= 0 or generation_energy <= 0:
        raise ValueError("Green metric denominators must be positive")
    return {
        "r1": (load_energy - sell_energy - buy_energy) / generation_energy,
        "r2": (generation_energy - sell_energy) / load_energy,
        "r3": sell_energy / generation_energy,
    }


def classify_green_metrics(metrics):
    status = {
        "新能源自发自用比例达标": metrics["r1"] > 0.60 + STRICT_TOLERANCE,
        "新能源供电占比达标": metrics["r2"] > 0.30 + STRICT_TOLERANCE,
        "上网电量比例达标": metrics["r3"] < 0.20 - STRICT_TOLERANCE,
    }
    count = int(sum(status.values()))
    if count == 3:
        conclusion = "全部达标"
    elif count == 0:
        conclusion = "全部不达标"
    else:
        conclusion = "部分达标"
    return status, count, conclusion


def is_semicontinuous(power, rated_power, tolerance=1e-9):
    return abs(power) <= tolerance or (
        0.1 * rated_power - tolerance <= power <= rated_power + tolerance
    )


def material_residual(alk_power, pem_power, ammonia_power):
    return 14.0 * alk_power + 16.0 * pem_power - 400.0 * ammonia_power


def required_ammonia_power_energy(output_t):
    return 0.5 * float(output_t)


def variable_slices():
    names = (
        "alk",
        "pem",
        "ammonia",
        "buy",
        "sell",
        "u_alk",
        "u_pem",
        "u_ammonia",
        "grid_mode",
    )
    return {name: slice(i * HOURS, (i + 1) * HOURS) for i, name in enumerate(names)}


def build_core_milp(conventional_load, generation, buy_prices, sell_price, output_t, om_costs):
    conventional_load = np.asarray(conventional_load, dtype=float)
    generation = np.asarray(generation, dtype=float)
    buy_prices = np.asarray(buy_prices, dtype=float)
    if conventional_load.shape != (HOURS,) or generation.shape != (HOURS,) or buy_prices.shape != (HOURS,):
        raise ValueError("Core MILP inputs must contain 24 hourly values")

    slices = variable_slices()
    n = HOURS * len(slices)
    primary = np.zeros(n, dtype=float)
    primary[slices["alk"]] = 1000.0 * om_costs["alk"]
    primary[slices["pem"]] = 1000.0 * om_costs["pem"]
    primary[slices["ammonia"]] = 1000.0 * om_costs["ammonia"]
    primary[slices["buy"]] = 1000.0 * buy_prices
    primary[slices["sell"]] = -1000.0 * sell_price

    lower = np.zeros(n, dtype=float)
    upper = np.full(n, np.inf, dtype=float)
    upper[slices["alk"]] = ALK_MAX_MW
    upper[slices["pem"]] = PEM_MAX_MW
    upper[slices["ammonia"]] = AMMONIA_MAX_MW
    upper[slices["buy"]] = conventional_load + ALK_MAX_MW + PEM_MAX_MW + AMMONIA_MAX_MW
    upper[slices["sell"]] = generation
    for name in ("u_alk", "u_pem", "u_ammonia", "grid_mode"):
        upper[slices[name]] = 1.0

    integrality = np.zeros(n, dtype=int)
    for name in ("u_alk", "u_pem", "u_ammonia", "grid_mode"):
        integrality[slices[name]] = 1

    rows = []
    cols = []
    vals = []
    row_lower = []
    row_upper = []
    row = 0

    def add(entries, lo, hi):
        nonlocal row
        for col, value in entries:
            rows.append(row)
            cols.append(col)
            vals.append(value)
        row_lower.append(lo)
        row_upper.append(hi)
        row += 1

    for t in range(HOURS):
        add(
            [
                (slices["alk"].start + t, 1.0),
                (slices["pem"].start + t, 1.0),
                (slices["ammonia"].start + t, 1.0),
                (slices["buy"].start + t, -1.0),
                (slices["sell"].start + t, 1.0),
            ],
            generation[t] - conventional_load[t],
            generation[t] - conventional_load[t],
        )
        add(
            [
                (slices["alk"].start + t, 14.0),
                (slices["pem"].start + t, 16.0),
                (slices["ammonia"].start + t, -400.0),
            ],
            0.0,
            0.0,
        )
        for power_name, state_name, minimum, maximum in (
            ("alk", "u_alk", ALK_MIN_MW, ALK_MAX_MW),
            ("pem", "u_pem", PEM_MIN_MW, PEM_MAX_MW),
            ("ammonia", "u_ammonia", AMMONIA_MIN_MW, AMMONIA_MAX_MW),
        ):
            add(
                [
                    (slices[power_name].start + t, 1.0),
                    (slices[state_name].start + t, -minimum),
                ],
                0.0,
                np.inf,
            )
            add(
                [
                    (slices[power_name].start + t, 1.0),
                    (slices[state_name].start + t, -maximum),
                ],
                -np.inf,
                0.0,
            )
        add(
            [
                (slices["buy"].start + t, 1.0),
                (slices["grid_mode"].start + t, -(conventional_load[t] + 41.5)),
            ],
            -np.inf,
            0.0,
        )
        add(
            [
                (slices["sell"].start + t, 1.0),
                (slices["grid_mode"].start + t, generation[t]),
            ],
            -np.inf,
            generation[t],
        )

    add(
        [(slices["ammonia"].start + t, 1.0) for t in range(HOURS)],
        required_ammonia_power_energy(output_t),
        required_ammonia_power_energy(output_t),
    )

    matrix = coo_matrix((vals, (rows, cols)), shape=(row, n)).tocsr()
    constraint = LinearConstraint(matrix, np.asarray(row_lower), np.asarray(row_upper))
    return primary, integrality, Bounds(lower, upper), constraint, slices


def add_lexicographic_caps(base_constraint, primary, cost_cap=None, buy_vector=None, buy_cap=None):
    matrices = [base_constraint.A]
    lower = list(np.asarray(base_constraint.lb, dtype=float))
    upper = list(np.asarray(base_constraint.ub, dtype=float))
    if cost_cap is not None:
        matrices.append(csr_matrix(np.asarray(primary, dtype=float).reshape(1, -1)))
        lower.append(-np.inf)
        upper.append(float(cost_cap))
    if buy_cap is not None:
        matrices.append(csr_matrix(np.asarray(buy_vector, dtype=float).reshape(1, -1)))
        lower.append(-np.inf)
        upper.append(float(buy_cap))
    return LinearConstraint(vstack(matrices, format="csr"), np.asarray(lower), np.asarray(upper))


def run_milp(objective, integrality, bounds, constraints):
    return milp(
        c=objective,
        integrality=integrality,
        bounds=bounds,
        constraints=constraints,
        options={"mip_rel_gap": 1e-10, "time_limit": 60.0, "presolve": True},
    )


def solve_core_case(conventional_load, generation, buy_prices, sell_price, output_t, om_costs):
    primary, integrality, bounds, constraint, slices = build_core_milp(
        conventional_load, generation, buy_prices, sell_price, output_t, om_costs
    )
    result = run_milp(primary, integrality, bounds, constraint)
    if not result.success:
        raise RuntimeError(result.message)
    return result, slices, primary


def solve_lexicographic_case(conventional_load, generation, buy_prices, sell_price, output_t, om_costs):
    primary, integrality, bounds, base_constraint, slices = build_core_milp(
        conventional_load, generation, buy_prices, sell_price, output_t, om_costs
    )
    first = run_milp(primary, integrality, bounds, base_constraint)
    if not first.success:
        raise RuntimeError(first.message)

    buy_objective = np.zeros_like(primary)
    buy_objective[slices["buy"]] = 1.0
    cost_tolerance = max(1e-5, abs(float(first.fun)) * 1e-10)
    second_constraint = add_lexicographic_caps(
        base_constraint, primary, cost_cap=float(first.fun) + cost_tolerance
    )
    second = run_milp(buy_objective, integrality, bounds, second_constraint)
    if not second.success:
        raise RuntimeError(second.message)

    sell_objective = np.zeros_like(primary)
    sell_objective[slices["sell"]] = 1.0
    third_constraint = add_lexicographic_caps(
        base_constraint,
        primary,
        cost_cap=float(first.fun) + cost_tolerance,
        buy_vector=buy_objective,
        buy_cap=float(second.fun) + 1e-5,
    )
    third = run_milp(sell_objective, integrality, bounds, third_constraint)
    if not third.success:
        raise RuntimeError(third.message)
    return third, slices, primary, first, second


def clean_array(values, tolerance=1e-8):
    array = np.asarray(values, dtype=float).copy()
    array[np.abs(array) < tolerance] = 0.0
    return array


def cost_components(data, wind_power, pv_power, alk, pem, ammonia, buy, sell, output_t):
    wind_energy = float(np.sum(wind_power) * TIME_STEP_H)
    pv_energy = float(np.sum(pv_power) * TIME_STEP_H)
    wind_cost = 1000.0 * data["wind_cost"] * wind_energy
    pv_cost = 1000.0 * data["pv_cost"] * pv_energy
    buy_cost = float(1000.0 * np.sum(buy * data["buy_prices"]) * TIME_STEP_H)
    sell_revenue = float(1000.0 * data["sell_price"] * np.sum(sell) * TIME_STEP_H)
    alk_om = float(1000.0 * data["alk_om"] * np.sum(alk) * TIME_STEP_H)
    pem_om = float(1000.0 * data["pem_om"] * np.sum(pem) * TIME_STEP_H)
    ammonia_om = float(1000.0 * data["ammonia_om"] * np.sum(ammonia) * TIME_STEP_H)
    capital = float(
        data["ammonia_investment"]
        * HYDROGEN_RATE_KG_PER_H
        / data["ammonia_lifetime"]
        / 365.0
    )
    net_cost = wind_cost + pv_cost + buy_cost - sell_revenue + alk_om + pem_om + ammonia_om + capital
    return {
        "风电发电成本(¥/day)": wind_cost,
        "光伏发电成本(¥/day)": pv_cost,
        "购电成本(¥/day)": buy_cost,
        "售电收入(¥/day)": sell_revenue,
        "碱性电解槽运维成本(¥/day)": alk_om,
        "PEM电解槽运维成本(¥/day)": pem_om,
        "合成氨装置运维成本(¥/day)": ammonia_om,
        "合成氨装置资本分摊(¥/day)": capital,
        "日净成本(¥/day)": net_cost,
        "净吨氨成本(¥/t)": net_cost / output_t,
    }


def semicontinuous_violation(power, state, minimum, maximum):
    lower_violation = np.maximum(minimum * state - power, 0.0)
    upper_violation = np.maximum(power - maximum * state, 0.0)
    return float(max(np.max(lower_violation), np.max(upper_violation)))


def solve_case(data, wind_power, pv_power, output_t, scenario_name, wind_index, pv_index):
    generation = wind_power + pv_power
    om_costs = {
        "alk": data["alk_om"],
        "pem": data["pem_om"],
        "ammonia": data["ammonia_om"],
    }
    result, slices, primary, first, second = solve_lexicographic_case(
        data["conventional_load"],
        generation,
        data["buy_prices"],
        data["sell_price"],
        output_t,
        om_costs,
    )
    x = clean_array(result.x)
    alk = x[slices["alk"]]
    pem = x[slices["pem"]]
    ammonia = x[slices["ammonia"]]
    u_alk = np.rint(x[slices["u_alk"]]).astype(int)
    u_pem = np.rint(x[slices["u_pem"]]).astype(int)
    u_ammonia = np.rint(x[slices["u_ammonia"]]).astype(int)
    grid_mode = np.rint(x[slices["grid_mode"]]).astype(int)
    total_load = data["conventional_load"] + alk + pem + ammonia
    net_load = total_load - generation
    buy = np.maximum(net_load, 0.0)
    sell = np.maximum(-net_load, 0.0)

    load_energy = float(np.sum(total_load) * TIME_STEP_H)
    generation_energy = float(np.sum(generation) * TIME_STEP_H)
    buy_energy = float(np.sum(buy) * TIME_STEP_H)
    sell_energy = float(np.sum(sell) * TIME_STEP_H)
    metrics = calculate_green_metrics(load_energy, generation_energy, buy_energy, sell_energy)
    metric_status, metric_count, conclusion = classify_green_metrics(metrics)
    costs = cost_components(data, wind_power, pv_power, alk, pem, ammonia, buy, sell, output_t)

    hydrogen_total = float(np.sum(14.0 * alk + 16.0 * pem))
    pem_hydrogen_share = float(np.sum(16.0 * pem) / hydrogen_total) if hydrogen_total > 0 else 0.0
    high_mask = np.isclose(data["buy_prices"], np.max(data["buy_prices"]))
    high_process_energy = float(np.sum((alk + pem + ammonia)[high_mask]))

    summary = {
        "场景": scenario_name,
        "风电场景": wind_index,
        "光伏场景": pv_index,
        "日氨产量(t/day)": float(output_t),
        "总负荷电量(MWh)": load_energy,
        "常规负荷电量(MWh)": float(np.sum(data["conventional_load"])),
        "生产负荷电量(MWh)": float(np.sum(alk + pem + ammonia)),
        "碱性电解槽用电量(MWh)": float(np.sum(alk)),
        "PEM电解槽用电量(MWh)": float(np.sum(pem)),
        "合成氨装置用电量(MWh)": float(np.sum(ammonia)),
        "碱性电解槽运行时数(h)": int(np.sum(u_alk)),
        "PEM电解槽运行时数(h)": int(np.sum(u_pem)),
        "合成氨装置运行时数(h)": int(np.sum(u_ammonia)),
        "碱性电解槽利用率(%)": float(np.sum(alk) / (ALK_MAX_MW * HOURS) * 100.0),
        "PEM电解槽利用率(%)": float(np.sum(pem) / (PEM_MAX_MW * HOURS) * 100.0),
        "合成氨装置利用率(%)": float(np.sum(ammonia) / (AMMONIA_MAX_MW * HOURS) * 100.0),
        "PEM制氢占比(%)": pem_hydrogen_share * 100.0,
        "高电价时段生产用电量(MWh)": high_process_energy,
        "风电电量(MWh)": float(np.sum(wind_power)),
        "光伏电量(MWh)": float(np.sum(pv_power)),
        "新能源发电量(MWh)": generation_energy,
        "购电量(MWh)": buy_energy,
        "售电量(MWh)": sell_energy,
        "新能源自发自用比例(%)": metrics["r1"] * 100.0,
        "新能源供电占比(%)": metrics["r2"] * 100.0,
        "上网电量比例(%)": metrics["r3"] * 100.0,
        **metric_status,
        "达标指标数": metric_count,
        "达标状态": conclusion,
        **costs,
        "求解器": "SciPy HiGHS MILP",
        "一级最优变动净成本(¥/day)": float(first.fun),
        "二级最小购电量(MWh)": float(second.fun),
        "三级最小售电量(MWh)": float(result.fun),
    }

    hourly = []
    for t in range(HOURS):
        hourly.append(
            {
                "场景": scenario_name,
                "风电场景": wind_index,
                "光伏场景": pv_index,
                "日氨产量(t/day)": float(output_t),
                "时段": data["time_labels"][t],
                "时段序号(h)": t + 1,
                "常规负荷功率(MW)": float(data["conventional_load"][t]),
                "风电功率(MW)": float(wind_power[t]),
                "光伏功率(MW)": float(pv_power[t]),
                "新能源总功率(MW)": float(generation[t]),
                "碱性电解槽功率(MW)": float(alk[t]),
                "PEM电解槽功率(MW)": float(pem[t]),
                "合成氨装置功率(MW)": float(ammonia[t]),
                "总负荷功率(MW)": float(total_load[t]),
                "购电功率(MW)": float(buy[t]),
                "售电功率(MW)": float(sell[t]),
                "碱性电解槽状态": int(u_alk[t]),
                "PEM电解槽状态": int(u_pem[t]),
                "合成氨装置状态": int(u_ammonia[t]),
                "电网购电模式": int(grid_mode[t]),
                "碱性制氢量(kg/h)": float(14.0 * alk[t]),
                "PEM制氢量(kg/h)": float(16.0 * pem[t]),
                "合成氨产量(kg/h)": float(2000.0 * ammonia[t]),
                "购电电价(¥/kWh)": float(data["buy_prices"][t]),
                "售电电价(¥/kWh)": float(data["sell_price"]),
            }
        )

    material = material_residual(alk, pem, ammonia)
    power = generation + buy - total_load - sell
    variable_cost = (
        costs["购电成本(¥/day)"]
        - costs["售电收入(¥/day)"]
        + costs["碱性电解槽运维成本(¥/day)"]
        + costs["PEM电解槽运维成本(¥/day)"]
        + costs["合成氨装置运维成本(¥/day)"]
    )
    validation = {
        "场景": scenario_name,
        "日氨产量(t/day)": float(output_t),
        "最大功率平衡残差(MW)": float(np.max(np.abs(power))),
        "最大氢氨物料平衡残差(kg/h)": float(np.max(np.abs(material))),
        "日产量约束残差(MWh)": float(abs(np.sum(ammonia) - required_ammonia_power_energy(output_t))),
        "半连续边界最大违反量(MW)": max(
            semicontinuous_violation(alk, u_alk, ALK_MIN_MW, ALK_MAX_MW),
            semicontinuous_violation(pem, u_pem, PEM_MIN_MW, PEM_MAX_MW),
            semicontinuous_violation(ammonia, u_ammonia, AMMONIA_MIN_MW, AMMONIA_MAX_MW),
        ),
        "购售电同时发生最大值(MW2)": float(np.max(buy * sell)),
        "一级目标重算残差(¥/day)": float(abs(primary @ x - variable_cost)),
        "三级成本容差(¥/day)": float(primary @ x - first.fun),
    }
    return summary, hourly, validation


def scenario_features(data):
    rows = []
    conventional = data["conventional_load"]
    high_mask = np.isclose(data["buy_prices"], np.max(data["buy_prices"]))
    for wi, wind in enumerate(data["wind_scenarios"], start=1):
        for pi, pv in enumerate(data["pv_scenarios"], start=1):
            generation = wind + pv
            generation_energy = float(np.sum(generation))
            residual = conventional - generation
            deficit = float(np.sum(np.maximum(residual, 0.0)))
            surplus = float(np.sum(np.maximum(-residual, 0.0)))
            high_energy = float(np.sum(generation[high_mask]))
            rows.append(
                {
                    "场景": f"风电{wi}-光伏{pi}",
                    "风电场景": wi,
                    "光伏场景": pi,
                    "新能源总发电量(MWh)": generation_energy,
                    "高电价时段新能源电量(MWh)": high_energy,
                    "高电价时段新能源占比(%)": high_energy / generation_energy * 100.0,
                    "原始缺电面积(MWh)": deficit,
                    "原始余电面积(MWh)": surplus,
                }
            )
    return pd.DataFrame(rows)


def annual_summary(summary):
    rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = summary[summary["日氨产量(t/day)"] == output_t]
        annual_cost = float(subset["日净成本(¥/day)"].sum() * SCENARIO_DAYS)
        annual_production = float(360.0 * output_t)
        load = float(subset["总负荷电量(MWh)"].sum() * SCENARIO_DAYS)
        generation = float(subset["新能源发电量(MWh)"].sum() * SCENARIO_DAYS)
        buy = float(subset["购电量(MWh)"].sum() * SCENARIO_DAYS)
        sell = float(subset["售电量(MWh)"].sum() * SCENARIO_DAYS)
        metrics = calculate_green_metrics(load, generation, buy, sell)
        rows.append(
            {
                "日氨产量(t/day)": output_t,
                "年氨产量(t/year)": annual_production,
                "年综合净成本(¥/year)": annual_cost,
                "综合净吨氨成本(¥/t)": annual_cost / annual_production,
                "全部达标天数(day)": int((subset["达标状态"] == "全部达标").sum() * SCENARIO_DAYS),
                "部分达标天数(day)": int((subset["达标状态"] == "部分达标").sum() * SCENARIO_DAYS),
                "全部不达标天数(day)": int((subset["达标状态"] == "全部不达标").sum() * SCENARIO_DAYS),
                "全年新能源自发自用比例(%)": metrics["r1"] * 100.0,
                "全年新能源供电占比(%)": metrics["r2"] * 100.0,
                "全年上网电量比例(%)": metrics["r3"] * 100.0,
            }
        )
    return pd.DataFrame(rows)


def annual_tonne_cost_distribution(summary, output_t, scenario_days=SCENARIO_DAYS):
    values = summary.loc[
        summary["日氨产量(t/day)"] == output_t,
        "净吨氨成本(¥/t)",
    ].to_numpy(dtype=float)
    return np.sort(np.repeat(values, scenario_days))[::-1]


def scenario_operating_summary(summary, annual):
    rows = []
    for output_t in sorted(summary["日氨产量(t/day)"].unique(), reverse=True):
        subset = summary[summary["日氨产量(t/day)"] == output_t]
        rows.append(
            {
                "日氨产量(t/day)": output_t,
                "平均购电量(MWh/day)": float(subset["购电量(MWh)"].mean()),
                "购电量最小值(MWh/day)": float(subset["购电量(MWh)"].min()),
                "购电量最大值(MWh/day)": float(subset["购电量(MWh)"].max()),
                "平均售电量(MWh/day)": float(subset["售电量(MWh)"].mean()),
                "售电量最小值(MWh/day)": float(subset["售电量(MWh)"].min()),
                "售电量最大值(MWh/day)": float(subset["售电量(MWh)"].max()),
                "平均吨氨成本(¥/t)": float(subset["净吨氨成本(¥/t)"].mean()),
                "吨氨成本最小值(¥/t)": float(subset["净吨氨成本(¥/t)"].min()),
                "吨氨成本最大值(¥/t)": float(subset["净吨氨成本(¥/t)"].max()),
            }
        )
    overview = pd.DataFrame(rows)
    annual_columns = [
        "日氨产量(t/day)",
        "全年新能源自发自用比例(%)",
        "全年新能源供电占比(%)",
        "全年上网电量比例(%)",
        "全部达标天数(day)",
        "部分达标天数(day)",
        "全部不达标天数(day)",
    ]
    return overview.merge(annual[annual_columns], on="日氨产量(t/day)", how="left")


def representative_scenario_analysis(merged, lowest_cost_output):
    subset = merged[merged["日氨产量(t/day)"] == lowest_cost_output].copy()
    median_cost = float(subset["净吨氨成本(¥/t)"].median())
    indices = [
        ("最低成本", subset["净吨氨成本(¥/t)"].idxmin()),
        ("中位成本", (subset["净吨氨成本(¥/t)"] - median_cost).abs().idxmin()),
        ("最高成本", subset["净吨氨成本(¥/t)"].idxmax()),
    ]
    rows = []
    for label, index in indices:
        row = subset.loc[index]
        rows.append(
            {
                "代表类型": label,
                "场景": row["场景"],
                "日氨产量(t/day)": row["日氨产量(t/day)"],
                "新能源总发电量(MWh)": row["新能源总发电量(MWh)"],
                "高电价时段新能源占比(%)": row["高电价时段新能源占比(%)"],
                "原始缺电面积(MWh)": row["原始缺电面积(MWh)"],
                "购电量(MWh)": row["购电量(MWh)"],
                "售电量(MWh)": row["售电量(MWh)"],
                "购电成本(¥/day)": row["购电成本(¥/day)"],
                "售电收入(¥/day)": row["售电收入(¥/day)"],
                "设备运维成本(¥/day)": row["碱性电解槽运维成本(¥/day)"]
                + row["PEM电解槽运维成本(¥/day)"]
                + row["合成氨装置运维成本(¥/day)"],
                "净吨氨成本(¥/t)": row["净吨氨成本(¥/t)"],
                "新能源自发自用比例(%)": row["新能源自发自用比例(%)"],
                "新能源供电占比(%)": row["新能源供电占比(%)"],
                "上网电量比例(%)": row["上网电量比例(%)"],
                "达标状态": row["达标状态"],
            }
        )
    return pd.DataFrame(rows)


def annual_green_comparison(discrete_summary, continuous_summary, scenario_days=SCENARIO_DAYS):
    rows = []
    outputs = sorted(
        set(discrete_summary["日氨产量(t/day)"]).intersection(continuous_summary["日氨产量(t/day)"]),
        reverse=True,
    )
    for output_t in outputs:
        discrete = discrete_summary[discrete_summary["日氨产量(t/day)"] == output_t]
        continuous = continuous_summary[continuous_summary["日氨产量(t/day)"] == output_t]
        metrics = {}
        for name, subset in (("问题二", discrete), ("问题三", continuous)):
            values = calculate_green_metrics(
                float(subset["总负荷电量(MWh)"].sum()),
                float(subset["新能源发电量(MWh)"].sum()),
                float(subset["购电量(MWh)"].sum()),
                float(subset["售电量(MWh)"].sum()),
            )
            metrics[name] = {key: value * 100.0 for key, value in values.items()}
        row = {
            "日氨产量(t/day)": output_t,
            "问题二新能源自发自用比例(%)": metrics["问题二"]["r1"],
            "问题三新能源自发自用比例(%)": metrics["问题三"]["r1"],
            "新能源自发自用比例变化(百分点)": metrics["问题三"]["r1"] - metrics["问题二"]["r1"],
            "问题二新能源供电占比(%)": metrics["问题二"]["r2"],
            "问题三新能源供电占比(%)": metrics["问题三"]["r2"],
            "新能源供电占比变化(百分点)": metrics["问题三"]["r2"] - metrics["问题二"]["r2"],
            "问题二上网电量比例(%)": metrics["问题二"]["r3"],
            "问题三上网电量比例(%)": metrics["问题三"]["r3"],
            "上网电量比例变化(百分点)": metrics["问题三"]["r3"] - metrics["问题二"]["r3"],
        }
        for status in ("全部达标", "部分达标", "全部不达标"):
            row[f"问题二{status}天数(day)"] = int((discrete["达标状态"] == status).sum() * scenario_days)
            row[f"问题三{status}天数(day)"] = int((continuous["达标状态"] == status).sum() * scenario_days)
        rows.append(row)
    return pd.DataFrame(rows)


def descriptive_analysis(summary, features, best_output):
    merged = summary.merge(features, on=["场景", "风电场景", "光伏场景"], how="left")
    feature_columns = [
        "新能源总发电量(MWh)",
        "高电价时段新能源占比(%)",
        "原始缺电面积(MWh)",
    ]
    result_columns = [
        "购电量(MWh)",
        "售电量(MWh)",
        "净吨氨成本(¥/t)",
        "新能源自发自用比例(%)",
        "新能源供电占比(%)",
        "上网电量比例(%)",
    ]
    correlation_rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = merged[merged["日氨产量(t/day)"] == output_t]
        corr = subset[feature_columns + result_columns].corr(method="spearman")
        for feature in feature_columns:
            for result in result_columns:
                correlation_rows.append(
                    {
                        "日氨产量(t/day)": output_t,
                        "场景特征": feature,
                        "运行结果": result,
                        "Spearman描述性相关系数": float(corr.loc[feature, result]),
                    }
                )

    best_subset = merged[merged["日氨产量(t/day)"] == best_output]
    marginal_rows = []
    for dimension, column in (("风电", "风电场景"), ("光伏", "光伏场景")):
        for level, group in best_subset.groupby(column):
            marginal_rows.append(
                {
                    "最低成本对应日产量(t/day)": best_output,
                    "因素": dimension,
                    "水平": int(level),
                    "平均吨氨成本(¥/t)": float(group["净吨氨成本(¥/t)"].mean()),
                    "平均购电量(MWh)": float(group["购电量(MWh)"].mean()),
                    "平均售电量(MWh)": float(group["售电量(MWh)"].mean()),
                    "平均达标指标数": float(group["达标指标数"].mean()),
                }
            )
    return merged, pd.DataFrame(correlation_rows), pd.DataFrame(marginal_rows)


def compare_with_problem2(project_root, summary):
    path = project_root / "outputs" / "问题二计算结果" / "问题二全部场景汇总.csv"
    if not path.exists():
        raise FileNotFoundError(path)
    discrete_source = pd.read_csv(path)
    annual_comparison = annual_green_comparison(discrete_source, summary)
    columns = [
        "场景",
        "日氨产量(t/day)",
        "日净成本(¥/day)",
        "净吨氨成本(¥/t)",
        "购电量(MWh)",
        "售电量(MWh)",
        "新能源自发自用比例(%)",
        "新能源供电占比(%)",
        "上网电量比例(%)",
        "达标指标数",
        "达标状态",
    ]
    discrete = discrete_source[columns].rename(columns={column: f"问题二{column}" for column in columns[2:]})
    continuous = summary[columns].rename(columns={column: f"问题三{column}" for column in columns[2:]})
    comparison = discrete.merge(continuous, on=["场景", "日氨产量(t/day)"], how="inner")
    if len(comparison) != 120:
        raise ValueError(f"Expected 120 paired cases, found {len(comparison)}")
    comparison["日成本节约(¥/day)"] = comparison["问题二日净成本(¥/day)"] - comparison["问题三日净成本(¥/day)"]
    comparison["相对成本节约率(%)"] = comparison["日成本节约(¥/day)"] / comparison["问题二日净成本(¥/day)"] * 100.0
    comparison["吨氨成本下降(¥/t)"] = comparison["问题二净吨氨成本(¥/t)"] - comparison["问题三净吨氨成本(¥/t)"]
    comparison["购电量变化(MWh)"] = comparison["问题三购电量(MWh)"] - comparison["问题二购电量(MWh)"]
    comparison["售电量变化(MWh)"] = comparison["问题三售电量(MWh)"] - comparison["问题二售电量(MWh)"]
    for metric in ("新能源自发自用比例(%)", "新能源供电占比(%)", "上网电量比例(%)"):
        comparison[f"{metric}变化"] = comparison[f"问题三{metric}"] - comparison[f"问题二{metric}"]
    comparison["达标指标数变化"] = comparison["问题三达标指标数"] - comparison["问题二达标指标数"]

    annual_rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = comparison[comparison["日氨产量(t/day)"] == output_t]
        annual_value = float(subset["日成本节约(¥/day)"].sum() * SCENARIO_DAYS)
        annual_rows.append(
            {
                "日氨产量(t/day)": output_t,
                "年柔性价值(¥/year)": annual_value,
                "单位吨氨柔性价值(¥/t)": annual_value / (360.0 * output_t),
                "平均日成本节约(¥/day)": float(subset["日成本节约(¥/day)"].mean()),
                "中位日成本节约(¥/day)": float(subset["日成本节约(¥/day)"].median()),
                "最小日成本节约(¥/day)": float(subset["日成本节约(¥/day)"].min()),
                "最大日成本节约(¥/day)": float(subset["日成本节约(¥/day)"].max()),
                "成本改善场景数": int((subset["日成本节约(¥/day)"] > 1e-4).sum()),
                "成本相同场景数": int((subset["日成本节约(¥/day)"].abs() <= 1e-4).sum()),
                "平均购电量变化(MWh)": float(subset["购电量变化(MWh)"].mean()),
                "平均售电量变化(MWh)": float(subset["售电量变化(MWh)"].mean()),
            }
        )
    return comparison, pd.DataFrame(annual_rows), annual_comparison


def write_excel(path, sheets):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="5271AE")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 28)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    workbook.save(path)


def save_figure(fig, directory, stem):
    fig.savefig(directory / f"{stem}.png", dpi=300, bbox_inches="tight")
    fig.savefig(directory / f"{stem}.pdf", bbox_inches="tight")
    plt.close(fig)


def _exchange_stair_series(buy, sell):
    buy = np.asarray(buy, dtype=float)
    sell = np.asarray(sell, dtype=float)
    if buy.shape != sell.shape:
        raise ValueError("Buy and sell series must have the same shape")
    if np.any(buy < -1e-9) or np.any(sell < -1e-9):
        raise ValueError("Buy and sell power must be nonnegative")
    if np.any(buy * sell > 1e-8):
        raise ValueError("Buy and sell power must be mutually exclusive")
    edges = np.arange(buy.size + 1, dtype=float)
    return edges, np.maximum(buy, 0.0), -np.maximum(sell, 0.0)


def heatmap(ax, matrix, title, fmt=".1f", cmap=None):
    image = ax.imshow(matrix, aspect="auto", cmap=cmap)
    ax.set_xticks(range(4), [f"光伏{i}" for i in range(1, 5)])
    ax.set_yticks(range(6), [f"风电{i}" for i in range(1, 7)])
    ax.set_title(title)
    threshold = float(np.nanmean(matrix))
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            value = matrix[i, j]
            color = "white" if value > threshold else "black"
            ax.text(j, i, format(value, fmt), ha="center", va="center", fontsize=8, color=color)
    return image


def generate_figures(summary, hourly, annual, comparison, flexibility, best_output, figure_dir):
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    cmap = LinearSegmentedColormap.from_list("soft", [PALETTE["blue"], PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"]])

    best_subset = summary[summary["日氨产量(t/day)"] == best_output].copy()
    median_cost = float(best_subset["净吨氨成本(¥/t)"].median())
    representative = best_subset.iloc[(best_subset["净吨氨成本(¥/t)"] - median_cost).abs().argmin()]["场景"]
    frame = hourly[(hourly["日氨产量(t/day)"] == best_output) & (hourly["场景"] == representative)].sort_values("时段序号(h)")
    hours = np.arange(len(frame), dtype=float) + 0.5
    edges, buy_step, sell_step = _exchange_stair_series(
        frame["购电功率(MW)"].to_numpy(),
        frame["售电功率(MW)"].to_numpy(),
    )
    fig, axes = plt.subplots(2, 1, figsize=(12, 7), sharex=True, gridspec_kw={"height_ratios": [2, 1]})
    axes[0].stackplot(
        hours,
        frame["常规负荷功率(MW)"],
        frame["PEM电解槽功率(MW)"],
        frame["碱性电解槽功率(MW)"],
        frame["合成氨装置功率(MW)"],
        labels=["常规负荷", "PEM电解槽", "碱性电解槽", "合成氨装置"],
        colors=[PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"]],
        alpha=0.78,
    )
    axes[0].plot(hours, frame["新能源总功率(MW)"], color=PALETTE["blue"], linewidth=2.3, label="风光总功率")
    axes[0].set_ylabel("功率 / MW")
    axes[0].legend(ncol=5, loc="upper center", fontsize=9)
    axes[0].grid(axis="y", alpha=0.25)
    axes[1].axhline(0, color="#555555", linewidth=0.8)
    axes[1].stairs(
        buy_step,
        edges,
        baseline=0.0,
        fill=True,
        color=PALETTE["orange"],
        alpha=0.85,
        linewidth=1.4,
        label="购电功率",
    )
    axes[1].stairs(
        sell_step,
        edges,
        baseline=0.0,
        fill=True,
        color=PALETTE["blue"],
        alpha=0.85,
        linewidth=1.4,
        label="售电功率",
    )
    axes[1].set_ylabel("电网交互 / MW")
    axes[1].set_xlabel("时刻 / h")
    axes[1].set_xlim(0.0, float(HOURS))
    axes[1].set_xticks(np.arange(0, HOURS + 1, 2))
    axes[1].legend(ncol=2, loc="upper left")
    axes[1].grid(axis="y", alpha=0.25)
    save_figure(fig, figure_dir, "问题三代表场景连续调度")

    subset = best_subset.sort_values(["风电场景", "光伏场景"])
    fig, axes = plt.subplots(2, 2, figsize=(11, 9))
    specs = [
        ("净吨氨成本(¥/t)", "净吨氨成本 / ¥/t", ".0f"),
        ("购电量(MWh)", "购电量 / MWh", ".1f"),
        ("售电量(MWh)", "售电量 / MWh", ".1f"),
        ("达标指标数", "达标指标数", ".0f"),
    ]
    for ax, (column, title, fmt) in zip(axes.flat, specs):
        matrix = subset.pivot(index="风电场景", columns="光伏场景", values=column).to_numpy()
        image = heatmap(ax, matrix, title, fmt=fmt, cmap=cmap)
        fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04)
    fig.tight_layout()
    save_figure(fig, figure_dir, "问题三风光场景运行热力图")

    fig, ax = plt.subplots(figsize=(10, 5.6))
    colors = [PALETTE["red"], PALETTE["orange"], PALETTE["yellow"], PALETTE["light_blue"], PALETTE["blue"]]
    for color, output_t in zip(colors, DAILY_OUTPUTS_T):
        distribution = annual_tonne_cost_distribution(summary, output_t)
        ax.plot(np.arange(1, 361), distribution, color=color, linewidth=1.8, label=f"{output_t} t/day")
    ax.set_xlabel("持续天数 / day")
    ax.set_ylabel("净吨氨成本 / ¥/t")
    ax.set_xlim(1, 360)
    ax.grid(alpha=0.25)
    ax.legend(ncol=5, loc="upper center")
    save_figure(fig, figure_dir, "问题三全年吨氨成本分布曲线")
    for suffix in ("png", "pdf"):
        stale = figure_dir / f"问题三全年成本持续曲线.{suffix}"
        if stale.exists():
            stale.unlink()

    fig, axes = plt.subplots(1, 2, figsize=(12, 4.8))
    flex_sorted = flexibility.sort_values("日氨产量(t/day)")
    axes[0].bar(
        flex_sorted["日氨产量(t/day)"].astype(str),
        flex_sorted["年柔性价值(¥/year)"],
        color=[PALETTE["blue"], PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"]],
    )
    axes[0].set_xlabel("日氨产量 / t/day")
    axes[0].set_ylabel("年柔性价值 / ¥")
    axes[0].grid(axis="y", alpha=0.25)
    delta = comparison[comparison["日氨产量(t/day)"] == best_output].copy()
    delta[["风电场景", "光伏场景"]] = delta["场景"].str.extract(r"风电(\d+)-光伏(\d+)").astype(int)
    matrix = delta.pivot(index="风电场景", columns="光伏场景", values="吨氨成本下降(¥/t)").to_numpy()
    image = heatmap(axes[1], matrix, "吨氨成本下降 / ¥/t", fmt=".1f", cmap=cmap)
    fig.colorbar(image, ax=axes[1], fraction=0.046, pad=0.04)
    fig.tight_layout()
    save_figure(fig, figure_dir, "问题三连续调节柔性价值")


def dataframe_records(frame):
    return json.loads(frame.to_json(orient="records", force_ascii=False))


def write_report(
    path,
    annual,
    summary,
    validation,
    correlation,
    marginal,
    comparison,
    flexibility,
    operating_overview,
    representative_scenarios,
    annual_comparison,
    best_output,
):
    best = annual.loc[annual["日氨产量(t/day)"] == best_output].iloc[0]
    max_validation = validation.select_dtypes(include=[np.number]).drop(columns=["日氨产量(t/day)"], errors="ignore").max()
    best_cases = summary[summary["日氨产量(t/day)"] == best_output]
    lowest = best_cases.loc[best_cases["净吨氨成本(¥/t)"].idxmin()]
    highest = best_cases.loc[best_cases["净吨氨成本(¥/t)"].idxmax()]
    flex_best = flexibility.loc[flexibility["日氨产量(t/day)"] == best_output].iloc[0]
    corr_best = correlation[correlation["日氨产量(t/day)"] == best_output].copy()
    corr_best["绝对相关系数"] = corr_best["Spearman描述性相关系数"].abs()
    strongest = corr_best.sort_values("绝对相关系数", ascending=False).head(6).drop(columns="绝对相关系数")
    negative_count = int((comparison["日成本节约(¥/day)"] < -0.05).sum())
    representative_report = representative_scenarios[
        [
            "代表类型",
            "场景",
            "新能源总发电量(MWh)",
            "高电价时段新能源占比(%)",
            "原始缺电面积(MWh)",
            "购电量(MWh)",
            "售电量(MWh)",
            "购电成本(¥/day)",
            "售电收入(¥/day)",
            "净吨氨成本(¥/t)",
            "新能源自发自用比例(%)",
            "新能源供电占比(%)",
            "上网电量比例(%)",
            "达标状态",
        ]
    ]
    annual_metric_change = annual_comparison[
        [
            "日氨产量(t/day)",
            "问题二新能源自发自用比例(%)",
            "问题三新能源自发自用比例(%)",
            "新能源自发自用比例变化(百分点)",
            "问题二新能源供电占比(%)",
            "问题三新能源供电占比(%)",
            "新能源供电占比变化(百分点)",
            "问题二上网电量比例(%)",
            "问题三上网电量比例(%)",
            "上网电量比例变化(百分点)",
        ]
    ]
    annual_status_change = annual_comparison[
        [
            "日氨产量(t/day)",
            "问题二全部达标天数(day)",
            "问题三全部达标天数(day)",
            "问题二部分达标天数(day)",
            "问题三部分达标天数(day)",
            "问题二全部不达标天数(day)",
            "问题三全部不达标天数(day)",
        ]
    ]
    representative_low = representative_scenarios[representative_scenarios["代表类型"] == "最低成本"].iloc[0]
    representative_high = representative_scenarios[representative_scenarios["代表类型"] == "最高成本"].iloc[0]
    text = f"""# 问题三计算结果

## 方法与口径

采用半连续混合整数线性规划。ALK、PEM和合成氨装置分别允许停机或在额定功率的10%至100%之间连续调节，不设置电储能和氢储能。每组场景依次执行“净成本最小、购电量最小、售电量最小”的词典序求解。风光成本按实际发电量核算，合成氨投资折旧单独计入固定成本，设备运维按实际功率核算。

## 问题三（1）年度结果

{annual.to_markdown(index=False, floatfmt=".3f")}

在给定五档候选产量、且仅以综合净吨氨成本最低为标准时，最低成本对应日产量为 **{best_output:.0f} t/day**，综合净吨氨成本为 **{best['综合净吨氨成本(¥/t)']:.2f} ¥/t**，年综合净成本为 **{best['年综合净成本(¥/year)']:.2f} ¥**。该结论不是兼顾产能规模与政策指标后的综合产能推荐。

最低成本对应产量下，最低吨氨成本场景为 **{lowest['场景']}**，成本 **{lowest['净吨氨成本(¥/t)']:.2f} ¥/t**；最高成本场景为 **{highest['场景']}**，成本 **{highest['净吨氨成本(¥/t)']:.2f} ¥/t**。

![全年吨氨成本分布曲线](../figures/问题三计算结果/问题三全年吨氨成本分布曲线.png)

## 问题三（2）场景机理分析

### 五档产量运行概览

{operating_overview.to_markdown(index=False, floatfmt=".3f")}

五档产量的24种场景均已纳入统计。产量下降时，连续调节空间扩大，生产负荷可以更多转移到风光充足或电价较低的时段；与此同时，较低生产负荷也可能扩大余电上网，因此三项绿电指标必须结合购电量、售电量和新能源总量共同判断。

### 最低成本对应产量的代表场景

{representative_report.to_markdown(index=False, floatfmt=".3f")}

在 {best_output:.0f} t/day 下，最低成本场景 **{representative_low['场景']}** 的购电量为 **{representative_low['购电量(MWh)']:.2f} MWh**、售电量为 **{representative_low['售电量(MWh)']:.2f} MWh**；最高成本场景 **{representative_high['场景']}** 的购电量为 **{representative_high['购电量(MWh)']:.2f} MWh**、售电量为 **{representative_high['售电量(MWh)']:.2f} MWh**。场景成本差异主要由新能源总量、风光在高电价时段的分布以及原始缺电面积共同形成：新能源总量决定总体供需缺口，高电价时段新能源量决定昂贵购电的替代程度，原始缺电面积刻画风光与常规负荷的时序错配。售电收入会降低净成本，但余电过多也会抬高上网电量比例并压低新能源自发自用比例。

![最低成本对应产量代表场景连续调度](../figures/问题三计算结果/问题三代表场景连续调度.png)

![最低成本对应产量风光场景运行热力图](../figures/问题三计算结果/问题三风光场景运行热力图.png)

### 描述性关联

使用新能源总发电量、高电价时段新能源占比和原始缺电面积三个外生特征。以下相关系数固定在最低成本对应产量 {best_output:.0f} t/day，仅用于描述性关联，不解释为因果贡献；全部五档产量的完整相关结果保存在结果工作簿中。

### 最强描述性关联

{strongest.to_markdown(index=False, floatfmt=".4f")}

### 风光边际均值

{marginal.to_markdown(index=False, floatfmt=".3f")}

## 问题三（3）相对问题二的柔性价值

{flexibility.to_markdown(index=False, floatfmt=".3f")}

最低成本对应产量下年柔性价值为 **{flex_best['年柔性价值(¥/year)']:.2f} ¥/year**，单位吨氨柔性价值为 **{flex_best['单位吨氨柔性价值(¥/t)']:.2f} ¥/t**。连续模型成本高于离散模型超过0.05 ¥/day的配对数为 **{negative_count}**。

### 全年绿电指标变化

{annual_metric_change.to_markdown(index=False, floatfmt=".3f")}

### 全年达标状态变化

{annual_status_change.to_markdown(index=False, floatfmt=".0f")}

72 t/day 下三套设备需要全天满负荷运行，连续调节可行域退化为问题二的全开状态，因此成本和绿电指标不变。其余产量下，连续模型通过降低全开全停的粒度损失、优先使用制氢效率较高的PEM并由ALK补充、把生产负荷迁移到风光充足或电价更合适的时段，平均减少购电和售电，提高新能源自发自用比例与新能源供电占比，并降低上网电量比例。购售电量变化不预设方向，以120组配对结果为准。

![连续调节柔性价值](../figures/问题三计算结果/问题三连续调节柔性价值.png)

## 约束与一致性校验

{max_validation.to_frame("120组最大值").reset_index(names="校验项").to_markdown(index=False, floatfmt=".3e")}

- 场景—产量组合数：{len(summary)}。
- 逐时结果行数：{len(summary) * HOURS}。
- 每档产量的全年状态天数均为360 day。
- 连续模型可行域包含问题二离散模型；若出现显著负节约，应检查成本或约束口径。

## 文件说明

- `outputs/问题三计算结果/问题三计算结果.xlsx`：全部结果工作簿。
- `outputs/问题三计算结果/问题三全部场景汇总.csv`：120组汇总结果。
- `outputs/问题三计算结果/问题三全部场景逐时调度.csv`：2880行逐时结果。
- `outputs/问题三计算结果/问题三五档产量运行概览.csv`：五档产量的购售电、成本和全年指标概览。
- `outputs/问题三计算结果/问题三代表场景机理.csv`：最低、中位和最高成本代表场景。
- `outputs/问题三计算结果/问题二与问题三比较.csv`：120组配对结果。
- `outputs/问题三计算结果/问题二三全年指标变化.csv`：问题二、三全年绿电指标和达标状态变化。
- `figures/问题三计算结果/`：四组论文可用PNG和PDF图形，报告正文均已引用。

## 可复现运行方式

```powershell
python -X utf8 code/问题三.py --self-test
python -X utf8 code/问题三.py
```
"""
    path.write_text(text, encoding="utf-8")


def solve_toy_case():
    conventional = np.full(HOURS, 3.0)
    generation = np.concatenate([np.zeros(12), np.full(12, 30.0)])
    prices = np.concatenate([np.full(12, 0.34), np.full(12, 0.80)])
    om_costs = {"alk": 0.10, "pem": 0.15, "ammonia": 0.02}
    result, slices, primary = solve_core_case(conventional, generation, prices, 0.38, 36.0, om_costs)
    x = result.x
    alk = x[slices["alk"]]
    pem = x[slices["pem"]]
    ammonia = x[slices["ammonia"]]
    buy = x[slices["buy"]]
    sell = x[slices["sell"]]
    load = conventional + alk + pem + ammonia
    discrete_u = np.concatenate([np.ones(12), np.zeros(12)])
    discrete_load = conventional + 41.5 * discrete_u
    discrete_buy = np.maximum(discrete_load - generation, 0.0)
    discrete_sell = np.maximum(generation - discrete_load, 0.0)
    discrete_cost = float(
        1000.0
        * np.sum(
            prices * discrete_buy
            - 0.38 * discrete_sell
            + om_costs["alk"] * 20.0 * discrete_u
            + om_costs["pem"] * 20.0 * discrete_u
            + om_costs["ammonia"] * 1.5 * discrete_u
        )
    )
    return {
        "success": bool(result.success),
        "max_material_residual": float(np.max(np.abs(material_residual(alk, pem, ammonia)))),
        "max_power_residual": float(np.max(np.abs(generation + buy - load - sell))),
        "continuous_cost": float(primary @ x),
        "discrete_cost": discrete_cost,
    }


def run_self_tests():
    edges, buy_step, sell_step = _exchange_stair_series(
        np.array([5.0, 0.0, 3.0]),
        np.array([0.0, 4.0, 0.0]),
    )
    assert np.array_equal(edges, np.array([0.0, 1.0, 2.0, 3.0]))
    assert np.array_equal(buy_step, np.array([5.0, 0.0, 3.0]))
    assert np.array_equal(sell_step, np.array([0.0, -4.0, 0.0]))

    metrics = calculate_green_metrics(100.0, 80.0, 30.0, 10.0)
    assert math.isclose(metrics["r1"], 0.75, abs_tol=1e-12)
    assert math.isclose(metrics["r2"], 0.70, abs_tol=1e-12)
    assert math.isclose(metrics["r3"], 0.125, abs_tol=1e-12)
    assert is_semicontinuous(0.0, 20.0)
    assert is_semicontinuous(2.0, 20.0)
    assert is_semicontinuous(20.0, 20.0)
    assert not is_semicontinuous(1.0, 20.0)
    assert not is_semicontinuous(20.1, 20.0)
    assert math.isclose(material_residual(20.0, 20.0, 1.5), 0.0, abs_tol=1e-12)
    assert math.isclose(required_ammonia_power_energy(72.0), 36.0, abs_tol=1e-12)
    toy = solve_toy_case()
    assert toy["success"]
    assert toy["max_material_residual"] < 1e-7
    assert toy["max_power_residual"] < 1e-7
    assert toy["continuous_cost"] <= toy["discrete_cost"] + 1e-6

    duration_builder = globals().get("annual_tonne_cost_distribution")
    assert duration_builder is not None, "Missing annual tonne-cost distribution builder"
    duration_sample = pd.DataFrame(
        {
            "日氨产量(t/day)": [36.0, 36.0],
            "净吨氨成本(¥/t)": [100.0, 200.0],
        }
    )
    duration_values = duration_builder(duration_sample, 36.0, scenario_days=2)
    assert np.array_equal(duration_values, np.array([200.0, 200.0, 100.0, 100.0]))

    overview_builder = globals().get("scenario_operating_summary")
    assert overview_builder is not None, "Missing five-output operating overview builder"
    overview_sample = pd.DataFrame(
        {
            "日氨产量(t/day)": [36.0, 36.0],
            "购电量(MWh)": [10.0, 30.0],
            "售电量(MWh)": [4.0, 8.0],
            "净吨氨成本(¥/t)": [100.0, 200.0],
        }
    )
    annual_sample = pd.DataFrame(
        {
            "日氨产量(t/day)": [36.0],
            "全部达标天数(day)": [15],
            "部分达标天数(day)": [15],
            "全部不达标天数(day)": [0],
            "全年新能源自发自用比例(%)": [70.0],
            "全年新能源供电占比(%)": [40.0],
            "全年上网电量比例(%)": [10.0],
        }
    )
    overview = overview_builder(overview_sample, annual_sample)
    assert math.isclose(overview.iloc[0]["平均购电量(MWh/day)"], 20.0, abs_tol=1e-12)
    assert math.isclose(overview.iloc[0]["吨氨成本最大值(¥/t)"], 200.0, abs_tol=1e-12)

    annual_comparison_builder = globals().get("annual_green_comparison")
    assert annual_comparison_builder is not None, "Missing annual green-metric comparison builder"
    discrete_sample = pd.DataFrame(
        {
            "日氨产量(t/day)": [36.0],
            "总负荷电量(MWh)": [100.0],
            "新能源发电量(MWh)": [80.0],
            "购电量(MWh)": [30.0],
            "售电量(MWh)": [10.0],
            "达标状态": ["全部达标"],
        }
    )
    continuous_sample = discrete_sample.copy()
    continuous_sample["购电量(MWh)"] = 20.0
    continuous_sample["售电量(MWh)"] = 5.0
    annual_comparison = annual_comparison_builder(discrete_sample, continuous_sample, scenario_days=15)
    assert math.isclose(annual_comparison.iloc[0]["新能源自发自用比例变化(百分点)"], 18.75, abs_tol=1e-12)
    assert math.isclose(annual_comparison.iloc[0]["新能源供电占比变化(百分点)"], 5.0, abs_tol=1e-12)
    assert math.isclose(annual_comparison.iloc[0]["上网电量比例变化(百分点)"], -6.25, abs_tol=1e-12)
    print("All problem 3 self-tests passed.")


def run_full_model(project_root):
    data = load_inputs(project_root)
    output_dir = project_root / "outputs" / "问题三计算结果"
    figure_dir = project_root / "figures" / "问题三计算结果"
    report_path = project_root / "reports" / "问题三计算结果.md"
    output_dir.mkdir(parents=True, exist_ok=True)
    figure_dir.mkdir(parents=True, exist_ok=True)

    summaries = []
    hourly_rows = []
    validations = []
    total_cases = 24 * len(DAILY_OUTPUTS_T)
    completed = 0
    for wi, wind in enumerate(data["wind_scenarios"], start=1):
        for pi, pv in enumerate(data["pv_scenarios"], start=1):
            scenario_name = f"风电{wi}-光伏{pi}"
            for output_t in DAILY_OUTPUTS_T:
                summary, hourly, validation = solve_case(
                    data, wind, pv, output_t, scenario_name, wi, pi
                )
                summaries.append(summary)
                hourly_rows.extend(hourly)
                validations.append(validation)
                completed += 1
                if completed % 10 == 0 or completed == total_cases:
                    print(f"Solved {completed}/{total_cases} cases", flush=True)

    summary = pd.DataFrame(summaries)
    hourly = pd.DataFrame(hourly_rows)
    validation = pd.DataFrame(validations)
    annual = annual_summary(summary)
    best_output = float(annual.loc[annual["综合净吨氨成本(¥/t)"].idxmin(), "日氨产量(t/day)"])
    features = scenario_features(data)
    merged, correlation, marginal = descriptive_analysis(summary, features, best_output)
    operating_overview = scenario_operating_summary(summary, annual)
    representative_scenarios = representative_scenario_analysis(merged, best_output)
    comparison, flexibility, annual_comparison = compare_with_problem2(project_root, summary)

    if len(summary) != 120 or len(hourly) != 2880:
        raise AssertionError("Unexpected result dimensions")
    if not np.all(annual[["全部达标天数(day)", "部分达标天数(day)", "全部不达标天数(day)"]].sum(axis=1) == 360):
        raise AssertionError("Annual status days do not sum to 360")
    if validation["最大功率平衡残差(MW)"].max() > 1e-6:
        raise AssertionError("Power balance validation failed")
    if validation["最大氢氨物料平衡残差(kg/h)"].max() > 1e-6:
        raise AssertionError("Material balance validation failed")
    if comparison["日成本节约(¥/day)"].min() < -0.05:
        raise AssertionError("Continuous cost exceeds discrete cost")

    frames = {
        "场景汇总": summary,
        "逐时调度": hourly,
        "全年汇总": annual,
        "约束校验": validation,
        "场景特征": features,
        "描述性相关": correlation,
        "风光边际均值": marginal,
        "五档产量运行概览": operating_overview,
        "代表场景机理": representative_scenarios,
        "问题二三比较": comparison,
        "全年指标变化": annual_comparison,
        "柔性价值": flexibility,
    }
    write_excel(output_dir / "问题三计算结果.xlsx", frames)
    summary.to_csv(output_dir / "问题三全部场景汇总.csv", index=False, encoding="utf-8-sig")
    hourly.to_csv(output_dir / "问题三全部场景逐时调度.csv", index=False, encoding="utf-8-sig")
    annual.to_csv(output_dir / "问题三年度汇总.csv", index=False, encoding="utf-8-sig")
    validation.to_csv(output_dir / "问题三约束校验.csv", index=False, encoding="utf-8-sig")
    features.to_csv(output_dir / "问题三场景特征.csv", index=False, encoding="utf-8-sig")
    correlation.to_csv(output_dir / "问题三描述性相关.csv", index=False, encoding="utf-8-sig")
    marginal.to_csv(output_dir / "问题三风光边际均值.csv", index=False, encoding="utf-8-sig")
    operating_overview.to_csv(output_dir / "问题三五档产量运行概览.csv", index=False, encoding="utf-8-sig")
    representative_scenarios.to_csv(output_dir / "问题三代表场景机理.csv", index=False, encoding="utf-8-sig")
    comparison.to_csv(output_dir / "问题二与问题三比较.csv", index=False, encoding="utf-8-sig")
    annual_comparison.to_csv(output_dir / "问题二三全年指标变化.csv", index=False, encoding="utf-8-sig")
    flexibility.to_csv(output_dir / "问题三柔性价值.csv", index=False, encoding="utf-8-sig")
    payload = {
        "候选产量中最低吨氨成本对应日产量": best_output,
        "年度汇总": dataframe_records(annual),
        "五档产量运行概览": dataframe_records(operating_overview),
        "问题二三全年指标变化": dataframe_records(annual_comparison),
        "柔性价值": dataframe_records(flexibility),
        "最大约束残差": dataframe_records(validation.max(numeric_only=True).to_frame().T),
    }
    (output_dir / "问题三完整计算结果.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    generate_figures(summary, hourly, annual, comparison, flexibility, best_output, figure_dir)
    write_report(
        report_path,
        annual,
        summary,
        validation,
        correlation,
        marginal,
        comparison,
        flexibility,
        operating_overview,
        representative_scenarios,
        annual_comparison,
        best_output,
    )
    print(f"Best output: {best_output:.0f} t/day")
    print(f"Minimum paired saving: {comparison['日成本节约(¥/day)'].min():.6f} ¥/day")
    print(f"Maximum power residual: {validation['最大功率平衡残差(MW)'].max():.3e} MW")
    print(f"Maximum material residual: {validation['最大氢氨物料平衡残差(kg/h)'].max():.3e} kg/h")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        run_self_tests()
        return
    project_root = Path(__file__).resolve().parents[1]
    run_full_model(project_root)


if __name__ == "__main__":
    main()
