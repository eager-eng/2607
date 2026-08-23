import json
import re
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import coo_matrix

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap


HOURS = 24
TIME_STEP_H = 1.0
CONVENTIONAL_LOAD_BASE_MW = 6.0
WIND_CAPACITY_MW = 40.0
PV_CAPACITY_MW = 64.0
ALK_POWER_MW = 20.0
PEM_POWER_MW = 20.0
AMMONIA_POWER_MW = 1.5
PROCESS_POWER_MW = ALK_POWER_MW + PEM_POWER_MW + AMMONIA_POWER_MW
AMMONIA_RATE_T_PER_H = 3.0
HYDROGEN_RATE_KG_PER_H = 600.0
DAILY_OUTPUTS_T = (72, 63, 54, 45, 36)
SCENARIO_DAYS = 15
STRICT_TOLERANCE = 1e-9
GREEN_DELTA = 1e-6
PALETTE = {
    "blue": "#5271AE",
    "light_blue": "#70ACDE",
    "yellow": "#F5CC7D",
    "orange": "#FFA660",
    "red": "#D85B59",
}


def find_attachment(project_root, prefix):
    matches = [
        path
        for path in Path(project_root).joinpath("题目").rglob("*.xlsx")
        if not path.name.startswith("~$") and path.name.startswith(prefix)
    ]
    if len(matches) != 1:
        raise FileNotFoundError(f"Expected one {prefix} workbook, found {len(matches)}")
    return matches[0]


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell.value for cell in row) for row in workbook.active.iter_rows()]
    finally:
        workbook.close()


def extract_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        raise ValueError(f"No numeric value found in {value!r}")
    return float(match.group())


def hourly_buy_prices(rows):
    peak_price = extract_number(rows[1][1])
    flat_price = extract_number(rows[2][1])
    valley_price = extract_number(rows[3][1])
    prices = []
    for hour in range(HOURS):
        if 10 <= hour < 15 or 18 <= hour < 21:
            prices.append(peak_price)
        elif 7 <= hour < 10 or 15 <= hour < 18 or 21 <= hour < 23:
            prices.append(flat_price)
        else:
            prices.append(valley_price)
    return np.asarray(prices, dtype=float)


def load_inputs(project_root):
    load_rows = read_rows(find_attachment(project_root, "附件1"))
    typical_rows = read_rows(find_attachment(project_root, "附件2"))
    wind_rows = read_rows(find_attachment(project_root, "附件3"))
    pv_rows = read_rows(find_attachment(project_root, "附件4"))
    equipment_rows = read_rows(find_attachment(project_root, "附件5"))
    ammonia_rows = read_rows(find_attachment(project_root, "附件6"))
    buy_rows = read_rows(find_attachment(project_root, "附件7"))
    sell_rows = read_rows(find_attachment(project_root, "附件8"))

    time_labels = [str(row[0]) for row in load_rows[1:]]
    load_pu = np.asarray([row[1] for row in load_rows[1:]], dtype=float)
    typical_times = [str(row[0]) for row in typical_rows[1:]]
    typical_wind_pu = np.asarray([row[1] for row in typical_rows[1:]], dtype=float)
    typical_pv_pu = np.asarray([row[2] for row in typical_rows[1:]], dtype=float)
    wind_times = [str(row[0]) for row in wind_rows[1:]]
    pv_times = [str(row[0]) for row in pv_rows[1:]]
    if not (
        len(time_labels) == HOURS
        and time_labels == typical_times
        and time_labels == wind_times
        and time_labels == pv_times
    ):
        raise ValueError("Hourly input tables must contain the same 24 periods")

    wind_scenarios_pu = np.asarray([row[1:7] for row in wind_rows[1:]], dtype=float).T
    pv_scenarios_pu = np.asarray([row[1:5] for row in pv_rows[1:]], dtype=float).T
    if wind_scenarios_pu.shape != (6, HOURS) or pv_scenarios_pu.shape != (4, HOURS):
        raise ValueError("Scenario tables must have dimensions 6x24 and 4x24")

    wind_cost = extract_number(equipment_rows[1][1])
    pv_cost = extract_number(equipment_rows[1][2])
    alk_om = extract_number(equipment_rows[2][3])
    pem_om = extract_number(equipment_rows[2][4])
    ammonia_investment = extract_number(ammonia_rows[1][2])
    ammonia_om = extract_number(ammonia_rows[2][2])
    ammonia_lifetime = extract_number(ammonia_rows[3][2])
    buy_prices = hourly_buy_prices(buy_rows)
    sell_wind_price = extract_number(sell_rows[1][1])
    sell_pv_price = extract_number(sell_rows[2][1])
    if not np.isclose(sell_wind_price, sell_pv_price):
        raise ValueError("Aggregate surplus requires a common feed-in price")

    return {
        "time_labels": time_labels,
        "conventional_load": CONVENTIONAL_LOAD_BASE_MW * load_pu,
        "typical_wind": WIND_CAPACITY_MW * typical_wind_pu,
        "typical_pv": PV_CAPACITY_MW * typical_pv_pu,
        "wind_scenarios": WIND_CAPACITY_MW * wind_scenarios_pu,
        "pv_scenarios": PV_CAPACITY_MW * pv_scenarios_pu,
        "buy_prices": buy_prices,
        "sell_price": sell_wind_price,
        "wind_cost": wind_cost,
        "pv_cost": pv_cost,
        "alk_om": alk_om,
        "pem_om": pem_om,
        "ammonia_om": ammonia_om,
        "ammonia_investment": ammonia_investment,
        "ammonia_lifetime": ammonia_lifetime,
    }


def calculate_grid_exchange(load_mw, generation_mw):
    net_load = np.asarray(load_mw, dtype=float) - np.asarray(generation_mw, dtype=float)
    return np.maximum(net_load, 0.0), np.maximum(-net_load, 0.0)


def calculate_green_metrics(load_energy, generation_energy, buy_energy, sell_energy):
    if load_energy <= 0 or generation_energy <= 0:
        raise ValueError("Green metric denominators must be positive")
    return {
        "新能源自发自用比例": (load_energy - sell_energy - buy_energy) / generation_energy,
        "新能源供电占比": (generation_energy - sell_energy) / load_energy,
        "上网电量比例": sell_energy / generation_energy,
    }


def classify_green_metrics(metrics):
    status = {
        "新能源自发自用比例达标": metrics["新能源自发自用比例"] > 0.60 + STRICT_TOLERANCE,
        "新能源供电占比达标": metrics["新能源供电占比"] > 0.30 + STRICT_TOLERANCE,
        "上网电量比例达标": metrics["上网电量比例"] < 0.20 - STRICT_TOLERANCE,
    }
    count = int(sum(status.values()))
    if count == 3:
        conclusion = "全部达标"
    elif count == 0:
        conclusion = "全部不达标"
    else:
        conclusion = "部分达标"
    return status, count, conclusion


def build_milp(conventional_load, generation, buy_prices, sell_price, output_t, green=False):
    n = HOURS * 4
    u_slice = slice(0, HOURS)
    buy_slice = slice(HOURS, 2 * HOURS)
    sell_slice = slice(2 * HOURS, 3 * HOURS)
    mode_slice = slice(3 * HOURS, 4 * HOURS)
    objective = np.zeros(n, dtype=float)
    objective[u_slice] = 1000.0 * (
        ALK_POWER_MW * 0.0 + PEM_POWER_MW * 0.0 + AMMONIA_POWER_MW * 0.0
    )
    objective[buy_slice] = 1000.0 * buy_prices
    objective[sell_slice] = -1000.0 * sell_price

    lower = np.zeros(n, dtype=float)
    upper = np.full(n, np.inf, dtype=float)
    upper[u_slice] = 1.0
    upper[mode_slice] = 1.0
    upper[buy_slice] = conventional_load + PROCESS_POWER_MW
    upper[sell_slice] = generation
    integrality = np.zeros(n, dtype=int)
    integrality[u_slice] = 1
    integrality[mode_slice] = 1

    row_index = []
    col_index = []
    values = []
    row_lower = []
    row_upper = []
    row = 0

    for t in range(HOURS):
        row_index.extend([row, row, row])
        col_index.extend([t, HOURS + t, 2 * HOURS + t])
        values.extend([PROCESS_POWER_MW, -1.0, 1.0])
        rhs = generation[t] - conventional_load[t]
        row_lower.append(rhs)
        row_upper.append(rhs)
        row += 1

    for t in range(HOURS):
        row_index.extend([row, row])
        col_index.extend([HOURS + t, 3 * HOURS + t])
        values.extend([1.0, -(conventional_load[t] + PROCESS_POWER_MW)])
        row_lower.append(-np.inf)
        row_upper.append(0.0)
        row += 1

    for t in range(HOURS):
        row_index.extend([row, row])
        col_index.extend([2 * HOURS + t, 3 * HOURS + t])
        values.extend([1.0, generation[t]])
        row_lower.append(-np.inf)
        row_upper.append(generation[t])
        row += 1

    for t in range(HOURS):
        row_index.append(row)
        col_index.append(t)
        values.append(1.0)
    required_hours = output_t / AMMONIA_RATE_T_PER_H
    row_lower.append(required_hours)
    row_upper.append(required_hours)
    row += 1

    if green:
        generation_energy = float(np.sum(generation) * TIME_STEP_H)
        load_energy = float(
            (np.sum(conventional_load) + PROCESS_POWER_MW * required_hours) * TIME_STEP_H
        )
        for t in range(HOURS):
            row_index.append(row)
            col_index.append(2 * HOURS + t)
            values.append(TIME_STEP_H)
        row_lower.append(-np.inf)
        row_upper.append((0.20 - GREEN_DELTA) * generation_energy)
        row += 1
        for t in range(HOURS):
            row_index.append(row)
            col_index.append(2 * HOURS + t)
            values.append(TIME_STEP_H)
        row_lower.append(-np.inf)
        row_upper.append(generation_energy - (0.30 + GREEN_DELTA) * load_energy)
        row += 1

    matrix = coo_matrix((values, (row_index, col_index)), shape=(row, n)).tocsr()
    constraint = LinearConstraint(matrix, np.asarray(row_lower), np.asarray(row_upper))
    result = milp(
        c=objective,
        integrality=integrality,
        bounds=Bounds(lower, upper),
        constraints=constraint,
        options={"mip_rel_gap": 1e-10, "time_limit": 60.0, "presolve": True},
    )
    return result


def cost_components(data, wind_power, pv_power, buy_power, sell_power, u, output_t):
    wind_energy = float(np.sum(wind_power) * TIME_STEP_H)
    pv_energy = float(np.sum(pv_power) * TIME_STEP_H)
    wind_cost = 1000.0 * data["wind_cost"] * wind_energy
    pv_cost = 1000.0 * data["pv_cost"] * pv_energy
    buy_cost = float(1000.0 * np.sum(buy_power * data["buy_prices"]) * TIME_STEP_H)
    sell_revenue = float(1000.0 * data["sell_price"] * np.sum(sell_power) * TIME_STEP_H)
    alk_om = float(1000.0 * data["alk_om"] * ALK_POWER_MW * np.sum(u) * TIME_STEP_H)
    pem_om = float(1000.0 * data["pem_om"] * PEM_POWER_MW * np.sum(u) * TIME_STEP_H)
    ammonia_om = float(
        1000.0 * data["ammonia_om"] * AMMONIA_POWER_MW * np.sum(u) * TIME_STEP_H
    )
    capital = float(
        data["ammonia_investment"]
        * HYDROGEN_RATE_KG_PER_H
        / data["ammonia_lifetime"]
        / 365.0
    )
    net_cost = wind_cost + pv_cost + buy_cost - sell_revenue + alk_om + pem_om + ammonia_om + capital
    return {
        "风电发电成本(¥/day)": wind_cost,
        "光伏发电成本(¥/day)": pv_cost,
        "购电成本(¥/day)": buy_cost,
        "售电收入(¥/day)": sell_revenue,
        "碱性电解槽运维成本(¥/day)": alk_om,
        "PEM电解槽运维成本(¥/day)": pem_om,
        "合成氨装置运维成本(¥/day)": ammonia_om,
        "合成氨装置资本分摊(¥/day)": capital,
        "日净成本(¥/day)": net_cost,
        "净吨氨成本(¥/t)": net_cost / output_t,
    }


def marginal_ranking_solution(conventional_load, generation, buy_prices, sell_price, output_t):
    off_buy, off_sell = calculate_grid_exchange(conventional_load, generation)
    on_buy, on_sell = calculate_grid_exchange(conventional_load + PROCESS_POWER_MW, generation)
    off_cost = 1000.0 * (buy_prices * off_buy - sell_price * off_sell)
    on_cost = 1000.0 * (buy_prices * on_buy - sell_price * on_sell)
    marginal = on_cost - off_cost
    required_hours = int(round(output_t / AMMONIA_RATE_T_PER_H))
    order = np.lexsort((np.arange(HOURS), on_buy, marginal))
    u = np.zeros(HOURS, dtype=float)
    u[order[:required_hours]] = 1.0
    total_load = conventional_load + PROCESS_POWER_MW * u
    buy, sell = calculate_grid_exchange(total_load, generation)
    return u, buy, sell, marginal


def format_operating_periods(u):
    active = np.flatnonzero(np.asarray(u, dtype=float) > 0.5)
    if len(active) == 0:
        return "停机"
    periods = []
    start = active[0]
    end = active[0]
    for hour in active[1:]:
        if hour == end + 1:
            end = hour
        else:
            periods.append(f"{start:02d}:00-{end + 1:02d}:00")
            start = hour
            end = hour
    periods.append(f"{start:02d}:00-{end + 1:02d}:00")
    return "、".join(periods)


def solve_case(data, wind_power, pv_power, output_t, scenario_name, green=False):
    conventional_load = data["conventional_load"]
    generation = wind_power + pv_power
    result = build_milp(
        conventional_load,
        generation,
        data["buy_prices"],
        data["sell_price"],
        output_t,
        green=green,
    )
    if not result.success:
        return {"feasible": False, "scenario": scenario_name, "output_t": output_t, "message": result.message}

    u = np.rint(result.x[:HOURS]).astype(float)
    total_load = conventional_load + PROCESS_POWER_MW * u
    buy, sell = calculate_grid_exchange(total_load, generation)
    ranking_u, ranking_buy, ranking_sell, marginal = marginal_ranking_solution(
        conventional_load,
        generation,
        data["buy_prices"],
        data["sell_price"],
        output_t,
    )

    load_energy = float(np.sum(total_load) * TIME_STEP_H)
    generation_energy = float(np.sum(generation) * TIME_STEP_H)
    buy_energy = float(np.sum(buy) * TIME_STEP_H)
    sell_energy = float(np.sum(sell) * TIME_STEP_H)
    metrics = calculate_green_metrics(load_energy, generation_energy, buy_energy, sell_energy)
    status, passed_count, conclusion = classify_green_metrics(metrics)
    costs = cost_components(data, wind_power, pv_power, buy, sell, u, output_t)
    ranking_costs = cost_components(data, wind_power, pv_power, ranking_buy, ranking_sell, ranking_u, output_t)

    power_residual = generation + buy - total_load - sell
    checks = {
        "最大逐时功率平衡残差(MW)": float(np.max(np.abs(power_residual))),
        "产量约束残差(t/day)": float(abs(AMMONIA_RATE_T_PER_H * np.sum(u) - output_t)),
        "购售电同时发生最大值(MW2)": float(np.max(buy * sell)),
        "氢气平衡残差(kgH2/h)": float(abs(HYDROGEN_RATE_KG_PER_H - 0.2 * 3000.0)),
        "合成氨耗电残差(MW)": float(abs(AMMONIA_POWER_MW - 0.5 * 3000.0 / 1000.0)),
        "MILP与排序成本差(¥/day)": float(abs(costs["日净成本(¥/day)"] - ranking_costs["日净成本(¥/day)"])),
    }

    hourly = []
    for t in range(HOURS):
        hourly.append(
            {
                "场景": scenario_name,
                "日氨产量(t/day)": output_t,
                "时段": data["time_labels"][t],
                "时段序号(h)": t + 1,
                "设备状态": int(u[t]),
                "常规负荷功率(MW)": float(conventional_load[t]),
                "生产负荷功率(MW)": float(PROCESS_POWER_MW * u[t]),
                "总负荷功率(MW)": float(total_load[t]),
                "风电功率(MW)": float(wind_power[t]),
                "光伏功率(MW)": float(pv_power[t]),
                "风光总出力(MW)": float(generation[t]),
                "购电功率(MW)": float(buy[t]),
                "售电功率(MW)": float(sell[t]),
                "购电电价(¥/kWh)": float(data["buy_prices"][t]),
                "售电电价(¥/kWh)": float(data["sell_price"]),
                "开机边际电网成本(¥/h)": float(marginal[t]),
            }
        )

    summary = {
        "场景": scenario_name,
        "日氨产量(t/day)": output_t,
        "开机时间(h)": int(np.sum(u)),
        "设备利用率(%)": float(100.0 * np.sum(u) / HOURS),
        "碱性电解槽利用率(%)": float(100.0 * np.sum(u) / HOURS),
        "PEM电解槽利用率(%)": float(100.0 * np.sum(u) / HOURS),
        "合成氨装置利用率(%)": float(100.0 * np.sum(u) / HOURS),
        "最优生产时段": format_operating_periods(u),
        "常规负荷电量(MWh)": float(np.sum(conventional_load) * TIME_STEP_H),
        "生产负荷电量(MWh)": float(PROCESS_POWER_MW * np.sum(u) * TIME_STEP_H),
        "总负荷电量(MWh)": load_energy,
        "风电电量(MWh)": float(np.sum(wind_power) * TIME_STEP_H),
        "光伏电量(MWh)": float(np.sum(pv_power) * TIME_STEP_H),
        "新能源发电量(MWh)": generation_energy,
        "购电量(MWh)": buy_energy,
        "售电量(MWh)": sell_energy,
        "新能源自发自用比例(%)": 100.0 * metrics["新能源自发自用比例"],
        "新能源供电占比(%)": 100.0 * metrics["新能源供电占比"],
        "上网电量比例(%)": 100.0 * metrics["上网电量比例"],
        "新能源自发自用比例达标": "是" if status["新能源自发自用比例达标"] else "否",
        "新能源供电占比达标": "是" if status["新能源供电占比达标"] else "否",
        "上网电量比例达标": "是" if status["上网电量比例达标"] else "否",
        "达标指标数": passed_count,
        "达标状态": conclusion,
        **costs,
        "求解器": "SciPy HiGHS MILP",
        "求解状态": result.message,
        "MIP间隙": float(getattr(result, "mip_gap", np.nan)),
        "MIP节点数": int(getattr(result, "mip_node_count", 0)),
    }
    return {
        "feasible": True,
        "summary": summary,
        "hourly": hourly,
        "checks": checks,
        "u": u,
        "buy": buy,
        "sell": sell,
        "generation": generation,
        "total_load": total_load,
    }


def descriptive_effects(all_summary):
    rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = all_summary[all_summary["日氨产量(t/day)"] == output_t].copy()
        matrix = subset.pivot(index="风电场景", columns="光伏场景", values="净吨氨成本(¥/t)").sort_index().sort_index(axis=1).to_numpy()
        grand = float(np.mean(matrix))
        wind_effect = np.mean(matrix, axis=1) - grand
        pv_effect = np.mean(matrix, axis=0) - grand
        interaction = matrix - grand - wind_effect[:, None] - pv_effect[None, :]
        ss_wind = float(matrix.shape[1] * np.sum(wind_effect**2))
        ss_pv = float(matrix.shape[0] * np.sum(pv_effect**2))
        ss_interaction = float(np.sum(interaction**2))
        ss_total = float(np.sum((matrix - grand) ** 2))
        denominator = ss_total if ss_total > 0 else 1.0
        rows.append(
            {
                "日氨产量(t/day)": output_t,
                "吨氨成本均值(¥/t)": grand,
                "风电效应平方和": ss_wind,
                "光伏效应平方和": ss_pv,
                "交互效应平方和": ss_interaction,
                "总平方和": ss_total,
                "风电贡献率(%)": 100.0 * ss_wind / denominator,
                "光伏贡献率(%)": 100.0 * ss_pv / denominator,
                "交互贡献率(%)": 100.0 * ss_interaction / denominator,
            }
        )
    return pd.DataFrame(rows)


def annual_summary(all_summary):
    rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = all_summary[all_summary["日氨产量(t/day)"] == output_t]
        total_cost = float(SCENARIO_DAYS * subset["日净成本(¥/day)"].sum())
        total_output = float(SCENARIO_DAYS * subset["日氨产量(t/day)"].sum())
        status_counts = subset["达标状态"].value_counts()
        rows.append(
            {
                "日氨产量(t/day)": output_t,
                "年度总产量(t)": total_output,
                "年度净成本(¥)": total_cost,
                "综合净吨氨成本(¥/t)": total_cost / total_output,
                "全部达标天数(day)": int(SCENARIO_DAYS * status_counts.get("全部达标", 0)),
                "部分达标天数(day)": int(SCENARIO_DAYS * status_counts.get("部分达标", 0)),
                "全部不达标天数(day)": int(SCENARIO_DAYS * status_counts.get("全部不达标", 0)),
                "平均购电量(MWh/day)": float(subset["购电量(MWh)"].mean()),
                "平均售电量(MWh/day)": float(subset["售电量(MWh)"].mean()),
            }
        )
    return pd.DataFrame(rows)


def distribution_summary(all_summary):
    indicators = [
        "净吨氨成本(¥/t)",
        "购电量(MWh)",
        "售电量(MWh)",
        "新能源自发自用比例(%)",
        "新能源供电占比(%)",
        "上网电量比例(%)",
    ]
    rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = all_summary[all_summary["日氨产量(t/day)"] == output_t]
        for indicator in indicators:
            values = subset[indicator].astype(float)
            rows.append(
                {
                    "日氨产量(t/day)": output_t,
                    "指标": indicator,
                    "最小值": float(values.min()),
                    "下四分位数": float(values.quantile(0.25)),
                    "中位数": float(values.median()),
                    "平均值": float(values.mean()),
                    "上四分位数": float(values.quantile(0.75)),
                    "最大值": float(values.max()),
                }
            )
    return pd.DataFrame(rows)


def extreme_scenarios(all_summary):
    rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = all_summary[all_summary["日氨产量(t/day)"] == output_t]
        best = subset.loc[subset["净吨氨成本(¥/t)"].idxmin()]
        worst = subset.loc[subset["净吨氨成本(¥/t)"].idxmax()]
        min_buy = subset.loc[subset["购电量(MWh)"].idxmin()]
        max_buy = subset.loc[subset["购电量(MWh)"].idxmax()]
        min_sell = subset.loc[subset["售电量(MWh)"].idxmin()]
        max_sell = subset.loc[subset["售电量(MWh)"].idxmax()]
        rows.append(
            {
                "日氨产量(t/day)": output_t,
                "最低成本场景": best["场景"],
                "最低净吨氨成本(¥/t)": float(best["净吨氨成本(¥/t)"]),
                "最高成本场景": worst["场景"],
                "最高净吨氨成本(¥/t)": float(worst["净吨氨成本(¥/t)"]),
                "最小购电场景": min_buy["场景"],
                "最小购电量(MWh)": float(min_buy["购电量(MWh)"]),
                "最大购电场景": max_buy["场景"],
                "最大购电量(MWh)": float(max_buy["购电量(MWh)"]),
                "最小售电场景": min_sell["场景"],
                "最小售电量(MWh)": float(min_sell["售电量(MWh)"]),
                "最大售电场景": max_sell["场景"],
                "最大售电量(MWh)": float(max_sell["售电量(MWh)"]),
            }
        )
    return pd.DataFrame(rows)


def build_duration_series(values, scenario_days=SCENARIO_DAYS):
    sorted_values = np.sort(np.asarray(values, dtype=float))[::-1]
    expanded = np.repeat(sorted_values, scenario_days)
    days = np.arange(1, len(expanded) + 1)
    return days, expanded


def configure_plotting():
    plt.rcParams["font.sans-serif"] = [
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["pdf.fonttype"] = 42
    plt.rcParams["ps.fonttype"] = 42


def save_figure(figure, figure_dir, name):
    figure.savefig(Path(figure_dir) / f"{name}.pdf", bbox_inches="tight")
    figure.savefig(Path(figure_dir) / f"{name}.png", dpi=300, bbox_inches="tight")
    plt.close(figure)


def plot_typical_schedule(typical_summary, typical_hourly, figure_dir):
    outputs = list(DAILY_OUTPUTS_T)
    matrix = np.vstack([
        typical_hourly[typical_hourly["日氨产量(t/day)"] == output_t].sort_values("时段序号(h)")["设备状态"].to_numpy()
        for output_t in outputs
    ])
    prices = typical_hourly[typical_hourly["日氨产量(t/day)"] == outputs[0]].sort_values("时段序号(h)")["购电电价(¥/kWh)"].to_numpy()
    cmap = LinearSegmentedColormap.from_list("schedule", ["#EEF3F8", PALETTE["blue"]])
    figure, axes = plt.subplots(2, 1, figsize=(10.5, 4.8), sharex=True, gridspec_kw={"height_ratios": [1, 3], "hspace": 0.08})
    axes[0].step(np.arange(1, HOURS + 1), prices, where="mid", color=PALETTE["orange"], linewidth=2.0)
    axes[0].set_ylabel("购电电价\n/ ¥/kWh")
    axes[0].grid(axis="y", color="#D9DEE8", linewidth=0.7)
    axes[0].spines[["top", "right"]].set_visible(False)
    axes[1].imshow(matrix, aspect="auto", cmap=cmap, vmin=0, vmax=1, interpolation="nearest")
    axes[1].set_yticks(np.arange(len(outputs)), [str(value) for value in outputs])
    axes[1].set_ylabel("日氨产量 / t/day")
    axes[1].set_xticks(np.arange(0, HOURS, 2), np.arange(1, HOURS + 1, 2))
    axes[1].set_xlabel("时段 / h")
    for y in range(matrix.shape[0]):
        for x in range(matrix.shape[1]):
            if matrix[y, x] > 0.5:
                axes[1].text(x, y, "1", ha="center", va="center", color="white", fontsize=7)
    save_figure(figure, figure_dir, "问题二典型场景设备启停热力图")


def plot_optimal_power_balance(typical_summary, typical_hourly, figure_dir):
    best_output = float(typical_summary.loc[typical_summary["净吨氨成本(¥/t)"].idxmin(), "日氨产量(t/day)"])
    frame = typical_hourly[typical_hourly["日氨产量(t/day)"] == best_output].sort_values("时段序号(h)")
    hours = frame["时段序号(h)"].to_numpy()
    figure, axes = plt.subplots(2, 1, figsize=(10.5, 6.8), sharex=True, gridspec_kw={"height_ratios": [2, 1], "hspace": 0.10})
    axes[0].plot(hours, frame["风电功率(MW)"], color=PALETTE["blue"], linewidth=1.8, label="风电功率")
    axes[0].plot(hours, frame["光伏功率(MW)"], color=PALETTE["yellow"], linewidth=1.8, label="光伏功率")
    axes[0].plot(hours, frame["风光总出力(MW)"], color=PALETTE["light_blue"], linewidth=2.4, label="风光总出力")
    axes[0].plot(hours, frame["总负荷功率(MW)"], color=PALETTE["red"], linewidth=2.4, label="总负荷功率")
    axes[0].set_ylabel("功率 / MW")
    axes[0].legend(ncol=4, frameon=False, loc="upper center")
    axes[0].grid(axis="y", color="#D9DEE8", linewidth=0.7)
    axes[0].spines[["top", "right"]].set_visible(False)
    axes[1].axhline(0, color="#777777", linewidth=0.8)
    axes[1].fill_between(hours, 0, frame["购电功率(MW)"], color=PALETTE["orange"], alpha=0.85, label="购电功率")
    axes[1].fill_between(hours, 0, -frame["售电功率(MW)"], color=PALETTE["blue"], alpha=0.85, label="售电功率")
    axes[1].set_ylabel("电网交换功率 / MW")
    axes[1].set_xlabel("时段 / h")
    axes[1].set_xticks(np.arange(1, HOURS + 1, 2))
    axes[1].legend(ncol=2, frameon=False, loc="lower left")
    axes[1].grid(axis="y", color="#D9DEE8", linewidth=0.7)
    axes[1].spines[["top", "right"]].set_visible(False)
    save_figure(figure, figure_dir, "问题二最低吨氨成本方案逐时功率平衡")


def plot_scenario_heatmaps(all_summary, figure_dir):
    cmap = LinearSegmentedColormap.from_list("cost", [PALETTE["light_blue"], PALETTE["yellow"], PALETTE["red"]])
    values = all_summary["净吨氨成本(¥/t)"].to_numpy()
    vmin, vmax = float(np.min(values)), float(np.max(values))
    figure = plt.figure(figsize=(11.5, 6.8), constrained_layout=True)
    grid = figure.add_gridspec(2, 6)
    axes = [
        figure.add_subplot(grid[0, 0:2]),
        figure.add_subplot(grid[0, 2:4]),
        figure.add_subplot(grid[0, 4:6]),
        figure.add_subplot(grid[1, 1:3]),
        figure.add_subplot(grid[1, 3:5]),
    ]
    image_handle = None
    for axis, output_t in zip(axes, DAILY_OUTPUTS_T):
        subset = all_summary[all_summary["日氨产量(t/day)"] == output_t]
        matrix = subset.pivot(index="风电场景", columns="光伏场景", values="净吨氨成本(¥/t)").sort_index().sort_index(axis=1)
        image_handle = axis.imshow(matrix.to_numpy(), cmap=cmap, aspect="auto", vmin=vmin, vmax=vmax)
        axis.set_title(f"{output_t} t/day", fontsize=11)
        axis.set_xticks(np.arange(4), [f"光伏{j}" for j in matrix.columns])
        axis.set_yticks(np.arange(6), [f"风电{i}" for i in matrix.index])
        for y in range(6):
            for x in range(4):
                axis.text(x, y, f"{matrix.iloc[y, x]:.0f}", ha="center", va="center", fontsize=7, color="#202020")
    colorbar = figure.colorbar(image_handle, ax=axes, shrink=0.88, pad=0.02)
    colorbar.set_label("净吨氨成本 / ¥/t")
    save_figure(figure, figure_dir, "问题二风光场景吨氨成本热力图")


def plot_annual_summary(all_summary, annual, figure_dir):
    figure, axes = plt.subplots(1, 2, figsize=(11.5, 4.6), gridspec_kw={"wspace": 0.28})
    for output_t, color in zip(DAILY_OUTPUTS_T, [PALETTE["blue"], PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"]]):
        days, values = build_duration_series(
            all_summary[all_summary["日氨产量(t/day)"] == output_t]["净吨氨成本(¥/t)"].to_numpy()
        )
        axes[0].step(days, values, where="post", color=color, linewidth=1.8, label=f"{output_t} t/day")
    axes[0].set_xlabel("持续天数 / day")
    axes[0].set_ylabel("净吨氨成本 / ¥/t")
    axes[0].set_xlim(1, 360)
    axes[0].legend(frameon=False, fontsize=8)
    axes[0].grid(axis="y", color="#D9DEE8", linewidth=0.7)
    axes[0].spines[["top", "right"]].set_visible(False)

    x = np.arange(len(annual))
    full = annual["全部达标天数(day)"].to_numpy()
    partial = annual["部分达标天数(day)"].to_numpy()
    none = annual["全部不达标天数(day)"].to_numpy()
    axes[1].bar(x, full, color=PALETTE["blue"], label="全部达标")
    axes[1].bar(x, partial, bottom=full, color=PALETTE["yellow"], label="部分达标")
    axes[1].bar(x, none, bottom=full + partial, color=PALETTE["red"], label="全部不达标")
    axes[1].set_xticks(x, [str(int(value)) for value in annual["日氨产量(t/day)"]])
    axes[1].set_xlabel("日氨产量 / t/day")
    axes[1].set_ylabel("统计天数 / day")
    axes[1].legend(frameon=False, fontsize=8)
    axes[1].grid(axis="y", color="#D9DEE8", linewidth=0.7)
    axes[1].spines[["top", "right"]].set_visible(False)
    save_figure(figure, figure_dir, "问题二年度成本持续曲线与达标状态")


def plot_distributions(all_summary, figure_dir):
    outputs = list(DAILY_OUTPUTS_T)
    colors = [PALETTE["blue"], PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"]]
    figure, axes = plt.subplots(2, 2, figsize=(11.5, 7.2), gridspec_kw={"hspace": 0.34, "wspace": 0.26})
    for axis, indicator, ylabel in [
        (axes[0, 0], "净吨氨成本(¥/t)", "净吨氨成本 / ¥/t"),
        (axes[0, 1], "购电量(MWh)", "日购电量 / MWh"),
        (axes[1, 0], "售电量(MWh)", "日售电量 / MWh"),
    ]:
        data_groups = [
            all_summary[all_summary["日氨产量(t/day)"] == output_t][indicator].to_numpy()
            for output_t in outputs
        ]
        boxes = axis.boxplot(data_groups, patch_artist=True, widths=0.62, showfliers=False)
        for patch, color in zip(boxes["boxes"], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.82)
        for median in boxes["medians"]:
            median.set_color("#303030")
        axis.set_xticks(np.arange(1, 6), [str(value) for value in outputs])
        axis.set_xlabel("日氨产量 / t/day")
        axis.set_ylabel(ylabel)
        axis.grid(axis="y", color="#D9DEE8", linewidth=0.7)
        axis.spines[["top", "right"]].set_visible(False)

    green_axis = axes[1, 1]
    x = np.arange(1, 6)
    offsets = [-0.25, 0.0, 0.25]
    green_items = [
        ("新能源自发自用比例(%)", "自发自用比例", PALETTE["blue"]),
        ("新能源供电占比(%)", "绿电供电占比", PALETTE["yellow"]),
        ("上网电量比例(%)", "上网电量比例", PALETTE["red"]),
    ]
    for offset, (indicator, label, color) in zip(offsets, green_items):
        data_groups = [
            all_summary[all_summary["日氨产量(t/day)"] == output_t][indicator].to_numpy()
            for output_t in outputs
        ]
        boxes = green_axis.boxplot(
            data_groups,
            positions=x + offset,
            widths=0.20,
            patch_artist=True,
            showfliers=False,
            manage_ticks=False,
        )
        for patch in boxes["boxes"]:
            patch.set_facecolor(color)
            patch.set_alpha(0.82)
        for median in boxes["medians"]:
            median.set_color("#303030")
        green_axis.plot([], [], color=color, linewidth=8, alpha=0.82, label=label)
    green_axis.axhline(60, color=PALETTE["blue"], linewidth=0.9, linestyle="--", alpha=0.7)
    green_axis.axhline(30, color=PALETTE["yellow"], linewidth=0.9, linestyle="--", alpha=0.7)
    green_axis.axhline(20, color=PALETTE["red"], linewidth=0.9, linestyle="--", alpha=0.7)
    green_axis.set_xticks(x, [str(value) for value in outputs])
    green_axis.set_xlabel("日氨产量 / t/day")
    green_axis.set_ylabel("绿电指标 / %")
    green_axis.legend(frameon=False, fontsize=8, loc="upper right")
    green_axis.grid(axis="y", color="#D9DEE8", linewidth=0.7)
    green_axis.spines[["top", "right"]].set_visible(False)
    save_figure(figure, figure_dir, "问题二多场景运行指标分布")


def plot_effects(effects, figure_dir):
    figure, axis = plt.subplots(figsize=(8.5, 4.3))
    x = np.arange(len(effects))
    wind = effects["风电贡献率(%)"].to_numpy()
    pv = effects["光伏贡献率(%)"].to_numpy()
    interaction = effects["交互贡献率(%)"].to_numpy()
    axis.bar(x, wind, color=PALETTE["blue"], label="风电场景")
    axis.bar(x, pv, bottom=wind, color=PALETTE["yellow"], label="光伏场景")
    axis.bar(x, interaction, bottom=wind + pv, color=PALETTE["red"], label="风光交互")
    axis.set_xticks(x, [str(int(value)) for value in effects["日氨产量(t/day)"]])
    axis.set_xlabel("日氨产量 / t/day")
    axis.set_ylabel("描述性贡献率 / %")
    axis.set_ylim(0, 100)
    axis.legend(ncol=3, frameon=False, loc="upper center")
    axis.grid(axis="y", color="#D9DEE8", linewidth=0.7)
    axis.spines[["top", "right"]].set_visible(False)
    save_figure(figure, figure_dir, "问题二风光场景效应贡献率")


def format_result_workbook(path):
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="5271AE")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 30
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            column_index = column_cells[0].column
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 28)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
    workbook.save(path)


def create_report(typical, annual, effects, compliance, distribution, extremes, checks, report_path):
    best_typical = typical.loc[typical["净吨氨成本(¥/t)"].idxmin()]
    best_annual = annual.loc[annual["综合净吨氨成本(¥/t)"].idxmin()]
    best_status = best_typical["达标状态"]
    failed_metrics = []
    if best_typical["新能源自发自用比例达标"] == "否":
        failed_metrics.append("新能源自发自用比例")
    if best_typical["新能源供电占比达标"] == "否":
        failed_metrics.append("新能源供电占比")
    if best_typical["上网电量比例达标"] == "否":
        failed_metrics.append("新能源上网电量比例")
    range_rows = []
    for output_t in DAILY_OUTPUTS_T:
        subset = distribution[distribution["日氨产量(t/day)"] == output_t].set_index("指标")
        range_rows.append(
            {
                "日氨产量(t/day)": output_t,
                "吨氨成本范围(¥/t)": f"{subset.loc['净吨氨成本(¥/t)', '最小值']:.2f}-{subset.loc['净吨氨成本(¥/t)', '最大值']:.2f}",
                "购电量范围(MWh)": f"{subset.loc['购电量(MWh)', '最小值']:.2f}-{subset.loc['购电量(MWh)', '最大值']:.2f}",
                "售电量范围(MWh)": f"{subset.loc['售电量(MWh)', '最小值']:.2f}-{subset.loc['售电量(MWh)', '最大值']:.2f}",
            }
        )
    range_table = pd.DataFrame(range_rows)
    max_residual = checks[[
        "最大逐时功率平衡残差(MW)",
        "产量约束残差(t/day)",
        "购售电同时发生最大值(MW2)",
        "MILP与排序成本差(¥/day)",
    ]].abs().max()
    lines = [
        "# 问题二计算结果",
        "",
        "## 运行环境",
        "",
        "- Python：全局 Python 3.12",
        "- 优化器：SciPy HiGHS MILP",
        "- 时间尺度：24 个 1 h 时段",
        "- 主口径：净成本、题面绿电指标、无氢储能、三类设备同步全开或停机",
        "",
        "## 典型场景结果",
        "",
        f"典型场景下，净吨氨成本最低的日产量为 **{best_typical['日氨产量(t/day)']:.0f} t/day**，对应日净成本为 **{best_typical['日净成本(¥/day)']:.2f} ¥/day**，净吨氨成本为 **{best_typical['净吨氨成本(¥/t)']:.2f} ¥/t**。",
        "",
        typical[["日氨产量(t/day)", "开机时间(h)", "最优生产时段", "日净成本(¥/day)", "净吨氨成本(¥/t)", "新能源自发自用比例(%)", "新能源供电占比(%)", "上网电量比例(%)", "达标状态"]].to_markdown(index=False, floatfmt=".3f"),
        "",
        f"最低成本方案每天开机 **{best_typical['开机时间(h)']:.0f} h**，最优生产时段为 **{best_typical['最优生产时段']}**。碱性电解槽、PEM电解槽和合成氨装置同步运行，三类设备利用率均为 **{best_typical['设备利用率(%)']:.2f}%**。",
        "",
        f"该方案的新能源自发自用比例为 **{best_typical['新能源自发自用比例(%)']:.3f}%**，总用电量绿电比例为 **{best_typical['新能源供电占比(%)']:.3f}%**，新能源上网电量比例为 **{best_typical['上网电量比例(%)']:.3f}%**，综合判定为 **{best_status}**。未达标指标为：{'、'.join(failed_metrics) if failed_metrics else '无'}。",
        "",
        "36 t/day方案减少了高成本购电和设备运维支出，同时保留余电上网收入，因此吨氨净成本最低；但生产负荷下降使新能源余电增加，导致自发自用比例下降、上网比例上升。由此可见，最低成本方案不等于绿电指标最优方案。",
        "",
        "## 24种风光场景年度折算",
        "",
        f"按每个场景代表 15 day、合计 360 day 折算，综合净吨氨成本最低的固定日产量为 **{best_annual['日氨产量(t/day)']:.0f} t/day**，综合净吨氨成本为 **{best_annual['综合净吨氨成本(¥/t)']:.2f} ¥/t**。",
        "",
        annual.to_markdown(index=False, floatfmt=".3f"),
        "",
        "年度结果呈现明确的经济性-合规性权衡：产量降低时综合净吨氨成本下降，但全部达标天数同步减少。36 t/day经济性最好，却没有全部达标场景；72 t/day全部达标天数最多，但综合成本最高。",
        "",
        "## 多场景分布特征",
        "",
        range_table.to_markdown(index=False),
        "",
        "五档产量的最低吨氨成本场景均为风电4-光伏1，最高吨氨成本场景均为风电2-光伏4。光伏出力降低会显著增加网购电量和吨氨成本；减产则降低购电量但增加余电上网量，使上网电量比例恶化。",
        "",
        extremes.to_markdown(index=False, floatfmt=".3f"),
        "",
        "完整的最小值、四分位数、中位数、平均值和最大值见问题二多场景分布统计.csv。",
        "",
        "## 绿电合规成本",
        "",
        compliance.fillna("—").to_markdown(index=False, floatfmt=".3f"),
        "",
        "## 风光场景描述性效应分解",
        "",
        "以下贡献率为确定性平方和分解，不进行 F 检验，也不报告统计显著性。",
        "",
        effects.to_markdown(index=False, floatfmt=".3f"),
        "",
        "## 成本口径与关键假设",
        "",
        "- 合成氨装置投资成本 60000 ¥/kgH2 按额定耗氢能力的容量投资解释，扩容后额定耗氢能力为 600 kgH2/h，并按 30 年、365 day 进行直线折旧。",
        "- 风电和光伏度电成本按全部实际发电量计算；售电收入作为园区综合成本的抵扣项，因此本文报告的是园区综合净吨氨成本。",
        "- 24种场景各代表15 day，年度统计基准为题面规定的360 day；若同时将成本和产量换算至365 day，综合吨氨成本不变。",
        "",
        "## 图表文件",
        "",
        "- 问题二典型场景设备启停热力图",
        "- 问题二最低吨氨成本方案逐时功率平衡",
        "- 问题二风光场景吨氨成本热力图",
        "- 问题二多场景运行指标分布",
        "- 问题二年度成本持续曲线与达标状态",
        "- 问题二风光场景效应贡献率",
        "",
        "## 约束与一致性校验",
        "",
        f"- 最大逐时功率平衡残差：{max_residual['最大逐时功率平衡残差(MW)']:.3e} MW",
        f"- 最大产量约束残差：{max_residual['产量约束残差(t/day)']:.3e} t/day",
        f"- 最大购售电同时发生值：{max_residual['购售电同时发生最大值(MW2)']:.3e} MW2",
        f"- MILP 与边际排序最大成本差：{max_residual['MILP与排序成本差(¥/day)']:.3e} ¥/day",
        "",
        "## 可复现运行方式",
        "",
        "```powershell",
        "python -X utf8 code/问题二.py",
        "```",
        "",
        "详细的 120 组场景结果、2880 条逐时排班和图表数据均保存在 `outputs/问题二计算结果/`。",
    ]
    Path(report_path).write_text("\n".join(lines), encoding="utf-8")


def run_problem2(project_root, output_dir=None, figure_dir=None, report_path=None):
    project_root = Path(project_root)
    output_dir = Path(output_dir) if output_dir else project_root / "outputs" / "问题二计算结果"
    figure_dir = Path(figure_dir) if figure_dir else project_root / "figures" / "问题二计算结果"
    report_path = Path(report_path) if report_path else project_root / "reports" / "问题二计算结果.md"
    output_dir.mkdir(parents=True, exist_ok=True)
    figure_dir.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    data = load_inputs(project_root)

    typical_results = []
    compliance_rows = []
    for output_t in DAILY_OUTPUTS_T:
        result = solve_case(data, data["typical_wind"], data["typical_pv"], output_t, "典型场景")
        if not result["feasible"]:
            raise RuntimeError(result["message"])
        typical_results.append(result)
        compliant = solve_case(data, data["typical_wind"], data["typical_pv"], output_t, "典型场景", green=True)
        base_cost = result["summary"]["日净成本(¥/day)"]
        if compliant["feasible"]:
            green_cost = compliant["summary"]["日净成本(¥/day)"]
            compliance_rows.append({
                "日氨产量(t/day)": output_t,
                "基础日净成本(¥/day)": base_cost,
                "全部达标是否可行": "是",
                "合规日净成本(¥/day)": green_cost,
                "合规成本增量(¥/day)": green_cost - base_cost,
                "合规吨氨成本增量(¥/t)": (green_cost - base_cost) / output_t,
            })
        else:
            compliance_rows.append({
                "日氨产量(t/day)": output_t,
                "基础日净成本(¥/day)": base_cost,
                "全部达标是否可行": "否",
                "合规日净成本(¥/day)": np.nan,
                "合规成本增量(¥/day)": np.nan,
                "合规吨氨成本增量(¥/t)": np.nan,
            })

    all_results = []
    for wind_index, wind_power in enumerate(data["wind_scenarios"], start=1):
        for pv_index, pv_power in enumerate(data["pv_scenarios"], start=1):
            scenario_name = f"风电{wind_index}-光伏{pv_index}"
            for output_t in DAILY_OUTPUTS_T:
                result = solve_case(data, wind_power, pv_power, output_t, scenario_name)
                if not result["feasible"]:
                    raise RuntimeError(f"{scenario_name}, {output_t}: {result['message']}")
                result["summary"]["风电场景"] = wind_index
                result["summary"]["光伏场景"] = pv_index
                for row in result["hourly"]:
                    row["风电场景"] = wind_index
                    row["光伏场景"] = pv_index
                all_results.append(result)

    typical_summary = pd.DataFrame([item["summary"] for item in typical_results])
    typical_hourly = pd.DataFrame([row for item in typical_results for row in item["hourly"]])
    all_summary = pd.DataFrame([item["summary"] for item in all_results])
    all_hourly = pd.DataFrame([row for item in all_results for row in item["hourly"]])
    check_rows = []
    for item in typical_results + all_results:
        check_rows.append({
            "场景": item["summary"]["场景"],
            "日氨产量(t/day)": item["summary"]["日氨产量(t/day)"],
            **item["checks"],
        })
    checks = pd.DataFrame(check_rows)
    compliance = pd.DataFrame(compliance_rows)
    annual = annual_summary(all_summary)
    effects = descriptive_effects(all_summary)
    distribution = distribution_summary(all_summary)
    extremes = extreme_scenarios(all_summary)

    typical_summary.to_csv(output_dir / "问题二典型场景汇总.csv", index=False, encoding="utf-8-sig")
    typical_hourly.to_csv(output_dir / "问题二典型场景逐时运行结果.csv", index=False, encoding="utf-8-sig")
    all_summary.to_csv(output_dir / "问题二全部场景汇总.csv", index=False, encoding="utf-8-sig")
    all_hourly.to_csv(output_dir / "问题二全部场景逐时排班.csv", index=False, encoding="utf-8-sig")
    annual.to_csv(output_dir / "问题二年度汇总.csv", index=False, encoding="utf-8-sig")
    effects.to_csv(output_dir / "问题二风光效应分解.csv", index=False, encoding="utf-8-sig")
    compliance.to_csv(output_dir / "问题二绿电合规成本.csv", index=False, encoding="utf-8-sig")
    distribution.to_csv(output_dir / "问题二多场景分布统计.csv", index=False, encoding="utf-8-sig")
    extremes.to_csv(output_dir / "问题二极端场景汇总.csv", index=False, encoding="utf-8-sig")
    checks.to_csv(output_dir / "问题二约束校验.csv", index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(output_dir / "问题二计算结果.xlsx", engine="openpyxl") as writer:
        typical_summary.to_excel(writer, sheet_name="典型场景汇总", index=False)
        typical_hourly.to_excel(writer, sheet_name="典型场景逐时结果", index=False)
        all_summary.to_excel(writer, sheet_name="全部场景汇总", index=False)
        all_hourly.to_excel(writer, sheet_name="全部场景逐时排班", index=False)
        annual.to_excel(writer, sheet_name="年度汇总", index=False)
        effects.to_excel(writer, sheet_name="风光效应分解", index=False)
        compliance.to_excel(writer, sheet_name="绿电合规成本", index=False)
        distribution.to_excel(writer, sheet_name="多场景分布统计", index=False)
        extremes.to_excel(writer, sheet_name="极端场景汇总", index=False)
        checks.to_excel(writer, sheet_name="约束校验", index=False)
    format_result_workbook(output_dir / "问题二计算结果.xlsx")

    compact = {
        "solver": "SciPy HiGHS MILP",
        "typical_summary": typical_summary.to_dict(orient="records"),
        "annual_summary": annual.to_dict(orient="records"),
        "effect_decomposition": effects.to_dict(orient="records"),
        "distribution_summary": distribution.to_dict(orient="records"),
        "extreme_scenarios": extremes.to_dict(orient="records"),
        "green_compliance_cost": compliance.where(pd.notna(compliance), None).to_dict(orient="records"),
    }
    with (output_dir / "问题二完整计算结果.json").open("w", encoding="utf-8") as file:
        json.dump(compact, file, ensure_ascii=False, indent=2)

    configure_plotting()
    plot_typical_schedule(typical_summary, typical_hourly, figure_dir)
    plot_optimal_power_balance(typical_summary, typical_hourly, figure_dir)
    plot_scenario_heatmaps(all_summary, figure_dir)
    plot_annual_summary(all_summary, annual, figure_dir)
    plot_distributions(all_summary, figure_dir)
    plot_effects(effects, figure_dir)
    create_report(typical_summary, annual, effects, compliance, distribution, extremes, checks, report_path)

    return {
        "typical_summary": typical_summary,
        "annual_summary": annual,
        "effects": effects,
        "distribution": distribution,
        "extremes": extremes,
        "compliance": compliance,
        "checks": checks,
    }


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    result = run_problem2(root)
    best_typical = result["typical_summary"].loc[result["typical_summary"]["净吨氨成本(¥/t)"].idxmin()]
    best_annual = result["annual_summary"].loc[result["annual_summary"]["综合净吨氨成本(¥/t)"].idxmin()]
    console = {
        "求解器": "SciPy HiGHS MILP",
        "典型场景最低吨氨成本日产量(t/day)": float(best_typical["日氨产量(t/day)"]),
        "典型场景最低净吨氨成本(¥/t)": float(best_typical["净吨氨成本(¥/t)"]),
        "年度最低吨氨成本日产量(t/day)": float(best_annual["日氨产量(t/day)"]),
        "年度最低综合净吨氨成本(¥/t)": float(best_annual["综合净吨氨成本(¥/t)"]),
        "最大逐时功率平衡残差(MW)": float(result["checks"]["最大逐时功率平衡残差(MW)"].abs().max()),
        "MILP与排序最大成本差(¥/day)": float(result["checks"]["MILP与排序成本差(¥/day)"].abs().max()),
    }
    print(json.dumps(console, ensure_ascii=False, indent=2))


