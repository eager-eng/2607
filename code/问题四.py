import argparse
import itertools
import json
import math
import re
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, linprog, milp
from scipy.sparse import coo_matrix, csr_matrix, vstack

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap


ALK_MAX_MW = 20.0
PEM_MAX_MW = 20.0
AMMONIA_MAX_MW = 1.5
HOURS = 24
SCENARIO_DAYS = 15
WIND_CAPACITY_MW = 40.0
PV_CAPACITY_MW = 64.0
CONVENTIONAL_LOAD_BASE_MW = 6.0
HYDROGEN_RATE_KG_PER_H = 600.0
MAIN_C_RATE = 0.5
DAILY_SELF_LOSS = 0.002
PALETTE = {
    "blue": "#5271AE",
    "light_blue": "#70ACDE",
    "yellow": "#F5CC7D",
    "orange": "#FFA660",
    "red": "#D85B59",
}


def find_attachment(project_root, prefix):
    matches = [
        path
        for path in Path(project_root).joinpath("题目").rglob("*.xlsx")
        if not path.name.startswith("~$") and path.name.startswith(prefix)
    ]
    if len(matches) != 1:
        raise FileNotFoundError(f"Expected one {prefix} workbook, found {len(matches)}")
    return matches[0]


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell.value for cell in row) for row in workbook.active.iter_rows()]
    finally:
        workbook.close()


def extract_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        raise ValueError(f"No numeric value found in {value!r}")
    return float(match.group())


def hourly_buy_prices(rows):
    peak_price = extract_number(rows[1][1])
    flat_price = extract_number(rows[2][1])
    valley_price = extract_number(rows[3][1])
    prices = []
    for hour in range(HOURS):
        if 10 <= hour < 15 or 18 <= hour < 21:
            prices.append(peak_price)
        elif 7 <= hour < 10 or 15 <= hour < 18 or 21 <= hour < 23:
            prices.append(flat_price)
        else:
            prices.append(valley_price)
    return np.asarray(prices, dtype=float)


def load_inputs(project_root):
    load_rows = read_rows(find_attachment(project_root, "附件1"))
    wind_rows = read_rows(find_attachment(project_root, "附件3"))
    pv_rows = read_rows(find_attachment(project_root, "附件4"))
    equipment_rows = read_rows(find_attachment(project_root, "附件5"))
    storage_rows = read_rows(find_attachment(project_root, "附件6"))
    buy_rows = read_rows(find_attachment(project_root, "附件7"))
    sell_rows = read_rows(find_attachment(project_root, "附件8"))
    time_labels = [str(row[0]) for row in load_rows[1:]]
    wind_times = [str(row[0]) for row in wind_rows[1:]]
    pv_times = [str(row[0]) for row in pv_rows[1:]]
    if not (len(time_labels) == HOURS and time_labels == wind_times and time_labels == pv_times):
        raise ValueError("Hourly input tables must contain the same 24 periods")
    load_pu = np.asarray([row[1] for row in load_rows[1:]], dtype=float)
    wind_pu = np.asarray([row[1:7] for row in wind_rows[1:]], dtype=float).T
    pv_pu = np.asarray([row[1:5] for row in pv_rows[1:]], dtype=float).T
    if wind_pu.shape != (6, HOURS) or pv_pu.shape != (4, HOURS):
        raise ValueError("Scenario tables must have dimensions 6x24 and 4x24")
    sell_wind_price = extract_number(sell_rows[1][1])
    sell_pv_price = extract_number(sell_rows[2][1])
    if not np.isclose(sell_wind_price, sell_pv_price):
        raise ValueError("Aggregate surplus requires a common feed-in price")
    return {
        "time_labels": time_labels,
        "conventional_load": CONVENTIONAL_LOAD_BASE_MW * load_pu,
        "wind_scenarios": WIND_CAPACITY_MW * wind_pu,
        "pv_scenarios": PV_CAPACITY_MW * pv_pu,
        "buy_prices": hourly_buy_prices(buy_rows),
        "sell_price": sell_wind_price,
        "wind_cost": extract_number(equipment_rows[1][1]),
        "pv_cost": extract_number(equipment_rows[1][2]),
        "alk_om": extract_number(equipment_rows[2][3]),
        "pem_om": extract_number(equipment_rows[2][4]),
        "battery_investment": extract_number(storage_rows[1][1]),
        "battery_om": extract_number(storage_rows[2][1]),
        "battery_lifetime": extract_number(storage_rows[3][1]),
        "eta_charge": 0.90,
        "eta_discharge": 0.90,
        "daily_self_loss": extract_number(storage_rows[5][1]) / 100.0,
        "ammonia_investment": extract_number(storage_rows[1][2]),
        "ammonia_om": extract_number(storage_rows[2][2]),
        "ammonia_lifetime": extract_number(storage_rows[3][2]),
    }


def hourly_self_loss(daily_loss):
    if not 0.0 <= daily_loss < 1.0:
        raise ValueError("Daily self-loss must be in [0, 1)")
    return 1.0 - (1.0 - float(daily_loss)) ** (1.0 / 24.0)


def material_residual(alk_power, pem_power, ammonia_power):
    return 14.0 * alk_power + 16.0 * pem_power - 400.0 * ammonia_power


def battery_daily_fixed_cost(capacity_mwh, unit_cost_yuan_per_kwh=1000.0, lifetime_years=15.0):
    if capacity_mwh < 0.0:
        raise ValueError("Battery capacity cannot be negative")
    return 1000.0 * unit_cost_yuan_per_kwh * capacity_mwh / (lifetime_years * 365.0)


def _solve_hour_state(generation, conventional_load, state, om_costs):
    bounds = []
    for enabled, rated in zip(state, (ALK_MAX_MW, PEM_MAX_MW, AMMONIA_MAX_MW)):
        bounds.append((0.1 * rated, rated) if enabled else (0.0, 0.0))
    bounds.extend([(0.0, conventional_load), (0.0, None)])
    base_a_eq = np.asarray(
        [
            [14.0, 16.0, -400.0, 0.0, 0.0],
            [1.0, 1.0, 1.0, -1.0, 1.0],
        ],
        dtype=float,
    )
    base_b_eq = np.asarray([0.0, generation - conventional_load], dtype=float)
    objectives = (
        np.asarray([0.0, 0.0, 0.0, 1.0, 0.0]),
        np.asarray([0.0, 0.0, -1.0, 0.0, 0.0]),
        np.asarray([0.0, 0.0, 0.0, 0.0, 1.0]),
        1000.0
        * np.asarray([om_costs["alk"], om_costs["pem"], om_costs["ammonia"], 0.0, 0.0]),
    )
    a_eq = base_a_eq.copy()
    b_eq = base_b_eq.copy()
    result = None
    for objective in objectives:
        result = linprog(objective, A_eq=a_eq, b_eq=b_eq, bounds=bounds, method="highs")
        if not result.success:
            return None
        value = float(objective @ result.x)
        a_eq = np.vstack([a_eq, objective])
        b_eq = np.append(b_eq, value)
    return result.x


def solve_offgrid_hour(
    conventional_load,
    wind_power,
    pv_power,
    wind_cost,
    pv_cost,
    alk_om,
    pem_om,
    ammonia_om,
):
    del wind_cost, pv_cost
    generation = float(wind_power) + float(pv_power)
    om_costs = {"alk": alk_om, "pem": pem_om, "ammonia": ammonia_om}
    candidates = []
    for state in itertools.product((0, 1), repeat=3):
        solution = _solve_hour_state(generation, float(conventional_load), state, om_costs)
        if solution is None:
            continue
        alk, pem, ammonia, shed, curtail = solution
        operating_cost = 1000.0 * (alk_om * alk + pem_om * pem + ammonia_om * ammonia)
        key = (round(shed, 9), -round(ammonia, 9), round(curtail, 9), round(operating_cost, 6))
        candidates.append((key, state, solution, operating_cost))
    if not candidates:
        raise RuntimeError("No feasible off-grid hourly dispatch")
    _, state, solution, operating_cost = min(candidates, key=lambda item: item[0])
    alk, pem, ammonia, shed, curtail = solution
    return {
        "alk": float(alk),
        "pem": float(pem),
        "ammonia": float(ammonia),
        "shed": float(shed),
        "curtail": float(curtail),
        "u_alk": int(state[0]),
        "u_pem": int(state[1]),
        "u_ammonia": int(state[2]),
        "operating_cost": float(operating_cost),
    }


def _group_slices(names, hours=24):
    return {name: slice(i * hours, (i + 1) * hours) for i, name in enumerate(names)}


def _append_objective_lock(constraint, objective, value, tolerance=1e-7):
    matrix = vstack(
        [constraint.A, csr_matrix(np.asarray(objective, dtype=float).reshape(1, -1))],
        format="csr",
    )
    lower = np.append(np.asarray(constraint.lb, dtype=float), value - tolerance)
    upper = np.append(np.asarray(constraint.ub, dtype=float), value + tolerance)
    return LinearConstraint(matrix, lower, upper)


def _run_milp(objective, integrality, bounds, constraint, time_limit=120.0):
    result = milp(
        c=np.asarray(objective, dtype=float),
        integrality=integrality,
        bounds=bounds,
        constraints=constraint,
        options={"mip_rel_gap": 1e-9, "time_limit": time_limit, "presolve": True},
    )
    if not result.success:
        raise RuntimeError(result.message)
    return result


def _solve_lexicographic(objectives, integrality, bounds, base_constraint):
    constraint = base_constraint
    result = None
    values = []
    for objective in objectives:
        result = _run_milp(objective, integrality, bounds, constraint)
        value = float(np.asarray(objective) @ result.x)
        values.append(value)
        tolerance = max(1e-7, abs(value) * 1e-9)
        constraint = _append_objective_lock(constraint, objective, value, tolerance)
    return result, values


def _clean(values, tolerance=1e-8):
    array = np.asarray(values, dtype=float).copy()
    array[np.abs(array) < tolerance] = 0.0
    return array


def build_storage_milp(
    conventional_load,
    generation,
    capacity_mwh,
    c_rate,
    self_loss_hourly,
    eta_charge,
    eta_discharge,
    om_costs,
):
    conventional_load = np.asarray(conventional_load, dtype=float)
    generation = np.asarray(generation, dtype=float)
    if conventional_load.shape != (24,) or generation.shape != (24,):
        raise ValueError("Storage MILP inputs must contain 24 hourly values")
    if capacity_mwh < 0.0 or c_rate <= 0.0:
        raise ValueError("Storage capacity and C-rate are invalid")
    names = (
        "alk",
        "pem",
        "ammonia",
        "shed",
        "curtail",
        "charge",
        "discharge",
        "energy",
        "u_alk",
        "u_pem",
        "u_ammonia",
        "battery_mode",
    )
    slices = _group_slices(names)
    n = 24 * len(names)
    lower = np.zeros(n, dtype=float)
    upper = np.full(n, np.inf, dtype=float)
    upper[slices["alk"]] = ALK_MAX_MW
    upper[slices["pem"]] = PEM_MAX_MW
    upper[slices["ammonia"]] = AMMONIA_MAX_MW
    upper[slices["shed"]] = conventional_load
    upper[slices["curtail"]] = generation + c_rate * capacity_mwh
    upper[slices["charge"]] = c_rate * capacity_mwh
    upper[slices["discharge"]] = c_rate * capacity_mwh
    lower[slices["energy"]] = 0.1 * capacity_mwh
    upper[slices["energy"]] = 0.9 * capacity_mwh
    for name in ("u_alk", "u_pem", "u_ammonia", "battery_mode"):
        upper[slices[name]] = 1.0
    integrality = np.zeros(n, dtype=int)
    for name in ("u_alk", "u_pem", "u_ammonia", "battery_mode"):
        integrality[slices[name]] = 1

    rows = []
    cols = []
    vals = []
    row_lower = []
    row_upper = []
    row = 0

    def add(entries, lo, hi):
        nonlocal row
        for col, value in entries:
            rows.append(row)
            cols.append(col)
            vals.append(value)
        row_lower.append(lo)
        row_upper.append(hi)
        row += 1

    for t in range(24):
        add(
            [
                (slices["alk"].start + t, 1.0),
                (slices["pem"].start + t, 1.0),
                (slices["ammonia"].start + t, 1.0),
                (slices["shed"].start + t, -1.0),
                (slices["curtail"].start + t, 1.0),
                (slices["charge"].start + t, 1.0),
                (slices["discharge"].start + t, -1.0),
            ],
            generation[t] - conventional_load[t],
            generation[t] - conventional_load[t],
        )
        add(
            [
                (slices["alk"].start + t, 14.0),
                (slices["pem"].start + t, 16.0),
                (slices["ammonia"].start + t, -400.0),
            ],
            0.0,
            0.0,
        )
        for power_name, state_name, minimum, maximum in (
            ("alk", "u_alk", 2.0, ALK_MAX_MW),
            ("pem", "u_pem", 2.0, PEM_MAX_MW),
            ("ammonia", "u_ammonia", 0.15, AMMONIA_MAX_MW),
        ):
            add(
                [
                    (slices[power_name].start + t, 1.0),
                    (slices[state_name].start + t, -minimum),
                ],
                0.0,
                np.inf,
            )
            add(
                [
                    (slices[power_name].start + t, 1.0),
                    (slices[state_name].start + t, -maximum),
                ],
                -np.inf,
                0.0,
            )
        add(
            [
                (slices["charge"].start + t, 1.0),
                (slices["battery_mode"].start + t, -c_rate * capacity_mwh),
            ],
            -np.inf,
            0.0,
        )
        add(
            [
                (slices["discharge"].start + t, 1.0),
                (slices["battery_mode"].start + t, c_rate * capacity_mwh),
            ],
            -np.inf,
            c_rate * capacity_mwh,
        )
        previous = 23 if t == 0 else t - 1
        add(
            [
                (slices["energy"].start + t, 1.0),
                (slices["energy"].start + previous, -(1.0 - self_loss_hourly)),
                (slices["charge"].start + t, -eta_charge),
                (slices["discharge"].start + t, 1.0 / eta_discharge),
            ],
            0.0,
            0.0,
        )

    matrix = coo_matrix((vals, (rows, cols)), shape=(row, n)).tocsr()
    constraint = LinearConstraint(matrix, np.asarray(row_lower), np.asarray(row_upper))
    objectives = []
    shed_objective = np.zeros(n)
    shed_objective[slices["shed"]] = 1.0
    objectives.append(shed_objective)
    production_objective = np.zeros(n)
    production_objective[slices["ammonia"]] = -1.0
    objectives.append(production_objective)
    curtail_objective = np.zeros(n)
    curtail_objective[slices["curtail"]] = 1.0
    objectives.append(curtail_objective)
    cost_objective = np.zeros(n)
    cost_objective[slices["alk"]] = 1000.0 * om_costs["alk"]
    cost_objective[slices["pem"]] = 1000.0 * om_costs["pem"]
    cost_objective[slices["ammonia"]] = 1000.0 * om_costs["ammonia"]
    cost_objective[slices["discharge"]] = 1000.0 * om_costs["battery"]
    objectives.append(cost_objective)
    return objectives, integrality, Bounds(lower, upper), constraint, slices


def solve_storage_case(
    conventional_load,
    wind_power,
    pv_power,
    capacity_mwh,
    c_rate,
    self_loss_hourly,
    eta_charge,
    eta_discharge,
    om_costs,
):
    generation = np.asarray(wind_power, dtype=float) + np.asarray(pv_power, dtype=float)
    objectives, integrality, bounds, constraint, slices = build_storage_milp(
        conventional_load,
        generation,
        capacity_mwh,
        c_rate,
        self_loss_hourly,
        eta_charge,
        eta_discharge,
        om_costs,
    )
    result, objective_values = _solve_lexicographic(objectives, integrality, bounds, constraint)
    x = _clean(result.x)
    values = {name: x[slice_] for name, slice_ in slices.items()}
    values["u_alk"] = np.rint(values["u_alk"]).astype(int)
    values["u_pem"] = np.rint(values["u_pem"]).astype(int)
    values["u_ammonia"] = np.rint(values["u_ammonia"]).astype(int)
    values["battery_mode"] = np.rint(values["battery_mode"]).astype(int)
    values["power_residual"] = (
        generation
        + values["discharge"]
        - np.asarray(conventional_load)
        + values["shed"]
        - values["alk"]
        - values["pem"]
        - values["ammonia"]
        - values["charge"]
        - values["curtail"]
    )
    values["material_residual"] = material_residual(values["alk"], values["pem"], values["ammonia"])
    values["objective_values"] = objective_values
    return values


def build_grid_milp(conventional_load, generation, buy_prices, sell_price, output_t, om_costs):
    conventional_load = np.asarray(conventional_load, dtype=float)
    generation = np.asarray(generation, dtype=float)
    buy_prices = np.asarray(buy_prices, dtype=float)
    names = ("alk", "pem", "ammonia", "buy", "sell", "u_alk", "u_pem", "u_ammonia", "grid_mode")
    slices = _group_slices(names)
    n = 24 * len(names)
    lower = np.zeros(n)
    upper = np.full(n, np.inf)
    upper[slices["alk"]] = ALK_MAX_MW
    upper[slices["pem"]] = PEM_MAX_MW
    upper[slices["ammonia"]] = AMMONIA_MAX_MW
    upper[slices["buy"]] = conventional_load + ALK_MAX_MW + PEM_MAX_MW + AMMONIA_MAX_MW
    upper[slices["sell"]] = generation
    for name in ("u_alk", "u_pem", "u_ammonia", "grid_mode"):
        upper[slices[name]] = 1.0
    integrality = np.zeros(n, dtype=int)
    for name in ("u_alk", "u_pem", "u_ammonia", "grid_mode"):
        integrality[slices[name]] = 1
    rows = []
    cols = []
    vals = []
    row_lower = []
    row_upper = []
    row = 0

    def add(entries, lo, hi):
        nonlocal row
        for col, value in entries:
            rows.append(row)
            cols.append(col)
            vals.append(value)
        row_lower.append(lo)
        row_upper.append(hi)
        row += 1

    for t in range(24):
        add(
            [
                (slices["alk"].start + t, 1.0),
                (slices["pem"].start + t, 1.0),
                (slices["ammonia"].start + t, 1.0),
                (slices["buy"].start + t, -1.0),
                (slices["sell"].start + t, 1.0),
            ],
            generation[t] - conventional_load[t],
            generation[t] - conventional_load[t],
        )
        add(
            [
                (slices["alk"].start + t, 14.0),
                (slices["pem"].start + t, 16.0),
                (slices["ammonia"].start + t, -400.0),
            ],
            0.0,
            0.0,
        )
        for power_name, state_name, minimum, maximum in (
            ("alk", "u_alk", 2.0, ALK_MAX_MW),
            ("pem", "u_pem", 2.0, PEM_MAX_MW),
            ("ammonia", "u_ammonia", 0.15, AMMONIA_MAX_MW),
        ):
            add([(slices[power_name].start + t, 1.0), (slices[state_name].start + t, -minimum)], 0.0, np.inf)
            add([(slices[power_name].start + t, 1.0), (slices[state_name].start + t, -maximum)], -np.inf, 0.0)
        add(
            [(slices["buy"].start + t, 1.0), (slices["grid_mode"].start + t, -(conventional_load[t] + 41.5))],
            -np.inf,
            0.0,
        )
        add(
            [(slices["sell"].start + t, 1.0), (slices["grid_mode"].start + t, generation[t])],
            -np.inf,
            generation[t],
        )
    add([(slices["ammonia"].start + t, 1.0) for t in range(24)], 0.5 * output_t, 0.5 * output_t)
    matrix = coo_matrix((vals, (rows, cols)), shape=(row, n)).tocsr()
    constraint = LinearConstraint(matrix, np.asarray(row_lower), np.asarray(row_upper))
    objective = np.zeros(n)
    objective[slices["alk"]] = 1000.0 * om_costs["alk"]
    objective[slices["pem"]] = 1000.0 * om_costs["pem"]
    objective[slices["ammonia"]] = 1000.0 * om_costs["ammonia"]
    objective[slices["buy"]] = 1000.0 * buy_prices
    objective[slices["sell"]] = -1000.0 * sell_price
    return objective, integrality, Bounds(lower, upper), constraint, slices


def solve_grid_same_output(
    conventional_load,
    wind_power,
    pv_power,
    buy_prices,
    sell_price,
    output_t,
    om_costs,
):
    generation = np.asarray(wind_power) + np.asarray(pv_power)
    objective, integrality, bounds, constraint, slices = build_grid_milp(
        conventional_load, generation, buy_prices, sell_price, output_t, om_costs
    )
    result = _run_milp(objective, integrality, bounds, constraint)
    x = _clean(result.x)
    values = {name: x[slice_] for name, slice_ in slices.items()}
    for name in ("u_alk", "u_pem", "u_ammonia", "grid_mode"):
        values[name] = np.rint(values[name]).astype(int)
    values["power_residual"] = (
        generation
        + values["buy"]
        - np.asarray(conventional_load)
        - values["alk"]
        - values["pem"]
        - values["ammonia"]
        - values["sell"]
    )
    values["material_residual"] = material_residual(values["alk"], values["pem"], values["ammonia"])
    return values


def daily_cost_components(
    data,
    wind_power,
    pv_power,
    alk,
    pem,
    ammonia,
    battery_discharge=None,
    battery_capacity_mwh=0.0,
    buy=None,
    sell=None,
):
    wind_power = np.asarray(wind_power, dtype=float)
    pv_power = np.asarray(pv_power, dtype=float)
    alk = np.asarray(alk, dtype=float)
    pem = np.asarray(pem, dtype=float)
    ammonia = np.asarray(ammonia, dtype=float)
    battery_discharge = np.zeros(HOURS) if battery_discharge is None else np.asarray(battery_discharge, dtype=float)
    buy = np.zeros(HOURS) if buy is None else np.asarray(buy, dtype=float)
    sell = np.zeros(HOURS) if sell is None else np.asarray(sell, dtype=float)
    wind_cost = 1000.0 * data["wind_cost"] * float(np.sum(wind_power))
    pv_cost = 1000.0 * data["pv_cost"] * float(np.sum(pv_power))
    alk_om = 1000.0 * data["alk_om"] * float(np.sum(alk))
    pem_om = 1000.0 * data["pem_om"] * float(np.sum(pem))
    ammonia_om = 1000.0 * data["ammonia_om"] * float(np.sum(ammonia))
    ammonia_capital = (
        data["ammonia_investment"]
        * HYDROGEN_RATE_KG_PER_H
        / data["ammonia_lifetime"]
        / 365.0
    )
    battery_om = 1000.0 * data["battery_om"] * float(np.sum(battery_discharge))
    battery_capital = battery_daily_fixed_cost(
        battery_capacity_mwh,
        data["battery_investment"],
        data["battery_lifetime"],
    )
    buy_cost = 1000.0 * float(np.sum(buy * data["buy_prices"]))
    sell_revenue = 1000.0 * data["sell_price"] * float(np.sum(sell))
    total = (
        wind_cost
        + pv_cost
        + alk_om
        + pem_om
        + ammonia_om
        + ammonia_capital
        + battery_om
        + battery_capital
        + buy_cost
        - sell_revenue
    )
    return {
        "风电发电成本(¥/day)": wind_cost,
        "光伏发电成本(¥/day)": pv_cost,
        "碱性电解槽运维成本(¥/day)": alk_om,
        "PEM电解槽运维成本(¥/day)": pem_om,
        "合成氨装置运维成本(¥/day)": ammonia_om,
        "合成氨装置资本分摊(¥/day)": ammonia_capital,
        "储能运维成本(¥/day)": battery_om,
        "储能资本分摊(¥/day)": battery_capital,
        "购电成本(¥/day)": buy_cost,
        "售电收入(¥/day)": sell_revenue,
        "日净成本(¥/day)": total,
    }


def scenario_iter(data):
    for wind_index, wind_power in enumerate(data["wind_scenarios"], start=1):
        for pv_index, pv_power in enumerate(data["pv_scenarios"], start=1):
            yield {
                "场景": f"风电{wind_index}-光伏{pv_index}",
                "风电场景": wind_index,
                "光伏场景": pv_index,
                "wind": wind_power,
                "pv": pv_power,
            }


def summarize_dispatch(data, scenario, dispatch, battery_capacity_mwh=0.0, mode="离网无储能"):
    wind = np.asarray(scenario["wind"], dtype=float)
    pv = np.asarray(scenario["pv"], dtype=float)
    conventional = data["conventional_load"]
    output_t = 2.0 * float(np.sum(dispatch["ammonia"]))
    curtail = np.asarray(dispatch.get("curtail", np.zeros(HOURS)), dtype=float)
    shed = np.asarray(dispatch.get("shed", np.zeros(HOURS)), dtype=float)
    charge = np.asarray(dispatch.get("charge", np.zeros(HOURS)), dtype=float)
    discharge = np.asarray(dispatch.get("discharge", np.zeros(HOURS)), dtype=float)
    buy = np.asarray(dispatch.get("buy", np.zeros(HOURS)), dtype=float)
    sell = np.asarray(dispatch.get("sell", np.zeros(HOURS)), dtype=float)
    costs = daily_cost_components(
        data,
        wind,
        pv,
        dispatch["alk"],
        dispatch["pem"],
        dispatch["ammonia"],
        battery_discharge=discharge,
        battery_capacity_mwh=battery_capacity_mwh,
        buy=buy,
        sell=sell,
    )
    generation_energy = float(np.sum(wind + pv))
    conventional_energy = float(np.sum(conventional))
    curtail_energy = float(np.sum(curtail))
    shed_energy = float(np.sum(shed))
    summary = {
        "运行模式": mode,
        "场景": scenario["场景"],
        "风电场景": scenario["风电场景"],
        "光伏场景": scenario["光伏场景"],
        "日氨产量(t/day)": output_t,
        "新能源发电量(MWh)": generation_energy,
        "弃电量(MWh)": curtail_energy,
        "弃电率(%)": 100.0 * curtail_energy / generation_energy if generation_energy > 0 else 0.0,
        "常规负荷失供量(MWh)": shed_energy,
        "能源自治率(%)": 100.0 * (1.0 - shed_energy / conventional_energy),
        "新能源利用率(%)": 100.0 * (1.0 - curtail_energy / generation_energy) if generation_energy > 0 else 100.0,
        "碱性电解槽利用率(%)": 100.0 * float(np.sum(dispatch["alk"])) / (HOURS * ALK_MAX_MW),
        "PEM电解槽利用率(%)": 100.0 * float(np.sum(dispatch["pem"])) / (HOURS * PEM_MAX_MW),
        "制氢装置平均利用率(%)": 100.0 * float(np.sum(dispatch["alk"] + dispatch["pem"])) / (HOURS * (ALK_MAX_MW + PEM_MAX_MW)),
        "合成氨装置利用率(%)": 100.0 * float(np.sum(dispatch["ammonia"])) / (HOURS * AMMONIA_MAX_MW),
        "储能容量(MWh)": battery_capacity_mwh,
        "充电量(MWh)": float(np.sum(charge)),
        "放电量(MWh)": float(np.sum(discharge)),
        "购电量(MWh)": float(np.sum(buy)),
        "售电量(MWh)": float(np.sum(sell)),
    }
    summary.update(costs)
    summary["净吨氨成本(¥/t)"] = costs["日净成本(¥/day)"] / output_t if output_t > 1e-9 else np.nan
    return summary


def hourly_frame(data, scenario, dispatch, battery_capacity_mwh=0.0, mode="离网无储能"):
    rows = []
    zeros = np.zeros(HOURS)
    energy = np.asarray(dispatch.get("energy", zeros), dtype=float)
    for t in range(HOURS):
        rows.append(
            {
                "运行模式": mode,
                "场景": scenario["场景"],
                "风电场景": scenario["风电场景"],
                "光伏场景": scenario["光伏场景"],
                "时段序号(h)": t + 1,
                "时段": data["time_labels"][t],
                "常规负荷功率(MW)": data["conventional_load"][t],
                "风电功率(MW)": scenario["wind"][t],
                "光伏功率(MW)": scenario["pv"][t],
                "碱性电解槽功率(MW)": dispatch["alk"][t],
                "PEM电解槽功率(MW)": dispatch["pem"][t],
                "合成氨装置功率(MW)": dispatch["ammonia"][t],
                "常规负荷失供功率(MW)": dispatch.get("shed", zeros)[t],
                "弃电功率(MW)": dispatch.get("curtail", zeros)[t],
                "储能充电功率(MW)": dispatch.get("charge", zeros)[t],
                "储能放电功率(MW)": dispatch.get("discharge", zeros)[t],
                "储能电量(MWh)": energy[t],
                "储能SOC(%)": 100.0 * energy[t] / battery_capacity_mwh if battery_capacity_mwh > 0 else 0.0,
                "购电功率(MW)": dispatch.get("buy", zeros)[t],
                "售电功率(MW)": dispatch.get("sell", zeros)[t],
            }
        )
    return rows


def solve_no_storage_scenario(data, scenario):
    values = {name: [] for name in ("alk", "pem", "ammonia", "shed", "curtail", "u_alk", "u_pem", "u_ammonia")}
    for t in range(HOURS):
        result = solve_offgrid_hour(
            data["conventional_load"][t],
            scenario["wind"][t],
            scenario["pv"][t],
            data["wind_cost"],
            data["pv_cost"],
            data["alk_om"],
            data["pem_om"],
            data["ammonia_om"],
        )
        for name in values:
            values[name].append(result[name])
    dispatch = {name: np.asarray(value) for name, value in values.items()}
    generation = scenario["wind"] + scenario["pv"]
    dispatch["power_residual"] = (
        generation
        - data["conventional_load"]
        + dispatch["shed"]
        - dispatch["alk"]
        - dispatch["pem"]
        - dispatch["ammonia"]
        - dispatch["curtail"]
    )
    dispatch["material_residual"] = material_residual(dispatch["alk"], dispatch["pem"], dispatch["ammonia"])
    summary = summarize_dispatch(data, scenario, dispatch)
    return summary, hourly_frame(data, scenario, dispatch), dispatch


def capacity_scale_grid(lower, upper, points):
    if math.isclose(lower, upper, rel_tol=0.0, abs_tol=1e-12):
        return np.asarray([float(upper)])
    return np.linspace(lower, upper, points)


def autonomy_capacity_frontier(data, points=401):
    constraints = []
    demand = data["conventional_load"] + ALK_MAX_MW + PEM_MAX_MW + AMMONIA_MAX_MW
    for scenario in scenario_iter(data):
        for t in range(HOURS):
            constraints.append((scenario["wind"][t], scenario["pv"][t], demand[t]))
    no_solar_requirements = []
    for wind, pv, required in constraints:
        if pv <= 1e-12:
            if wind <= 1e-12:
                raise RuntimeError("Strict autonomy is impossible because wind and PV are both zero")
            no_solar_requirements.append(required / wind)
    wind_lower = max(no_solar_requirements, default=0.0)
    if any(wind <= 1e-12 for wind, _, _ in constraints):
        raise RuntimeError("Pure-wind strict autonomy is impossible in at least one hour")
    pure_wind = max(required / wind for wind, _, required in constraints)
    wind_scales = capacity_scale_grid(wind_lower, pure_wind, points)
    annual_wind_energy = SCENARIO_DAYS * sum(float(np.sum(scenario["wind"])) for scenario in scenario_iter(data))
    annual_pv_energy = SCENARIO_DAYS * sum(float(np.sum(scenario["pv"])) for scenario in scenario_iter(data))
    rows = []
    for wind_scale in wind_scales:
        pv_scale = 0.0
        feasible = True
        for wind, pv, required in constraints:
            residual = required - wind_scale * wind
            if residual <= 1e-10:
                continue
            if pv <= 1e-12:
                feasible = False
                break
            pv_scale = max(pv_scale, residual / pv)
        if not feasible:
            continue
        maximum_shortage = max(required - wind_scale * wind - pv_scale * pv for wind, pv, required in constraints)
        rows.append(
            {
                "风电缩放系数": wind_scale,
                "光伏缩放系数": pv_scale,
                "风电装机容量(MW)": WIND_CAPACITY_MW * wind_scale,
                "光伏装机容量(MW)": PV_CAPACITY_MW * pv_scale,
                "总装机容量(MW)": WIND_CAPACITY_MW * wind_scale + PV_CAPACITY_MW * pv_scale,
                "年风光供电成本(¥/year)": 1000.0
                * (
                    data["wind_cost"] * wind_scale * annual_wind_energy
                    + data["pv_cost"] * pv_scale * annual_pv_energy
                ),
                "最大严格自治缺口(MW)": maximum_shortage,
            }
        )
    frontier = pd.DataFrame(rows)
    if frontier.empty:
        raise RuntimeError("No strict-autonomy wind-PV capacity pair found")
    frontier["最小总装机点"] = False
    frontier["最小年供电成本点"] = False
    frontier.loc[frontier["总装机容量(MW)"].idxmin(), "最小总装机点"] = True
    frontier.loc[frontier["年风光供电成本(¥/year)"].idxmin(), "最小年供电成本点"] = True
    return frontier


def evaluate_storage_capacity(data, scenario, capacity_mwh, c_rate=MAIN_C_RATE, daily_self_loss=DAILY_SELF_LOSS):
    dispatch = solve_storage_case(
        data["conventional_load"],
        scenario["wind"],
        scenario["pv"],
        capacity_mwh,
        c_rate,
        hourly_self_loss(daily_self_loss),
        data["eta_charge"],
        data["eta_discharge"],
        {
            "alk": data["alk_om"],
            "pem": data["pem_om"],
            "ammonia": data["ammonia_om"],
            "battery": data["battery_om"],
        },
    )
    summary = summarize_dispatch(
        data,
        scenario,
        dispatch,
        battery_capacity_mwh=capacity_mwh,
        mode="离网配置储能",
    )
    summary["C-rate"] = c_rate
    summary["日自损耗率(%)"] = 100.0 * daily_self_loss
    return summary, dispatch


def select_capacity_result(frame, cost_multiplier=1.0):
    adjusted = frame.copy()
    adjusted["调整后日净成本(¥/day)"] = (
        adjusted["日净成本(¥/day)"]
        - adjusted["储能资本分摊(¥/day)"]
        + cost_multiplier * adjusted["储能资本分摊(¥/day)"]
    )
    adjusted["调整后净吨氨成本(¥/t)"] = adjusted["调整后日净成本(¥/day)"] / adjusted["日氨产量(t/day)"].replace(0.0, np.nan)
    minimum_shed = float(adjusted["常规负荷失供量(MWh)"].min())
    admissible = adjusted[adjusted["常规负荷失供量(MWh)"] <= minimum_shed + 1e-5]
    return adjusted.loc[admissible["调整后净吨氨成本(¥/t)"].idxmin()]


def search_storage_capacity(
    data,
    scenario,
    c_rate=MAIN_C_RATE,
    daily_self_loss=DAILY_SELF_LOSS,
    coarse_count=25,
    refine_count=11,
):
    baseline, _ = evaluate_storage_capacity(data, scenario, 0.0, c_rate, daily_self_loss)
    upper = max(10.0, 1.25 * baseline["弃电量(MWh)"] / 0.8)
    cache = {}

    def evaluate(capacity):
        key = round(max(0.0, float(capacity)), 8)
        if key not in cache:
            cache[key] = evaluate_storage_capacity(data, scenario, key, c_rate, daily_self_loss)
        return cache[key]

    for _ in range(4):
        capacities = np.linspace(0.0, upper, coarse_count)
        for capacity in capacities:
            evaluate(capacity)
        coarse_frame = pd.DataFrame([summary for summary, _ in cache.values()])
        selected = select_capacity_result(coarse_frame)
        at_upper = selected["储能容量(MWh)"] >= upper - upper / max(coarse_count - 1, 1)
        if selected["常规负荷失供量(MWh)"] <= 1e-5 or not at_upper:
            break
        upper *= 2.0
    coarse_step = upper / max(coarse_count - 1, 1)
    selected = select_capacity_result(pd.DataFrame([summary for summary, _ in cache.values()]))
    center = float(selected["储能容量(MWh)"])
    lower_refine = max(0.0, center - coarse_step)
    upper_refine = center + coarse_step
    for capacity in np.linspace(lower_refine, upper_refine, refine_count):
        evaluate(capacity)
    frame = pd.DataFrame([summary for summary, _ in cache.values()]).sort_values("储能容量(MWh)").reset_index(drop=True)
    selected = select_capacity_result(frame)
    best_capacity = float(selected["储能容量(MWh)"])
    best_summary, best_dispatch = cache[round(best_capacity, 8)]
    frame["是否最优容量"] = np.isclose(frame["储能容量(MWh)"], best_capacity, atol=1e-8)
    return frame, best_summary, best_dispatch


def aggregate_annual(summary, label):
    annual_cost = SCENARIO_DAYS * float(summary["日净成本(¥/day)"].sum())
    annual_production = SCENARIO_DAYS * float(summary["日氨产量(t/day)"].sum())
    generation = float(summary["新能源发电量(MWh)"].sum())
    curtail = float(summary["弃电量(MWh)"].sum())
    conventional_energy = 24.0 * float(np.mean(summary["常规负荷失供量(MWh)"].apply(lambda _: 0.0)))
    del conventional_energy
    return pd.DataFrame(
        [
            {
                "运行模式": label,
                "年氨产量(t/year)": annual_production,
                "年综合净成本(¥/year)": annual_cost,
                "综合净吨氨成本(¥/t)": annual_cost / annual_production if annual_production > 0 else np.nan,
                "全年弃电量(MWh/year)": SCENARIO_DAYS * curtail,
                "全年负荷失供量(MWh/year)": SCENARIO_DAYS * float(summary["常规负荷失供量(MWh)"].sum()),
                "全年购电量(MWh/year)": SCENARIO_DAYS * float(summary["购电量(MWh)"].sum()),
                "全年售电量(MWh/year)": SCENARIO_DAYS * float(summary["售电量(MWh)"].sum()),
                "新能源利用率(%)": 100.0 * (1.0 - curtail / generation) if generation > 0 else 100.0,
                "平均能源自治率(%)": float(summary["能源自治率(%)"].mean()),
                "制氢装置平均利用率(%)": float(summary["制氢装置平均利用率(%)"].mean()),
                "合成氨装置平均利用率(%)": float(summary["合成氨装置利用率(%)"].mean()),
            }
        ]
    )


def storage_dynamic_residual(dispatch, capacity_mwh, self_loss_hourly, eta_charge, eta_discharge):
    if capacity_mwh <= 0.0:
        return np.zeros(HOURS)
    energy = np.asarray(dispatch["energy"])
    previous = np.roll(energy, 1)
    return energy - (
        (1.0 - self_loss_hourly) * previous
        + eta_charge * np.asarray(dispatch["charge"])
        - np.asarray(dispatch["discharge"]) / eta_discharge
    )


def validate_dispatch(data, scenario, dispatch, mode, capacity_mwh=0.0, c_rate=MAIN_C_RATE):
    power_residual = np.asarray(dispatch["power_residual"], dtype=float)
    material = np.asarray(dispatch["material_residual"], dtype=float)
    semi = []
    for power, state, minimum, maximum in (
        (dispatch["alk"], dispatch["u_alk"], 2.0, ALK_MAX_MW),
        (dispatch["pem"], dispatch["u_pem"], 2.0, PEM_MAX_MW),
        (dispatch["ammonia"], dispatch["u_ammonia"], 0.15, AMMONIA_MAX_MW),
    ):
        semi.extend(np.maximum(minimum * state - power, 0.0))
        semi.extend(np.maximum(power - maximum * state, 0.0))
    simultaneous = 0.0
    soc_violation = 0.0
    dynamic = 0.0
    if "charge" in dispatch:
        simultaneous = float(np.max(np.minimum(dispatch["charge"], dispatch["discharge"])))
        if capacity_mwh > 0.0:
            low = np.maximum(0.1 * capacity_mwh - dispatch["energy"], 0.0)
            high = np.maximum(dispatch["energy"] - 0.9 * capacity_mwh, 0.0)
            power_high = np.maximum(dispatch["charge"] - c_rate * capacity_mwh, 0.0)
            power_high = np.maximum(power_high, np.maximum(dispatch["discharge"] - c_rate * capacity_mwh, 0.0))
            soc_violation = float(max(np.max(low), np.max(high), np.max(power_high)))
            dynamic = float(
                np.max(
                    np.abs(
                        storage_dynamic_residual(
                            dispatch,
                            capacity_mwh,
                            hourly_self_loss(data["daily_self_loss"]),
                            data["eta_charge"],
                            data["eta_discharge"],
                        )
                    )
                )
            )
    return {
        "运行模式": mode,
        "场景": scenario["场景"],
        "最大功率平衡残差(MW)": float(np.max(np.abs(power_residual))),
        "最大氢氨物料平衡残差(kg/h)": float(np.max(np.abs(material))),
        "最大半连续边界违反(MW)": float(max(semi, default=0.0)),
        "最大同时充放电功率(MW)": simultaneous,
        "最大SOC及倍率边界违反(MWh)": soc_violation,
        "最大储能动态残差(MWh)": dynamic,
    }


def solve_grid_comparison(data, scenario, output_t):
    dispatch = solve_grid_same_output(
        data["conventional_load"],
        scenario["wind"],
        scenario["pv"],
        data["buy_prices"],
        data["sell_price"],
        output_t,
        {"alk": data["alk_om"], "pem": data["pem_om"], "ammonia": data["ammonia_om"]},
    )
    summary = summarize_dispatch(data, scenario, dispatch, mode="并网同产量")
    return summary, hourly_frame(data, scenario, dispatch, mode="并网同产量"), dispatch


def write_excel(path, sheets):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="5271AE")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 30)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    workbook.save(path)


def save_figure(fig, directory, stem):
    fig.savefig(directory / f"{stem}.png", dpi=300, bbox_inches="tight")
    fig.savefig(directory / f"{stem}.pdf", bbox_inches="tight")
    plt.close(fig)


def _heatmap(ax, matrix, xlabel, ylabel, fmt, cmap):
    image = ax.imshow(matrix, aspect="auto", cmap=cmap)
    ax.set_xticks(range(4), [f"光伏{i}" for i in range(1, 5)])
    ax.set_yticks(range(6), [f"风电{i}" for i in range(1, 7)])
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    threshold = float(np.nanmean(matrix))
    for i in range(6):
        for j in range(4):
            value = matrix[i, j]
            ax.text(j, i, format(value, fmt), ha="center", va="center", fontsize=8, color="white" if value > threshold else "black")
    return image


def generate_figures(
    no_storage_summary,
    no_storage_hourly,
    frontier,
    capacity_search,
    storage_summary,
    storage_hourly,
    comparison,
    grid_summary,
    max_curtail_scenario,
    figure_dir,
):
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    cmap = LinearSegmentedColormap.from_list(
        "soft",
        [PALETTE["blue"], PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"]],
    )
    fig, axes = plt.subplots(1, 3, figsize=(14, 4.6))
    for ax, column, label, fmt in (
        (axes[0], "日氨产量(t/day)", "日氨产量 / t/day", ".1f"),
        (axes[1], "弃电率(%)", "弃电率 / %", ".1f"),
        (axes[2], "净吨氨成本(¥/t)", "净吨氨成本 / ¥/t", ".0f"),
    ):
        matrix = no_storage_summary.pivot(index="风电场景", columns="光伏场景", values=column).to_numpy()
        image = _heatmap(ax, matrix, "光伏场景", "风电场景", fmt, cmap)
        ax.set_title(label)
        fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04)
    fig.tight_layout()
    save_figure(fig, figure_dir, "问题四无储能场景运行热力图")

    fig, ax = plt.subplots(figsize=(8.5, 5.4))
    if len(frontier) == 1:
        point = frontier.iloc[0]
        ax.bar(
            ["风电", "光伏"],
            [point["风电装机容量(MW)"], point["光伏装机容量(MW)"]],
            color=[PALETTE["blue"], PALETTE["yellow"]],
            width=0.55,
        )
        ax.text(0, point["风电装机容量(MW)"] * 1.01, f"{point['风电装机容量(MW)']:.1f}", ha="center")
        ax.text(
            0.5,
            0.88,
            "严格72 t/day约束下，夜间无光伏\n容量边界退化为纯风电点",
            transform=ax.transAxes,
            ha="center",
            va="center",
            color=PALETTE["red"],
        )
        ax.set_ylabel("装机容量 / MW")
        ax.grid(axis="y", alpha=0.25)
    else:
        ax.plot(frontier["风电装机容量(MW)"], frontier["光伏装机容量(MW)"], color=PALETTE["blue"], linewidth=2.2)
        for flag, color, label in (
            ("最小总装机点", PALETTE["red"], "最小总装机"),
            ("最小年供电成本点", PALETTE["orange"], "最小年供电成本"),
        ):
            point = frontier[frontier[flag]].iloc[0]
            ax.scatter(point["风电装机容量(MW)"], point["光伏装机容量(MW)"], s=80, color=color, label=label, zorder=3)
        ax.set_xlabel("风电装机容量 / MW")
        ax.set_ylabel("光伏装机容量 / MW")
        ax.grid(alpha=0.25)
        ax.legend()
    save_figure(fig, figure_dir, "问题四能源自治风光容量边界")

    before = no_storage_hourly[no_storage_hourly["场景"] == max_curtail_scenario].sort_values("时段序号(h)")
    after = storage_hourly[storage_hourly["场景"] == max_curtail_scenario].sort_values("时段序号(h)")
    hours = before["时段序号(h)"].to_numpy()
    fig, axes = plt.subplots(3, 1, figsize=(11, 8.5), sharex=True, gridspec_kw={"height_ratios": [2, 1, 1]})
    generation = before["风电功率(MW)"].to_numpy() + before["光伏功率(MW)"].to_numpy()
    process_before = before["碱性电解槽功率(MW)"].to_numpy() + before["PEM电解槽功率(MW)"].to_numpy() + before["合成氨装置功率(MW)"].to_numpy()
    process_after = after["碱性电解槽功率(MW)"].to_numpy() + after["PEM电解槽功率(MW)"].to_numpy() + after["合成氨装置功率(MW)"].to_numpy()
    axes[0].plot(hours, generation, color=PALETTE["blue"], linewidth=2.3, label="风光总功率")
    axes[0].plot(hours, process_before, color=PALETTE["red"], linestyle="--", linewidth=1.8, label="储能前生产功率")
    axes[0].plot(hours, process_after, color=PALETTE["orange"], linewidth=2.0, label="储能后生产功率")
    axes[0].set_ylabel("功率 / MW")
    axes[0].legend(ncol=3)
    axes[0].grid(alpha=0.25)
    axes[1].bar(hours, after["储能放电功率(MW)"], color=PALETTE["orange"], label="放电")
    axes[1].bar(hours, -after["储能充电功率(MW)"], color=PALETTE["blue"], label="充电")
    axes[1].axhline(0, color="#555555", linewidth=0.8)
    axes[1].set_ylabel("储能功率 / MW")
    axes[1].legend(ncol=2)
    axes[1].grid(axis="y", alpha=0.25)
    axes[2].plot(hours, after["储能SOC(%)"], color=PALETTE["red"], linewidth=2.0)
    axes[2].axhline(10, color="#777777", linestyle="--", linewidth=0.8)
    axes[2].axhline(90, color="#777777", linestyle="--", linewidth=0.8)
    axes[2].set_ylabel("SOC / %")
    axes[2].set_xlabel("时段 / h")
    axes[2].set_xticks(np.arange(1, 25, 2))
    axes[2].grid(alpha=0.25)
    save_figure(fig, figure_dir, "问题四最大弃电场景储能前后调度")

    fig, ax1 = plt.subplots(figsize=(9, 5.4))
    ax1.plot(capacity_search["储能容量(MWh)"], capacity_search["净吨氨成本(¥/t)"], color=PALETTE["red"], marker="o", markersize=3, label="净吨氨成本")
    ax1.set_xlabel("储能容量 / MWh")
    ax1.set_ylabel("净吨氨成本 / ¥/t", color=PALETTE["red"])
    ax2 = ax1.twinx()
    ax2.plot(capacity_search["储能容量(MWh)"], capacity_search["弃电率(%)"], color=PALETTE["blue"], linewidth=2.0, label="弃电率")
    ax2.set_ylabel("弃电率 / %", color=PALETTE["blue"])
    best = capacity_search[capacity_search["是否最优容量"]].iloc[0]
    ax1.axvline(best["储能容量(MWh)"], color=PALETTE["orange"], linestyle="--", label="最优容量")
    ax1.grid(alpha=0.25)
    lines = ax1.get_lines() + ax2.get_lines()
    ax1.legend(lines, [line.get_label() for line in lines], loc="best")
    save_figure(fig, figure_dir, "问题四储能容量经济性曲线")

    cost_columns = [
        "风电发电成本(¥/day)",
        "光伏发电成本(¥/day)",
        "碱性电解槽运维成本(¥/day)",
        "PEM电解槽运维成本(¥/day)",
        "合成氨装置运维成本(¥/day)",
        "合成氨装置资本分摊(¥/day)",
        "储能运维成本(¥/day)",
        "储能资本分摊(¥/day)",
        "购电成本(¥/day)",
    ]
    modes = ["离网配置储能", "并网同产量"]
    frames = [storage_summary, grid_summary]
    annual_components = np.asarray([[SCENARIO_DAYS * frame[column].sum() for column in cost_columns] for frame in frames])
    fig, ax = plt.subplots(figsize=(9.5, 5.7))
    bottom = np.zeros(2)
    colors = [PALETTE["blue"], PALETTE["light_blue"], PALETTE["yellow"], PALETTE["orange"], PALETTE["red"], "#8F7AB8", "#69A88D", "#A6A6A6", "#C98B5B"]
    for i, column in enumerate(cost_columns):
        ax.bar(modes, annual_components[:, i], bottom=bottom, color=colors[i], label=column.replace("(¥/day)", ""))
        bottom += annual_components[:, i]
    sell = np.asarray([SCENARIO_DAYS * frame["售电收入(¥/day)"].sum() for frame in frames])
    ax.bar(modes, -sell, color="#4E6E58", label="售电收入")
    ax.set_ylabel("年成本 / ¥")
    ax.grid(axis="y", alpha=0.25)
    ax.legend(ncol=3, fontsize=8, loc="upper center")
    save_figure(fig, figure_dir, "问题四并离网年成本构成")


def frame_to_markdown(frame, float_digits=3):
    def format_value(value):
        if pd.isna(value):
            return ""
        if isinstance(value, (float, np.floating)):
            return f"{float(value):.{float_digits}f}"
        return str(value)

    headers = [str(column).replace("|", "\\|") for column in frame.columns]
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in frame.itertuples(index=False, name=None):
        values = [format_value(value).replace("|", "\\|") for value in row]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def write_report(
    path,
    annual,
    no_storage_summary,
    frontier,
    max_curtail_scenario,
    capacity_search,
    storage_summary,
    comparison,
    sensitivity,
    validation,
):
    no_storage_annual = annual[annual["运行模式"] == "离网无储能"].iloc[0]
    storage_annual = annual[annual["运行模式"] == "离网配置储能"].iloc[0]
    grid_annual = annual[annual["运行模式"] == "并网同产量"].iloc[0]
    min_capacity = frontier[frontier["最小总装机点"]].iloc[0]
    min_cost = frontier[frontier["最小年供电成本点"]].iloc[0]
    best_capacity = capacity_search[capacity_search["是否最优容量"]].iloc[0]
    support = float(storage_annual["年综合净成本(¥/year)"] - grid_annual["年综合净成本(¥/year)"])
    support_t = support / float(storage_annual["年氨产量(t/year)"])
    frontier_explanation = (
        "严格72 t/day要求意味着合成氨装置必须24 h满功率运行；控制性约束出现在无光伏的夜间，故本数据下风光容量Pareto边界退化为纯风电单点。"
        if len(frontier) == 1
        else "风电和光伏之间存在替代关系，因此报告非支配容量边界。"
    )
    text = f"""# 问题四计算结果

## 方法与统一口径

无储能离网调度采用逐时8种设备启停状态枚举，并按“常规负荷失供最小、合成氨产量最大、弃电最小、运行成本最小”执行词典序求解。储能模型采用2 h、0.5C，SOC范围10%至90%，首末SOC自由循环，0.2%/day自损耗换算为小时自损耗。固定投资按365 day/year日化，24种场景各代表15 day，全年按360个典型运行日累计。

## 问题四（1）无储能离网运行

{frame_to_markdown(no_storage_annual.to_frame().T)}

24种场景中最大弃电场景为 **{max_curtail_scenario}**。日制氨量范围为 **{no_storage_summary['日氨产量(t/day)'].min():.3f}–{no_storage_summary['日氨产量(t/day)'].max():.3f} t/day**，场景吨氨成本范围为 **{no_storage_summary['净吨氨成本(¥/t)'].min():.2f}–{no_storage_summary['净吨氨成本(¥/t)'].max():.2f} ¥/t**。

![无储能场景运行热力图](../figures/问题四计算结果/问题四无储能场景运行热力图.png)

### 严格能源自治容量边界

- 最小总装机点：风电 **{min_capacity['风电装机容量(MW)']:.3f} MW**，光伏 **{min_capacity['光伏装机容量(MW)']:.3f} MW**，总装机 **{min_capacity['总装机容量(MW)']:.3f} MW**；
- 最小年供电成本点：风电 **{min_cost['风电装机容量(MW)']:.3f} MW**，光伏 **{min_cost['光伏装机容量(MW)']:.3f} MW**，年风光供电成本 **{min_cost['年风光供电成本(¥/year)']:.2f} ¥**。

第二个点依据度电成本计算，只称为“最小年供电成本点”，不称为最小投资点。

{frontier_explanation}

![能源自治风光容量边界](../figures/问题四计算结果/问题四能源自治风光容量边界.png)

## 问题四（2）储能配置

最大弃电场景的最优储能容量为 **{best_capacity['储能容量(MWh)']:.3f} MWh**，对应日氨产量 **{best_capacity['日氨产量(t/day)']:.3f} t/day**、弃电率 **{best_capacity['弃电率(%)']:.3f}%**、净吨氨成本 **{best_capacity['净吨氨成本(¥/t)']:.2f} ¥/t**。该容量固定应用于全部24种场景后，年度结果如下。

{frame_to_markdown(storage_annual.to_frame().T)}

![最大弃电场景储能前后调度](../figures/问题四计算结果/问题四最大弃电场景储能前后调度.png)

![储能容量经济性曲线](../figures/问题四计算结果/问题四储能容量经济性曲线.png)

### 灵敏度分析

{frame_to_markdown(sensitivity)}

## 问题四（3）并离网经济性

并网模式逐场景锁定为与离网储能模式相同的合成氨产量。两种模式年度结果如下。

{frame_to_markdown(annual[annual['运行模式'].isin(['离网配置储能', '并网同产量'])])}

电网年度系统支撑价值为 **{support:.2f} ¥/year**，单位制氨量支撑价值为 **{support_t:.2f} ¥/t**。该值表示在相同场景产量下，接入主电网后减少的年度净成本，不使用MILP对偶变量进行边际价格解释。

![并离网年成本构成](../figures/问题四计算结果/问题四并离网年成本构成.png)

## 约束校验

{frame_to_markdown(validation.max(numeric_only=True).to_frame('最大值').reset_index(names='校验项'), float_digits=8)}

全部输出均按逐时电力平衡、氢氨物料平衡、半连续功率边界、充放电互斥、SOC边界和储能循环状态重新回代校验。

## 可复现运行方式

```powershell
python -X utf8 code/问题四.py --self-test
python -X utf8 code/问题四.py
```
"""
    path.write_text(text, encoding="utf-8")


def dataframe_records(frame):
    return json.loads(frame.to_json(orient="records", force_ascii=False))


def run_full_model(project_root):
    data = load_inputs(project_root)
    output_dir = project_root / "outputs" / "问题四计算结果"
    figure_dir = project_root / "figures" / "问题四计算结果"
    report_path = project_root / "reports" / "问题四计算结果.md"
    output_dir.mkdir(parents=True, exist_ok=True)
    figure_dir.mkdir(parents=True, exist_ok=True)

    scenarios = list(scenario_iter(data))
    no_storage_summaries = []
    no_storage_hourly_rows = []
    no_storage_dispatches = {}
    for index, scenario in enumerate(scenarios, start=1):
        summary, hourly_rows, dispatch = solve_no_storage_scenario(data, scenario)
        no_storage_summaries.append(summary)
        no_storage_hourly_rows.extend(hourly_rows)
        no_storage_dispatches[scenario["场景"]] = dispatch
        print(f"Solved no-storage scenario {index}/24", flush=True)
    no_storage_summary = pd.DataFrame(no_storage_summaries)
    no_storage_hourly = pd.DataFrame(no_storage_hourly_rows)
    frontier = autonomy_capacity_frontier(data)
    max_curtail_row = no_storage_summary.loc[no_storage_summary["弃电量(MWh)"].idxmax()]
    max_curtail_scenario = str(max_curtail_row["场景"])
    max_scenario = next(scenario for scenario in scenarios if scenario["场景"] == max_curtail_scenario)

    print(f"Searching storage capacity for {max_curtail_scenario}", flush=True)
    capacity_search, best_capacity_summary, best_capacity_dispatch = search_storage_capacity(
        data,
        max_scenario,
        c_rate=MAIN_C_RATE,
        daily_self_loss=data["daily_self_loss"],
    )
    best_capacity = float(best_capacity_summary["储能容量(MWh)"])
    print(f"Selected storage capacity: {best_capacity:.4f} MWh", flush=True)

    storage_summaries = []
    storage_hourly_rows = []
    storage_dispatches = {}
    for index, scenario in enumerate(scenarios, start=1):
        if scenario["场景"] == max_curtail_scenario:
            dispatch = best_capacity_dispatch
            summary = summarize_dispatch(
                data,
                scenario,
                dispatch,
                battery_capacity_mwh=best_capacity,
                mode="离网配置储能",
            )
        else:
            summary, dispatch = evaluate_storage_capacity(
                data,
                scenario,
                best_capacity,
                MAIN_C_RATE,
                data["daily_self_loss"],
            )
        storage_summaries.append(summary)
        storage_hourly_rows.extend(
            hourly_frame(data, scenario, dispatch, battery_capacity_mwh=best_capacity, mode="离网配置储能")
        )
        storage_dispatches[scenario["场景"]] = dispatch
        print(f"Solved storage scenario {index}/24", flush=True)
    storage_summary = pd.DataFrame(storage_summaries)
    storage_hourly = pd.DataFrame(storage_hourly_rows)

    grid_summaries = []
    grid_hourly_rows = []
    grid_dispatches = {}
    for index, scenario in enumerate(scenarios, start=1):
        target = float(storage_summary.loc[storage_summary["场景"] == scenario["场景"], "日氨产量(t/day)"].iloc[0])
        summary, hourly_rows, dispatch = solve_grid_comparison(data, scenario, target)
        grid_summaries.append(summary)
        grid_hourly_rows.extend(hourly_rows)
        grid_dispatches[scenario["场景"]] = dispatch
        print(f"Solved grid comparison {index}/24", flush=True)
    grid_summary = pd.DataFrame(grid_summaries)
    grid_hourly = pd.DataFrame(grid_hourly_rows)

    comparison = storage_summary.merge(
        no_storage_summary,
        on=["场景", "风电场景", "光伏场景"],
        suffixes=("_储能后", "_储能前"),
    )
    comparison["日氨产量增加(t/day)"] = comparison["日氨产量(t/day)_储能后"] - comparison["日氨产量(t/day)_储能前"]
    comparison["弃电量减少(MWh)"] = comparison["弃电量(MWh)_储能前"] - comparison["弃电量(MWh)_储能后"]
    comparison["吨氨成本变化(¥/t)"] = comparison["净吨氨成本(¥/t)_储能后"] - comparison["净吨氨成本(¥/t)_储能前"]
    grid_comparison = storage_summary[
        ["场景", "风电场景", "光伏场景", "日氨产量(t/day)", "日净成本(¥/day)", "净吨氨成本(¥/t)"]
    ].merge(
        grid_summary[["场景", "日氨产量(t/day)", "日净成本(¥/day)", "净吨氨成本(¥/t)", "购电量(MWh)", "售电量(MWh)"]],
        on="场景",
        suffixes=("_离网", "_并网"),
    )
    grid_comparison["产量匹配残差(t/day)"] = grid_comparison["日氨产量(t/day)_并网"] - grid_comparison["日氨产量(t/day)_离网"]
    grid_comparison["电网日支撑价值(¥/day)"] = grid_comparison["日净成本(¥/day)_离网"] - grid_comparison["日净成本(¥/day)_并网"]
    grid_comparison["单位产量电网支撑价值(¥/t)"] = grid_comparison["电网日支撑价值(¥/day)"] / grid_comparison["日氨产量(t/day)_离网"]

    sensitivity_rows = []
    for duration_h, c_rate in ((1.0, 1.0), (2.0, 0.5), (4.0, 0.25)):
        if math.isclose(c_rate, MAIN_C_RATE):
            frame = capacity_search
            selected = select_capacity_result(frame)
        else:
            frame, _, _ = search_storage_capacity(
                data,
                max_scenario,
                c_rate=c_rate,
                daily_self_loss=data["daily_self_loss"],
                coarse_count=15,
                refine_count=7,
            )
            selected = select_capacity_result(frame)
        sensitivity_rows.append(
            {
                "敏感性类型": "储能时长",
                "参数": f"{duration_h:g} h/{c_rate:g}C",
                "最优储能容量(MWh)": selected["储能容量(MWh)"],
                "日氨产量(t/day)": selected["日氨产量(t/day)"],
                "弃电率(%)": selected["弃电率(%)"],
                "净吨氨成本(¥/t)": selected["净吨氨成本(¥/t)"],
            }
        )
    for daily_loss in (0.001, 0.002, 0.003):
        if math.isclose(daily_loss, data["daily_self_loss"]):
            frame = capacity_search
            selected = select_capacity_result(frame)
        else:
            frame, _, _ = search_storage_capacity(
                data,
                max_scenario,
                c_rate=MAIN_C_RATE,
                daily_self_loss=daily_loss,
                coarse_count=15,
                refine_count=7,
            )
            selected = select_capacity_result(frame)
        sensitivity_rows.append(
            {
                "敏感性类型": "日自损耗率",
                "参数": f"{100*daily_loss:.1f}%/day",
                "最优储能容量(MWh)": selected["储能容量(MWh)"],
                "日氨产量(t/day)": selected["日氨产量(t/day)"],
                "弃电率(%)": selected["弃电率(%)"],
                "净吨氨成本(¥/t)": selected["净吨氨成本(¥/t)"],
            }
        )
    for multiplier in (0.8, 1.0, 1.2):
        selected = select_capacity_result(capacity_search, multiplier)
        sensitivity_rows.append(
            {
                "敏感性类型": "储能投资成本",
                "参数": f"{multiplier:.1f}×",
                "最优储能容量(MWh)": selected["储能容量(MWh)"],
                "日氨产量(t/day)": selected["日氨产量(t/day)"],
                "弃电率(%)": selected["弃电率(%)"],
                "净吨氨成本(¥/t)": selected["调整后净吨氨成本(¥/t)"],
            }
        )
    sensitivity = pd.DataFrame(sensitivity_rows)

    annual = pd.concat(
        [
            aggregate_annual(no_storage_summary, "离网无储能"),
            aggregate_annual(storage_summary, "离网配置储能"),
            aggregate_annual(grid_summary, "并网同产量"),
        ],
        ignore_index=True,
    )
    validation_rows = []
    for scenario in scenarios:
        name = scenario["场景"]
        validation_rows.append(validate_dispatch(data, scenario, no_storage_dispatches[name], "离网无储能"))
        validation_rows.append(validate_dispatch(data, scenario, storage_dispatches[name], "离网配置储能", best_capacity, MAIN_C_RATE))
        validation_rows.append(validate_dispatch(data, scenario, grid_dispatches[name], "并网同产量"))
    validation = pd.DataFrame(validation_rows)

    if len(no_storage_summary) != 24 or len(storage_summary) != 24 or len(grid_summary) != 24:
        raise AssertionError("Each operating mode must contain 24 scenarios")
    if validation.select_dtypes(include=[np.number]).max().max() > 1e-5:
        raise AssertionError("Constraint validation failed")
    if grid_comparison["产量匹配残差(t/day)"].abs().max() > 1e-5:
        raise AssertionError("Grid and off-grid production matching failed")
    if frontier["最大严格自治缺口(MW)"].max() > 1e-6:
        raise AssertionError("Strict-autonomy capacity frontier is infeasible")

    frames = {
        "问题四1场景汇总": no_storage_summary,
        "问题四1逐时调度": no_storage_hourly,
        "能源自治容量边界": frontier,
        "储能容量搜索": capacity_search,
        "问题四2场景汇总": storage_summary,
        "问题四2逐时调度": storage_hourly,
        "储能前后比较": comparison,
        "问题四3并网汇总": grid_summary,
        "问题四3并网逐时": grid_hourly,
        "并离网同产量比较": grid_comparison,
        "年度汇总": annual,
        "灵敏度分析": sensitivity,
        "约束校验": validation,
    }
    write_excel(output_dir / "问题四计算结果.xlsx", frames)
    export_frames = {
        "问题四无储能场景汇总.csv": no_storage_summary,
        "问题四无储能逐时调度.csv": no_storage_hourly,
        "问题四能源自治容量边界.csv": frontier,
        "问题四储能容量搜索.csv": capacity_search,
        "问题四储能场景汇总.csv": storage_summary,
        "问题四储能逐时调度.csv": storage_hourly,
        "问题四并离网同产量比较.csv": grid_comparison,
        "问题四年度汇总.csv": annual,
        "问题四灵敏度分析.csv": sensitivity,
        "问题四约束校验.csv": validation,
    }
    for filename, frame in export_frames.items():
        frame.to_csv(output_dir / filename, index=False, encoding="utf-8-sig")
    payload = {
        "最大弃电场景": max_curtail_scenario,
        "最优储能容量(MWh)": best_capacity,
        "年度汇总": dataframe_records(annual),
        "能源自治容量关键点": dataframe_records(frontier[frontier["最小总装机点"] | frontier["最小年供电成本点"]]),
        "灵敏度分析": dataframe_records(sensitivity),
        "最大约束残差": dataframe_records(validation.max(numeric_only=True).to_frame().T),
    }
    (output_dir / "问题四完整计算结果.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    generate_figures(
        no_storage_summary,
        no_storage_hourly,
        frontier,
        capacity_search,
        storage_summary,
        storage_hourly,
        grid_comparison,
        grid_summary,
        max_curtail_scenario,
        figure_dir,
    )
    write_report(
        report_path,
        annual,
        no_storage_summary,
        frontier,
        max_curtail_scenario,
        capacity_search,
        storage_summary,
        grid_comparison,
        sensitivity,
        validation,
    )
    print(f"Maximum-curtailment scenario: {max_curtail_scenario}")
    print(f"Optimal storage capacity: {best_capacity:.6f} MWh")
    print(f"Maximum validation residual: {validation.select_dtypes(include=[np.number]).max().max():.3e}")


def run_self_tests():
    assert np.array_equal(capacity_scale_grid(2.0, 2.0, 401), np.array([2.0]))
    markdown = frame_to_markdown(pd.DataFrame({"甲": [1.2345], "乙": ["正常"]}), float_digits=2)
    assert "| 甲 | 乙 |" in markdown
    assert "| 1.23 | 正常 |" in markdown
    assert math.isclose(hourly_self_loss(0.002), 8.341329886207838e-05, rel_tol=0.0, abs_tol=1e-12)
    assert math.isclose(material_residual(20.0, 20.0, 1.5), 0.0, abs_tol=1e-12)
    assert math.isclose(battery_daily_fixed_cost(1.0), 1000000.0 / (15.0 * 365.0), abs_tol=1e-12)

    balanced = solve_offgrid_hour(
        conventional_load=6.0,
        wind_power=6.0,
        pv_power=0.0,
        wind_cost=0.15,
        pv_cost=0.12,
        alk_om=0.10,
        pem_om=0.15,
        ammonia_om=0.002,
    )
    assert math.isclose(balanced["shed"], 0.0, abs_tol=1e-8)
    assert math.isclose(balanced["ammonia"], 0.0, abs_tol=1e-8)
    assert math.isclose(balanced["curtail"], 0.0, abs_tol=1e-8)

    full = solve_offgrid_hour(
        conventional_load=6.0,
        wind_power=47.5,
        pv_power=0.0,
        wind_cost=0.15,
        pv_cost=0.12,
        alk_om=0.10,
        pem_om=0.15,
        ammonia_om=0.002,
    )
    assert math.isclose(full["alk"], 20.0, abs_tol=1e-7)
    assert math.isclose(full["pem"], 20.0, abs_tol=1e-7)
    assert math.isclose(full["ammonia"], 1.5, abs_tol=1e-7)
    assert math.isclose(full["shed"], 0.0, abs_tol=1e-8)
    assert math.isclose(full["curtail"], 0.0, abs_tol=1e-8)

    storage = solve_storage_case(
        conventional_load=np.full(24, 6.0),
        wind_power=np.full(24, 47.5),
        pv_power=np.zeros(24),
        capacity_mwh=0.0,
        c_rate=0.5,
        self_loss_hourly=hourly_self_loss(0.002),
        eta_charge=0.9,
        eta_discharge=0.9,
        om_costs={"alk": 0.10, "pem": 0.15, "ammonia": 0.002, "battery": 0.01},
    )
    assert math.isclose(float(np.sum(storage["ammonia"])) * 2.0, 72.0, abs_tol=1e-6)
    assert float(np.max(np.abs(storage["power_residual"]))) < 1e-7
    assert float(np.max(np.abs(storage["material_residual"]))) < 1e-7

    grid = solve_grid_same_output(
        conventional_load=np.full(24, 6.0),
        wind_power=np.zeros(24),
        pv_power=np.zeros(24),
        buy_prices=np.full(24, 0.5),
        sell_price=0.3,
        output_t=36.0,
        om_costs={"alk": 0.10, "pem": 0.15, "ammonia": 0.002},
    )
    assert math.isclose(float(np.sum(grid["ammonia"])) * 2.0, 36.0, abs_tol=1e-6)
    assert float(np.max(np.abs(grid["power_residual"]))) < 1e-7
    print("All problem 4 self-tests passed.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        run_self_tests()
        return
    project_root = Path(__file__).resolve().parents[1]
    run_full_model(project_root)


if __name__ == "__main__":
    main()
