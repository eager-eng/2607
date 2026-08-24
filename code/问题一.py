import csv
import json
import re
from pathlib import Path

import matplotlib
import numpy as np
from openpyxl import load_workbook

matplotlib.use("Agg")
import matplotlib.pyplot as plt


HOURS = 24
TIME_STEP_H = 1.0
CONVENTIONAL_LOAD_BASE_MW = 6.0
WIND_CAPACITY_MW = 40.0
PV_CAPACITY_MW = 64.0
ALK_POWER_MW = 10.0
PEM_POWER_MW = 10.0
AMMONIA_POWER_MW = 0.75
PALETTE = {
    "blue": "#5271AE",
    "light_blue": "#70ACDE",
    "yellow": "#F5CC7D",
    "orange": "#FFA660",
    "red": "#D85B59",
}


def calculate_hydrogen_output(power_mw, efficiency):
    return 1000.0 * np.asarray(power_mw, dtype=float) * efficiency / 50.0


def calculate_grid_exchange(load_mw, generation_mw):
    net_load = np.asarray(load_mw, dtype=float) - np.asarray(generation_mw, dtype=float)
    return np.maximum(net_load, 0.0), np.maximum(-net_load, 0.0)


def calculate_green_metrics(
    load_energy_mwh,
    generation_energy_mwh,
    buy_energy_mwh,
    sell_energy_mwh,
):
    if generation_energy_mwh <= 0 or load_energy_mwh <= 0:
        raise ValueError("Energy denominators must be positive")
    return {
        "新能源自发自用比例": (
            load_energy_mwh - sell_energy_mwh - buy_energy_mwh
        )
        / generation_energy_mwh,
        "新能源供电占比": (generation_energy_mwh - sell_energy_mwh)
        / load_energy_mwh,
        "上网电量比例": sell_energy_mwh / generation_energy_mwh,
    }


def calculate_daily_cost(
    wind_energy_mwh,
    pv_energy_mwh,
    buy_energy_mwh,
    buy_price_yuan_per_kwh,
    sell_energy_mwh,
    sell_price_yuan_per_kwh,
    alk_energy_mwh,
    pem_energy_mwh,
    ammonia_energy_mwh,
    rated_hydrogen_demand_kg_per_h,
    ammonia_output_t,
    wind_cost_yuan_per_kwh=0.15,
    pv_cost_yuan_per_kwh=0.12,
    alk_om_yuan_per_kwh=0.10,
    pem_om_yuan_per_kwh=0.15,
    ammonia_om_yuan_per_kwh=0.002,
    ammonia_investment_yuan_per_kg_h2=60000.0,
    ammonia_lifetime_years=30.0,
):
    buy_energy = np.asarray(buy_energy_mwh, dtype=float)
    buy_price = np.asarray(buy_price_yuan_per_kwh, dtype=float)
    sell_energy = np.asarray(sell_energy_mwh, dtype=float)
    sell_price = np.asarray(sell_price_yuan_per_kwh, dtype=float)
    if buy_energy.shape != buy_price.shape or sell_energy.shape != sell_price.shape:
        raise ValueError("Energy and price arrays must have matching shapes")
    if ammonia_output_t <= 0:
        raise ValueError("Ammonia output must be positive")

    wind_cost = wind_energy_mwh * 1000.0 * wind_cost_yuan_per_kwh
    pv_cost = pv_energy_mwh * 1000.0 * pv_cost_yuan_per_kwh
    buy_cost = float(np.sum(buy_energy * buy_price) * 1000.0)
    sell_revenue = float(np.sum(sell_energy * sell_price) * 1000.0)
    alk_om = alk_energy_mwh * 1000.0 * alk_om_yuan_per_kwh
    pem_om = pem_energy_mwh * 1000.0 * pem_om_yuan_per_kwh
    ammonia_om = ammonia_energy_mwh * 1000.0 * ammonia_om_yuan_per_kwh
    ammonia_capital = (
        ammonia_investment_yuan_per_kg_h2
        * rated_hydrogen_demand_kg_per_h
        / ammonia_lifetime_years
        / 365.0
    )
    net_cost = (
        wind_cost
        + pv_cost
        + buy_cost
        - sell_revenue
        + alk_om
        + pem_om
        + ammonia_om
        + ammonia_capital
    )
    return {
        "风电发电成本": float(wind_cost),
        "光伏发电成本": float(pv_cost),
        "购电成本": float(buy_cost),
        "售电收入": float(sell_revenue),
        "碱性电解槽运维成本": float(alk_om),
        "PEM电解槽运维成本": float(pem_om),
        "合成氨装置运维成本": float(ammonia_om),
        "合成氨装置资本分摊": float(ammonia_capital),
        "日净成本": float(net_cost),
        "净吨氨成本": float(net_cost / ammonia_output_t),
    }


def _find_attachment(project_root, prefix):
    matches = [
        path
        for path in Path(project_root).joinpath("题目").rglob("*.xlsx")
        if not path.name.startswith("~$") and path.name.startswith(prefix)
    ]
    if len(matches) != 1:
        raise FileNotFoundError(f"Expected one {prefix} workbook, found {len(matches)}")
    return matches[0]


def _read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell.value for cell in row) for row in workbook.active.iter_rows()]
    finally:
        workbook.close()


def _extract_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        raise ValueError(f"No numeric value found in {value!r}")
    return float(match.group())


def _hourly_buy_prices(parameter_rows):
    peak_price = _extract_number(parameter_rows[1][1])
    flat_price = _extract_number(parameter_rows[2][1])
    valley_price = _extract_number(parameter_rows[3][1])
    prices = []
    for hour in range(HOURS):
        if 10 <= hour < 15 or 18 <= hour < 21:
            prices.append(peak_price)
        elif 7 <= hour < 10 or 15 <= hour < 18 or 21 <= hour < 23:
            prices.append(flat_price)
        else:
            prices.append(valley_price)
    return np.asarray(prices, dtype=float)


def solve_problem1(project_root):
    project_root = Path(project_root)
    load_rows = _read_rows(_find_attachment(project_root, "附件1"))
    renewable_rows = _read_rows(_find_attachment(project_root, "附件2"))
    equipment_rows = _read_rows(_find_attachment(project_root, "附件5"))
    storage_ammonia_rows = _read_rows(_find_attachment(project_root, "附件6"))
    buy_price_rows = _read_rows(_find_attachment(project_root, "附件7"))
    sell_price_rows = _read_rows(_find_attachment(project_root, "附件8"))

    time_labels = [str(row[0]) for row in load_rows[1:]]
    load_pu = np.asarray([row[1] for row in load_rows[1:]], dtype=float)
    renewable_times = [str(row[0]) for row in renewable_rows[1:]]
    wind_pu = np.asarray([row[1] for row in renewable_rows[1:]], dtype=float)
    pv_pu = np.asarray([row[2] for row in renewable_rows[1:]], dtype=float)
    if len(time_labels) != HOURS or time_labels != renewable_times:
        raise ValueError("Hourly input tables must contain the same 24 time periods")

    wind_cost = _extract_number(equipment_rows[1][1])
    pv_cost = _extract_number(equipment_rows[1][2])
    alk_om = _extract_number(equipment_rows[2][3])
    pem_om = _extract_number(equipment_rows[2][4])
    alk_efficiency = _extract_number(equipment_rows[4][3]) / 100.0
    pem_efficiency = _extract_number(equipment_rows[4][4]) / 100.0
    ammonia_investment = _extract_number(storage_ammonia_rows[1][2])
    ammonia_om = _extract_number(storage_ammonia_rows[2][2])
    ammonia_lifetime = _extract_number(storage_ammonia_rows[3][2])
    buy_prices = _hourly_buy_prices(buy_price_rows)
    sell_wind_price = _extract_number(sell_price_rows[1][1])
    sell_pv_price = _extract_number(sell_price_rows[2][1])
    if not np.isclose(sell_wind_price, sell_pv_price):
        raise ValueError("Problem 1 requires a common feed-in price for aggregate surplus")
    sell_prices = np.full(HOURS, sell_wind_price, dtype=float)

    conventional_load = CONVENTIONAL_LOAD_BASE_MW * load_pu
    wind_power = WIND_CAPACITY_MW * wind_pu
    pv_power = PV_CAPACITY_MW * pv_pu
    generation = wind_power + pv_power
    process_load = np.full(
        HOURS, ALK_POWER_MW + PEM_POWER_MW + AMMONIA_POWER_MW, dtype=float
    )
    total_load = conventional_load + process_load
    buy_power, sell_power = calculate_grid_exchange(total_load, generation)

    alk_hydrogen = float(calculate_hydrogen_output(ALK_POWER_MW, alk_efficiency))
    pem_hydrogen = float(calculate_hydrogen_output(PEM_POWER_MW, pem_efficiency))
    total_hydrogen = alk_hydrogen + pem_hydrogen
    ammonia_output_kg_per_h = total_hydrogen / 0.2
    ammonia_output_t = ammonia_output_kg_per_h * HOURS / 1000.0

    conventional_energy = float(np.sum(conventional_load) * TIME_STEP_H)
    process_energy = float(np.sum(process_load) * TIME_STEP_H)
    load_energy = float(np.sum(total_load) * TIME_STEP_H)
    wind_energy = float(np.sum(wind_power) * TIME_STEP_H)
    pv_energy = float(np.sum(pv_power) * TIME_STEP_H)
    generation_energy = wind_energy + pv_energy
    buy_energy = float(np.sum(buy_power) * TIME_STEP_H)
    sell_energy = float(np.sum(sell_power) * TIME_STEP_H)

    metrics = calculate_green_metrics(
        load_energy,
        generation_energy,
        buy_energy,
        sell_energy,
    )
    thresholds = {
        "新能源自发自用比例": 0.60,
        "新能源供电占比": 0.30,
        "上网电量比例": 0.20,
    }
    metric_status = {
        "新能源自发自用比例": metrics["新能源自发自用比例"]
        > thresholds["新能源自发自用比例"],
        "新能源供电占比": metrics["新能源供电占比"]
        > thresholds["新能源供电占比"],
        "上网电量比例": metrics["上网电量比例"]
        < thresholds["上网电量比例"],
    }
    metric_gaps = {
        "新能源自发自用比例": metrics["新能源自发自用比例"] - 0.60,
        "新能源供电占比": metrics["新能源供电占比"] - 0.30,
        "上网电量比例": 0.20 - metrics["上网电量比例"],
    }

    alk_energy = ALK_POWER_MW * HOURS * TIME_STEP_H
    pem_energy = PEM_POWER_MW * HOURS * TIME_STEP_H
    ammonia_energy = AMMONIA_POWER_MW * HOURS * TIME_STEP_H
    costs = calculate_daily_cost(
        wind_energy_mwh=wind_energy,
        pv_energy_mwh=pv_energy,
        buy_energy_mwh=buy_power * TIME_STEP_H,
        buy_price_yuan_per_kwh=buy_prices,
        sell_energy_mwh=sell_power * TIME_STEP_H,
        sell_price_yuan_per_kwh=sell_prices,
        alk_energy_mwh=alk_energy,
        pem_energy_mwh=pem_energy,
        ammonia_energy_mwh=ammonia_energy,
        rated_hydrogen_demand_kg_per_h=total_hydrogen,
        ammonia_output_t=ammonia_output_t,
        wind_cost_yuan_per_kwh=wind_cost,
        pv_cost_yuan_per_kwh=pv_cost,
        alk_om_yuan_per_kwh=alk_om,
        pem_om_yuan_per_kwh=pem_om,
        ammonia_om_yuan_per_kwh=ammonia_om,
        ammonia_investment_yuan_per_kg_h2=ammonia_investment,
        ammonia_lifetime_years=ammonia_lifetime,
    )

    power_residual = generation + buy_power - total_load - sell_power
    energy_residual = generation_energy + buy_energy - load_energy - sell_energy
    hydrogen_residual = total_hydrogen - 0.2 * ammonia_output_kg_per_h
    ammonia_power_residual = AMMONIA_POWER_MW - 0.5 * ammonia_output_kg_per_h / 1000.0
    checks = {
        "最大逐时功率平衡残差": float(np.max(np.abs(power_residual))),
        "电量平衡残差": float(energy_residual),
        "购售电同时发生最大值": float(np.max(buy_power * sell_power)),
        "氢气平衡残差": float(hydrogen_residual),
        "合成氨耗电残差": float(ammonia_power_residual),
    }

    summary = {
        "常规负荷电量": conventional_energy,
        "电氢氨负荷电量": process_energy,
        "总负荷电量": load_energy,
        "风电电量": wind_energy,
        "光伏电量": pv_energy,
        "新能源发电量": generation_energy,
        "购电量": buy_energy,
        "上网电量": sell_energy,
        "氢气产量": total_hydrogen * HOURS,
        "氨产量": ammonia_output_t,
        "缺电时段数": int(np.sum(buy_power > 1e-12)),
        "余电时段数": int(np.sum(sell_power > 1e-12)),
        "最大购电功率": float(np.max(buy_power)),
        "最大上网功率": float(np.max(sell_power)),
        "新能源总量充裕系数": generation_energy / load_energy,
    }

    hourly = []
    for index in range(HOURS):
        hourly.append(
            {
                "时段": time_labels[index],
                "常规负荷标幺值": float(load_pu[index]),
                "常规负荷功率": float(conventional_load[index]),
                "风电标幺值": float(wind_pu[index]),
                "光伏标幺值": float(pv_pu[index]),
                "风电功率": float(wind_power[index]),
                "光伏功率": float(pv_power[index]),
                "电氢氨负荷功率": float(process_load[index]),
                "总负荷功率": float(total_load[index]),
                "风光总出力": float(generation[index]),
                "购电功率": float(buy_power[index]),
                "上网功率": float(sell_power[index]),
                "购电电价": float(buy_prices[index]),
                "上网电价": float(sell_prices[index]),
            }
        )

    parameters = {
        "常规负荷基准功率": CONVENTIONAL_LOAD_BASE_MW,
        "风电装机容量": WIND_CAPACITY_MW,
        "光伏装机容量": PV_CAPACITY_MW,
        "碱性电解槽功率": ALK_POWER_MW,
        "PEM电解槽功率": PEM_POWER_MW,
        "合成氨装置功率": AMMONIA_POWER_MW,
        "碱性电解槽效率": alk_efficiency,
        "PEM电解槽效率": pem_efficiency,
        "碱性电解槽产氢速率": alk_hydrogen,
        "PEM电解槽产氢速率": pem_hydrogen,
        "合成氨产量速率": ammonia_output_kg_per_h,
        "风电度电成本": wind_cost,
        "光伏度电成本": pv_cost,
        "碱性电解槽运维系数": alk_om,
        "PEM电解槽运维系数": pem_om,
        "合成氨装置运维系数": ammonia_om,
        "合成氨装置投资成本": ammonia_investment,
        "合成氨装置寿命": ammonia_lifetime,
        "上网电价": sell_wind_price,
    }
    return {
        "hourly": hourly,
        "summary": summary,
        "metrics": metrics,
        "thresholds": thresholds,
        "metric_status": metric_status,
        "metric_gaps": metric_gaps,
        "costs": costs,
        "checks": checks,
        "parameters": parameters,
    }


def _write_rows(path, fieldnames, rows):
    with Path(path).open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_key_value(path, values, units):
    rows = [
        {"指标": key, "数值": value, "单位": units.get(key, "-")}
        for key, value in values.items()
    ]
    _write_rows(path, ["指标", "数值", "单位"], rows)


def _exchange_stair_series(buy, sell):
    buy = np.asarray(buy, dtype=float)
    sell = np.asarray(sell, dtype=float)
    if buy.shape != sell.shape:
        raise ValueError("Buy and sell series must have the same shape")
    if np.any(buy < -1e-9) or np.any(sell < -1e-9):
        raise ValueError("Buy and sell power must be nonnegative")
    if np.any(buy * sell > 1e-8):
        raise ValueError("Buy and sell power must be mutually exclusive")
    edges = np.arange(buy.size + 1, dtype=float)
    return edges, np.maximum(buy, 0.0), -np.maximum(sell, 0.0)


def _plot_power_balance(hourly, figure_dir):
    hours = np.arange(len(hourly), dtype=float) + 0.5
    wind = np.asarray([row["风电功率"] for row in hourly], dtype=float)
    pv = np.asarray([row["光伏功率"] for row in hourly], dtype=float)
    renewable = np.asarray([row["风光总出力"] for row in hourly], dtype=float)
    load = np.asarray([row["总负荷功率"] for row in hourly], dtype=float)
    buy = np.asarray([row["购电功率"] for row in hourly], dtype=float)
    sell = np.asarray([row["上网功率"] for row in hourly], dtype=float)
    edges, buy_step, sell_step = _exchange_stair_series(buy, sell)

    plt.rcParams["font.sans-serif"] = [
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False
    figure, axes = plt.subplots(
        2,
        1,
        figsize=(10.5, 6.8),
        sharex=True,
        gridspec_kw={"height_ratios": [2.0, 1.0], "hspace": 0.12},
    )
    upper, lower = axes
    upper.plot(hours, wind, color=PALETTE["blue"], linewidth=1.8, label="风电功率")
    upper.plot(hours, pv, color=PALETTE["yellow"], linewidth=1.8, label="光伏功率")
    upper.plot(
        hours,
        renewable,
        color=PALETTE["light_blue"],
        linewidth=2.4,
        label="风光总出力",
    )
    upper.plot(hours, load, color=PALETTE["red"], linewidth=2.4, label="总负荷功率")
    upper.set_ylabel("功率 / MW")
    upper.legend(ncol=4, loc="upper center", frameon=False)
    upper.grid(axis="y", color="#D9DEE8", linewidth=0.7, alpha=0.8)
    upper.spines[["top", "right"]].set_visible(False)

    lower.axhline(0.0, color="#7A7A7A", linewidth=0.8)
    lower.stairs(
        buy_step,
        edges,
        baseline=0.0,
        fill=True,
        color=PALETTE["orange"],
        alpha=0.85,
        linewidth=1.4,
        label="购电功率",
    )
    lower.stairs(
        sell_step,
        edges,
        baseline=0.0,
        fill=True,
        color=PALETTE["blue"],
        alpha=0.85,
        linewidth=1.4,
        label="上网功率",
    )
    lower.set_xlabel("时刻 / h")
    lower.set_ylabel("电网交换功率 / MW")
    lower.set_xlim(0.0, float(len(hourly)))
    lower.set_xticks(np.arange(0, len(hourly) + 1, 2))
    lower.legend(ncol=2, loc="lower left", frameon=False)
    lower.grid(axis="y", color="#D9DEE8", linewidth=0.7, alpha=0.8)
    lower.spines[["top", "right"]].set_visible(False)

    figure.align_ylabels(axes)
    figure.savefig(Path(figure_dir) / "问题一逐时功率平衡.pdf", bbox_inches="tight")
    figure.savefig(
        Path(figure_dir) / "问题一逐时功率平衡.png",
        dpi=300,
        bbox_inches="tight",
    )
    plt.close(figure)


def export_plain_results(result, output_dir, figure_dir):
    output_dir = Path(output_dir)
    figure_dir = Path(figure_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    figure_dir.mkdir(parents=True, exist_ok=True)

    with (output_dir / "问题一完整计算结果.json").open("w", encoding="utf-8") as file:
        json.dump(result, file, ensure_ascii=False, indent=2)

    _write_rows(
        output_dir / "问题一逐时运行结果.csv",
        list(result["hourly"][0].keys()),
        result["hourly"],
    )
    summary_units = {
        "常规负荷电量": "MWh",
        "电氢氨负荷电量": "MWh",
        "总负荷电量": "MWh",
        "风电电量": "MWh",
        "光伏电量": "MWh",
        "新能源发电量": "MWh",
        "购电量": "MWh",
        "上网电量": "MWh",
        "氢气产量": "kgH2/day",
        "氨产量": "t/day",
        "缺电时段数": "h",
        "余电时段数": "h",
        "最大购电功率": "MW",
        "最大上网功率": "MW",
    }
    _write_key_value(
        output_dir / "问题一汇总结果.csv",
        result["summary"],
        summary_units,
    )

    metric_rows = []
    for name, value in result["metrics"].items():
        metric_rows.append(
            {
                "指标": name,
                "计算值": value,
                "阈值": result["thresholds"][name],
                "判定方向": ">" if name != "上网电量比例" else "<",
                "阈值缺口": result["metric_gaps"][name],
                "结论": "达标" if result["metric_status"][name] else "不达标",
            }
        )
    _write_rows(
        output_dir / "问题一绿电指标.csv",
        ["指标", "计算值", "阈值", "判定方向", "阈值缺口", "结论"],
        metric_rows,
    )

    cost_units = {
        "风电发电成本": "¥/day",
        "光伏发电成本": "¥/day",
        "购电成本": "¥/day",
        "售电收入": "¥/day",
        "碱性电解槽运维成本": "¥/day",
        "PEM电解槽运维成本": "¥/day",
        "合成氨装置运维成本": "¥/day",
        "合成氨装置资本分摊": "¥/day",
        "日净成本": "¥/day",
        "净吨氨成本": "¥/t",
    }
    _write_key_value(
        output_dir / "问题一成本明细.csv",
        result["costs"],
        cost_units,
    )
    check_units = {
        "最大逐时功率平衡残差": "MW",
        "电量平衡残差": "MWh",
        "购售电同时发生最大值": "MW2",
        "氢气平衡残差": "kgH2/h",
        "合成氨耗电残差": "MW",
    }
    _write_key_value(
        output_dir / "问题一约束校验.csv",
        result["checks"],
        check_units,
    )
    _plot_power_balance(result["hourly"], figure_dir)


def run_problem1(project_root, output_dir=None, figure_dir=None):
    project_root = Path(project_root)
    output_dir = Path(output_dir) if output_dir else project_root / "outputs" / "问题一计算结果"
    figure_dir = Path(figure_dir) if figure_dir else project_root / "figures" / "问题一计算结果"
    result = solve_problem1(project_root)
    export_plain_results(result, output_dir, figure_dir)
    return result


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    final_result = run_problem1(root)
    console_result = {
        "汇总结果": final_result["summary"],
        "绿电指标": final_result["metrics"],
        "达标结论": final_result["metric_status"],
        "成本结果": final_result["costs"],
        "约束校验": final_result["checks"],
    }
    print(json.dumps(console_result, ensure_ascii=False, indent=2))
